import sys
import os

views_path = r"c:\Sistema de Facturacion\administrar\views.py"

content_to_insert = """

# ==========================
# REPORTES CTA CTE PROVEEDOR
# ==========================

@login_required
@verificar_permiso('ctacte')
def cc_proveedor_imprimir(request, id):
    \"\"\"Vista para imprimir estado de cuenta del proveedor\"\"\"
    from datetime import datetime
    from .models import Proveedor, Compra, MovimientoCuentaCorrienteProveedor, Empresa
    from decimal import Decimal
    from django.template.loader import render_to_string
    from django.http import HttpResponse, JsonResponse
    from xhtml2pdf import pisa
    from io import BytesIO

    try:
        proveedor = Proveedor.objects.get(id=id)
    except Proveedor.DoesNotExist:
        return HttpResponse("Proveedor no encontrado", status=404)
    
    movimientos = []
    try:
        compras = Compra.objects.filter(proveedor=proveedor, estado='REGISTRADA')
        for c in compras:
            movimientos.append({
                'fecha': c.fecha,
                'tipo': 'COMPRA',
                'descripcion': f'Compra #{c.nro_comprobante or c.id}',
                'debe': 0,
                'haber': float(c.total),
            })
    except Exception as e:
        pass

    try:
        pagos = MovimientoCuentaCorrienteProveedor.objects.filter(proveedor=proveedor)
        for p in pagos:
             debe = float(p.monto) if p.tipo == 'DEBE' else 0
             haber = float(p.monto) if p.tipo == 'HABER' else 0
             movimientos.append({
                'fecha': p.fecha,
                'tipo': 'PAGO',
                'descripcion': p.descripcion,
                'debe': debe,
                'haber': haber,
            })
    except Exception as e:
        pass

    movimientos.sort(key=lambda x: x['fecha'])
    saldo = Decimal('0')
    for mov in movimientos:
        saldo += Decimal(str(mov['haber'])) - Decimal(str(mov['debe']))
        mov['saldo'] = float(saldo)

    empresa = Empresa.objects.first()

    context = {
        'proveedor': proveedor,
        'movimientos': movimientos,
        'saldo_actual': float(saldo),
        'empresa': empresa,
        'fecha_impresion': datetime.now().strftime('%d/%m/%Y %H:%M')
    }

    try:
        html_string = render_to_string('administrar/ctacte/proveedor_imprimir_final.html', context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            return response
        return HttpResponse("Error generando PDF", status=500)
    except Exception as e:
        # Fallback to plain text if template doesn't exist
        return HttpResponse(f"Error o plantilla inexistente: {e}", status=500)

@login_required
@verificar_permiso('ctacte')
def api_cc_proveedor_exportar_pdf(request, id):
    \"\"\"Exportar a PDF (exactamente la misma lógica)\"\"\"
    response = cc_proveedor_imprimir(request, id)
    if isinstance(response, HttpResponse):
        response['Content-Disposition'] = f'attachment; filename="estado_cuenta_prov_{id}.pdf"'
    return response

@login_required
@verificar_permiso('ctacte')
def api_cc_proveedor_exportar_excel(request, id):
    \"\"\"Exportar estado de cuenta proveedor a Excel\"\"\"
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse, JsonResponse
        from datetime import datetime
        from .models import Proveedor, Compra, MovimientoCuentaCorrienteProveedor
        from decimal import Decimal
        
        proveedor = Proveedor.objects.get(id=id)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estado de Cuenta"
        
        ws['A1'] = 'ESTADO DE CUENTA CORRIENTE PROVEEDORES'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Proveedor: {proveedor.nombre}'
        ws['A3'] = f'CUIT: {proveedor.cuit or "N/A"}'
        ws['A4'] = f'Fecha: {datetime.now().strftime("%d/%m/%Y")}'
        
        headers = ['Fecha', 'Tipo', 'Descripción', 'Debe', 'Haber', 'Saldo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        movimientos = []
        try:
            compras = Compra.objects.filter(proveedor=proveedor, estado='REGISTRADA')
            for c in compras:
                movimientos.append({
                    'fecha': c.fecha,
                    'tipo': 'COMPRA',
                    'descripcion': f'Compra #{c.nro_comprobante or c.id}',
                    'debe': 0,
                    'haber': float(c.total),
                })
        except Exception: pass
        try:
            pagos = MovimientoCuentaCorrienteProveedor.objects.filter(proveedor=proveedor)
            for p in pagos:
                debe = float(p.monto) if p.tipo == 'DEBE' else 0
                haber = float(p.monto) if p.tipo == 'HABER' else 0
                movimientos.append({
                    'fecha': p.fecha,
                    'tipo': 'PAGO',
                    'descripcion': p.descripcion,
                    'debe': debe,
                    'haber': haber,
                })
        except Exception: pass

        movimientos.sort(key=lambda x: x['fecha'])
        saldo = Decimal('0')
        row = 7
        for mov in movimientos:
            saldo += Decimal(str(mov['haber'])) - Decimal(str(mov['debe']))
            ws.cell(row=row, column=1, value=mov['fecha'].strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=mov['tipo'])
            ws.cell(row=row, column=3, value=mov['descripcion'])
            ws.cell(row=row, column=4, value=mov['debe']).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=5, value=mov['haber']).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=6, value=float(saldo)).number_format = '"$"#,##0.00'
            row += 1
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="estado_cuenta_prov_{proveedor.id}.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        return JsonResponse({'ok': False, 'error': 'openpyxl no está instalado'}, status=500)
    except Proveedor.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Proveedor no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

"""

with open(views_path, "a", encoding="utf-8") as f:
    f.write(content_to_insert)

print("Funciones insertadas correctamente en views.py")
