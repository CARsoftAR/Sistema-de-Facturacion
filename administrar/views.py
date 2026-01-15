from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.contrib import messages
from django.db import transaction
from django.core.management import call_command
from django.core.paginator import Paginator
import io
import datetime
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from decimal import Decimal
import json
import re
from django.db import IntegrityError
from django.views.decorators.http import require_POST, require_http_methods
from django.utils.crypto import get_random_string
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required

from .models import (
    Producto, Proveedor, Marca, Rubro,
    Provincia, Localidad, Cliente,
    OrdenCompra, DetalleOrdenCompra,
    Compra, DetalleCompra, MovimientoStock,
    MovimientoCaja, CajaDiaria, Venta, DetalleVenta, Unidad,
    MovimientoCuentaCorriente, MovimientoCuentaCorrienteProveedor, Pedido, DetallePedido,
    InvoiceTemplate, Empresa, PerfilUsuario,
    Remito, DetalleRemito, NotaCredito, DetalleNotaCredito, NotaDebito, DetalleNotaDebito,
    Presupuesto, DetallePresupuesto, ActiveSession
)
from .forms import InvoiceTemplateForm
from functools import wraps
from django.core.exceptions import PermissionDenied

def verificar_permiso(permiso):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            
            if request.user.is_staff:
                return view_func(request, *args, **kwargs)
                
            try:
                perfil = request.user.perfilusuario
                # Mapeo de permisos a campos del modelo
                campos_permisos = {
                    'ventas': perfil.acceso_ventas,
                    'compras': perfil.acceso_compras,
                    'productos': perfil.acceso_productos,
                    'clientes': perfil.acceso_clientes,
                    'proveedores': perfil.acceso_proveedores,
                    'caja': perfil.acceso_caja,
                    'contabilidad': perfil.acceso_contabilidad,
                    'configuracion': perfil.acceso_configuracion,
                    'usuarios': perfil.acceso_usuarios,
                    'reportes': perfil.acceso_reportes,
                    'pedidos': perfil.acceso_pedidos,
                    'bancos': perfil.acceso_bancos,
                    'ctacte': perfil.acceso_ctacte,
                    'remitos': perfil.acceso_remitos,
                }
                
                if campos_permisos.get(permiso, False):
                    return view_func(request, *args, **kwargs)
            except PerfilUsuario.DoesNotExist:
                pass
                
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.path.startswith('/api/'):
                return JsonResponse({'ok': False, 'error': 'No tienes permisos para acceder a esta secci贸n.'}, status=403)
                
            messages.error(request, 'No tienes permisos para acceder a esta secci贸n.')
            return redirect('menu')
            
        return _wrapped_view
    return decorator

#====================================
# AUTENTICACIN
#====================================
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        force_logout = request.POST.get('force_logout') == 'true'
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Verificar si ya existe una sesi贸n activa para este usuario
            # (Excluyendo la sesi贸n actual si es que ya tiene una cookie, aunque en login usualmente no)
            existing_sessions = ActiveSession.objects.filter(user=user)
            
            if existing_sessions.exists() and not force_logout:
                # CONFLICTO: Ya hay sesi贸n activa
                current_session = existing_sessions.first()
                
                # Preparamos datos para mostrar al usuario
                conflict_data = {
                    'ip': current_session.ip_address,
                    'last_activity': current_session.last_activity,
                    'user_agent': current_session.user_agent
                }
                
                return render(request, 'administrar/login.html', {
                    'conflict': True,
                    'conflict_data': conflict_data,
                    'username_value': username, # Para re-llenar el form oculto o visible
                    # No pasamos password por seguridad, el usuario deber谩 reingresarla si fuera un campo visible,
                    # pero en este flujo "Force Logout" asumimos que la validaci贸n ya pas贸 arriba.
                    # Sin embargo, para "Force Logout" necesitaremos enviar las credenciales de nuevo o un token.
                    # Simplificaci贸n: En el HTML pediremos confirmar "Cerrar sesi贸n" y re-enviaremos los datos.
                })
            
            # Si llegamos aqu铆 es porque:
            # 1. No hay sesiones activas OR
            # 2. force_logout es True (el usuario decidi贸 patear al otro)
            
            if force_logout:
                # Eliminar todas las sesiones anteriores
                existing_sessions.delete()
                
            # Login normal
            login(request, user)
            
            # Crear/Registrar la nueva sesi贸n activa lo maneja el user_logged_in signal
            # que ya tenemos en middleware.py (log_user_login)
            
            next_url = request.POST.get('next') or request.GET.get('next') or '/'
            return redirect(next_url)
        else:
            # messages.error(request, 'Usuario o contrase帽a incorrectos') # Eliminado para evitar duplicidad
            return render(request, 'administrar/login.html', {
                'form': {'errors': True}, 
                'username_value': username
            })
    
    return render(request, 'administrar/login.html')

def register_view(request):
    from django.contrib.auth.forms import UserCreationForm
    from .models import PerfilUsuario
    
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            # --- SECURITY ENHANCEMENT: Lock down new user ---
            # Create a profile with ALL permissions set to False by default
            # This ensures they can log in but can't access any sensitive module
            PerfilUsuario.objects.create(
                user=user,
                acceso_ventas=False,
                acceso_compras=False,
                acceso_productos=False,
                acceso_clientes=False,
                acceso_proveedores=False,
                acceso_caja=False,
                acceso_contabilidad=False,
                acceso_configuracion=False,
                acceso_usuarios=False,
                acceso_reportes=False
            )
            
            login(request, user)
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    
    return render(request, 'administrar/register.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.success(request, 'Has cerrado sesi贸n correctamente')
    return redirect('login')

#====================================
# VALIDAR CUIT
#====================================
def validar_cuit(cuit):
    c = re.sub(r"\D", "", cuit)
    if len(c) != 11:
        return False

    coef = [5,4,3,2,7,6,5,4,3,2]
    suma = sum(int(c[i]) * coef[i] for i in range(10))

    dv = 11 - (suma % 11)
    if dv == 11: dv = 0
    if dv == 10: dv = 9

    return dv == int(c[10])

# =======================================
#  DEMOS JSON (no cr铆tico, pero los dejo)
# =======================================
def index(request):
    my_dictionary = {"a": 1, "b": 2}
    return JsonResponse(my_dictionary)

def index2(request):
    my_array = [("a", 1), ("b", 2)]
    return JsonResponse(my_array, safe=False)


# =======================================
#  VISTAS PRINCIPALES DEL SISTEMA
# =======================================

@login_required
def menu(request):
    return render(request, "administrar/menu.html")

@login_required
@verificar_permiso('reportes')
def dashboard(request):
    from django.db.models import Sum, Count, F
    from datetime import date, timedelta, datetime
    
    hoy = date.today()
    
    # 1. KPIs Principales
    ventas_hoy = Venta.objects.filter(fecha__date=hoy).aggregate(total=Sum('total'))['total'] or 0
    caja_hoy = MovimientoCaja.objects.filter(fecha__date=hoy).aggregate(total=Sum('monto'))['total'] or 0
    pedidos_pendientes = Pedido.objects.filter(estado__in=['PENDIENTE', 'PREPARACION']).count()
    pedidos_listos = Pedido.objects.filter(estado='LISTO').count()
    stock_bajo_count = Producto.objects.filter(Q(stock__lte=F('stock_minimo')) | Q(stock__lte=10)).count() # Usando stock_minimo din谩mico O hardcode 10
    
    # 2. Gr谩fico: Ventas 煤ltimos 7 d铆as
    fecha_inicio_chart = hoy - timedelta(days=6)
    ventas_chart = Venta.objects.filter(fecha__date__gte=fecha_inicio_chart)\
        .extra(select={'day': 'date(fecha)'})\
        .values('day')\
        .annotate(total=Sum('total'))\
        .order_by('day')
        
    # Formatear para Chart.js
    chart_labels = []
    chart_data = []
    for i in range(7):
        dia_iter = fecha_inicio_chart + timedelta(days=i)
        dia_str = dia_iter.strftime('%Y-%m-%d')
        # Buscar valor
        found = next((v for v in ventas_chart if str(v['day']) == dia_str), None)
        chart_labels.append(dia_iter.strftime('%d/%m'))
        chart_data.append(float(found['total']) if found else 0)

    # 3. Actividad Reciente (ltimas 5 ventas o movimientos)
    recientes = []
    ultimas_ventas = Venta.objects.select_related('cliente').order_by('-fecha')[:5]
    for v in ultimas_ventas:
        recientes.append({
            'tipo': 'VENTA',
            'icono': 'bi-cart-check',
            'color': 'primary',
            'fecha': v.fecha,
            'texto': f"Venta #{v.id} - ${v.total}",
            'subtexto': v.cliente.nombre if v.cliente else 'Cliente Final'
        })
    
    # Podr铆amos mezclar con compras, pero por ahora solo ventas recientes para simplificar
    
    # 4. Top Productos (ltimos 30 d铆as)
    fecha_inicio_top = hoy - timedelta(days=30)
    top_productos = DetalleVenta.objects.filter(venta__fecha__date__gte=fecha_inicio_top)\
        .values('producto__descripcion')\
        .annotate(total_vendido=Sum('cantidad'))\
        .order_by('-total_vendido')[:5]

    context = {
        'ventas_hoy': ventas_hoy,
        'caja_hoy': caja_hoy,
        'pedidos_pendientes': pedidos_pendientes,
        'pedidos_listos': pedidos_listos,
        'stock_bajo': stock_bajo_count,
        # Chart Data
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        # Lists
        'actividad_reciente': recientes,
        'top_productos': top_productos,
    }
    
    return render(request, "administrar/dashboard.html", context)

@login_required
def estado_sistema(request):
    import psutil
    import platform
    import django
    import socket
    from django.db import connection
    from django.conf import settings
    from datetime import datetime
    
    # 1. Uso de Base de Datos (MySQL)
    db_size = 0
    db_name = settings.DATABASES['default']['NAME']
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT SUM(data_length + index_length) / 1024 / 1024 "
                "FROM information_schema.TABLES WHERE table_schema = %s", 
                [db_name]
            )
            row = cursor.fetchone()
            if row and row[0]:
                db_size = round(row[0], 2)
    except Exception as e:
        print(f"Error obteniendo tama帽o DB: {e}")
        db_size = "N/A"

    # 2. Conexi贸n AFIP (Simulada por ahora, o verificar si hay token v谩lido)
    # Como no hay m贸dulos implementados, mostramos el estado real
    afip_status = "No configurado" 
    afip_class = "text-muted"
    
    # 3. Uso del Servidor
    cpu_usage = psutil.cpu_percent(interval=None)
    ram_usage = psutil.virtual_memory().percent
    
    # 4. Informaci贸n del Sistema
    try:
        socket.create_connection(("www.google.com", 80), timeout=2)
        network_status = "Operativa"
        network_class = "bg-success"
    except OSError:
        network_status = "Sin conexi贸n"
        network_class = "bg-danger"

    context = {
        'db_size': db_size,
        'db_name': db_name,
        'cpu_usage': cpu_usage,
        'ram_usage': ram_usage,
        'python_version': platform.python_version(),
        'django_version': django.get_version(),
        'db_engine': settings.DATABASES['default']['ENGINE'].split('.')[-1],
        'network_status': network_status,
        'network_class': network_class,
        'afip_status': afip_status,
        'afip_class': afip_class,
        'server_software': request.META.get('SERVER_SOFTWARE', 'Desconocido'),
        'last_backup': "No disponible" # Implementar l贸gica real si existe registro de backups
    }
    
    return render(request, "administrar/estado.html", context)

@login_required
@verificar_permiso('configuracion')
def parametros(request):
    from .models import Empresa
    from django.contrib import messages
    
    # Obtener o crear la configuraci贸n de la empresa
    empresa, created = Empresa.objects.get_or_create(
        id=1,
        defaults={
            'nombre': 'Mi Empresa',
            'cuit': '00-00000000-0',
            'direccion': '',
            'condicion_fiscal': 'RI',
            'punto_venta': '0001',
        }
    )
    
    if request.method == 'POST':
        # Guardar los datos del formulario
        empresa.nombre = request.POST.get('nombre', '')
        empresa.cuit = request.POST.get('cuit', '')
        empresa.direccion = request.POST.get('direccion', '')
        empresa.iibb = request.POST.get('iibb', '')
        empresa.inicio_actividades = request.POST.get('inicio_actividades') or None
        empresa.telefono = request.POST.get('telefono', '')
        empresa.email = request.POST.get('email', '')
        empresa.condicion_fiscal = request.POST.get('condicion_fiscal', 'RI')
        empresa.punto_venta = request.POST.get('punto_venta', '0001')
        empresa.tipo_facturacion = request.POST.get('tipo_facturacion', 'AFIP')
        
        # Nuevos Campos
        empresa.localidad = request.POST.get('localidad', '')
        empresa.provincia = request.POST.get('provincia', '')
        empresa.habilita_remitos = request.POST.get('habilita_remitos') == 'on'
        empresa.moneda_predeterminada = request.POST.get('moneda', 'ARS')
        empresa.pie_factura = request.POST.get('pie_factura', '')
        
        # Manejar logo si se subi贸
        if 'logo' in request.FILES:
            empresa.logo = request.FILES['logo']
        
        empresa.save()
        messages.success(request, 'Par谩metros guardados correctamente.')
        
    context = {
        'empresa': empresa
    }
    return render(request, "administrar/parametros.html", context)


# =======================================
#  MDULOS COMERCIALES
# =======================================

@login_required
@verificar_permiso('ventas')
def ventas_lista(request):
    return render(request, "administrar/ventas.html")

@login_required
@verificar_permiso('ventas')
def venta_nueva(request):
    return render(request, "administrar/venta_nueva.html")

@login_required
@verificar_permiso('ventas')
def detalle_venta(request, venta_id):
    venta = get_object_or_404(Venta.objects.select_related('pedido_origen', 'cliente'), pk=venta_id)
    return render(request, 'administrar/ventas/detalle_venta.html', {'venta': venta})


@login_required
@verificar_permiso('compras')
def compras_lista(request):
    return render(request, "administrar/compras.html")


@login_required
@verificar_permiso('clientes')
def clientes_lista(request):
    clientes = Cliente.objects.all().order_by("nombre")
    return render(request, "administrar/clientes.html", {
        "clientes": clientes
    })


@login_required
@verificar_permiso('proveedores')
def proveedores_lista(request):
    proveedores = Proveedor.objects.select_related("provincia", "localidad").all().order_by("nombre")
    provincias = Provincia.objects.all().order_by("nombre")
    localidades = Localidad.objects.all().order_by("nombre")

    return render(request, "administrar/proveedores.html", {
        "proveedores": proveedores,
        "provincias": provincias,
        "localidades": localidades,
    })



@login_required
@verificar_permiso('contabilidad')
def contabilidad(request):
    return render(request, "administrar/contabilidad/menu.html")


# =======================================
#  API CAJA - MOVIMIENTOS
# =======================================

@login_required
@verificar_permiso('caja')
def api_caja_movimientos_lista(request):
    """API para listar movimientos de caja con paginaci贸n y filtros"""
    from django.db.models import Sum, Q
    from datetime import datetime
    
    try:
        # Paginaci贸n
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        # Filtros
        fecha_desde = request.GET.get('fecha_desde', '').strip()
        fecha_hasta = request.GET.get('fecha_hasta', '').strip()
        tipo = request.GET.get('tipo', '').strip()
        busqueda = request.GET.get('busqueda', '').strip()
        
        # Query base
        movimientos = MovimientoCaja.objects.select_related('usuario').all()
        
        # Filtros de fecha
        if fecha_desde:
            try:
                movimientos = movimientos.filter(fecha__date__gte=datetime.strptime(fecha_desde, '%Y-%m-%d').date())
            except ValueError: pass
        if fecha_hasta:
            try:
                movimientos = movimientos.filter(fecha__date__lte=datetime.strptime(fecha_hasta, '%Y-%m-%d').date())
            except ValueError: pass
        
        if tipo:
            movimientos = movimientos.filter(tipo=tipo)
        
        if busqueda:
            movimientos = movimientos.filter(
                Q(descripcion__icontains=busqueda) |
                Q(usuario__username__icontains=busqueda)
            )
        
        total = movimientos.count()
        movimientos_page = movimientos.order_by('-fecha')[(page - 1) * per_page : page * per_page]
        
        data = []
        for m in movimientos_page:
            data.append({
                'id': m.id,
                'fecha': m.fecha.strftime('%d/%m/%Y %H:%M'),
                'tipo': m.tipo,
                'descripcion': m.descripcion,
                'monto': float(m.monto),
                'usuario': m.usuario.username if m.usuario else 'Sistema',
            })
        
        # Saldo Global (Cash on Hand)
        global_ingresos = MovimientoCaja.objects.filter(tipo='Ingreso').aggregate(total=Sum('monto'))['total'] or 0
        global_egresos = MovimientoCaja.objects.filter(tipo='Egreso').aggregate(total=Sum('monto'))['total'] or 0
        saldo_global = float(global_ingresos) - float(global_egresos)

        # Saldo Filtrado (Current View)
        filt_ingresos = movimientos.filter(tipo='Ingreso').aggregate(total=Sum('monto'))['total'] or 0
        filt_egresos = movimientos.filter(tipo='Egreso').aggregate(total=Sum('monto'))['total'] or 0
        saldo_filtrado = float(filt_ingresos) - float(filt_egresos)
        
        return JsonResponse({
            'movimientos': data,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page,
            'saldo_actual': saldo_global, # Keep original key for compatibility
            'saldo_filtrado': saldo_filtrado, # New key
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
@login_required
@verificar_permiso('caja')
def api_caja_movimiento_crear(request):
    """API para crear un nuevo movimiento de caja"""
    # Verificar si hay una caja abierta
    caja = CajaDiaria.objects.filter(estado='ABIERTA').first()
    if not caja:
        return JsonResponse({'error': 'Debe abrir la caja antes de registrar movimientos'}, status=400)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inv谩lido'}, status=400)
    
    tipo = data.get('tipo', '').strip()
    descripcion = data.get('descripcion', '').strip()
    monto = data.get('monto', 0)
    
    # Validaciones
    if not tipo or tipo not in ['Ingreso', 'Egreso']:
        return JsonResponse({'error': 'Tipo inv谩lido. Debe ser Ingreso o Egreso'}, status=400)
    
    if not descripcion:
        return JsonResponse({'error': 'La descripci贸n es requerida'}, status=400)
    
    try:
        monto = Decimal(str(monto))
        if monto <= 0:
            return JsonResponse({'error': 'El monto debe ser mayor a 0'}, status=400)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Monto inv谩lido'}, status=400)
    
    try:
        movimiento = MovimientoCaja.objects.create(
            caja_diaria=caja,
            usuario=request.user,
            tipo=tipo,
            descripcion=descripcion,
            monto=monto,
        )

        try:
            from .services import AccountingService
            AccountingService.registrar_movimiento_caja(movimiento)
        except Exception as e:
            print(f"Error generando asiento caja: {e}")
        
        return JsonResponse({
            'ok': True,
            'id': movimiento.id,
            'message': 'Movimiento creado exitosamente'
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
@csrf_exempt
@require_POST
@login_required
@verificar_permiso('caja')
def api_caja_movimiento_editar(request, id):
    """API para editar un movimiento de caja existente"""
    try:
        movimiento = MovimientoCaja.objects.get(id=id)
    except MovimientoCaja.DoesNotExist:
        return JsonResponse({'error': 'Movimiento no encontrado'}, status=404)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inv谩lido'}, status=400)
    
    tipo = data.get('tipo', '').strip()
    descripcion = data.get('descripcion', '').strip()
    monto = data.get('monto', 0)
    
    # Validaciones
    if not tipo or tipo not in ['Ingreso', 'Egreso']:
        return JsonResponse({'error': 'Tipo inv谩lido'}, status=400)
    
    if not descripcion:
        return JsonResponse({'error': 'La descripci贸n es requerida'}, status=400)
    
    try:
        monto = Decimal(str(monto))
        if monto <= 0:
            return JsonResponse({'error': 'El monto debe ser mayor a 0'}, status=400)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Monto inv谩lido'}, status=400)
    
    try:
        movimiento.tipo = tipo
        movimiento.descripcion = descripcion
        movimiento.monto = monto
        movimiento.save()
        
        return JsonResponse({
            'ok': True,
            'message': 'Movimiento actualizado exitosamente'
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
@csrf_exempt
@require_POST
@login_required
@verificar_permiso('caja')
def api_caja_movimiento_eliminar(request, id):
    """API para eliminar un movimiento de caja"""
    try:
        movimiento = MovimientoCaja.objects.get(id=id)
    except MovimientoCaja.DoesNotExist:
        return JsonResponse({'error': 'Movimiento no encontrado'}, status=404)
    
    try:
        movimiento.delete()
        return JsonResponse({
            'ok': True,
            'message': 'Movimiento eliminado exitosamente'
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@verificar_permiso('caja')
def api_caja_movimiento_detalle(request, id):
    """API para obtener detalle de un movimiento de caja"""
    try:
        movimiento = MovimientoCaja.objects.get(id=id)
    except MovimientoCaja.DoesNotExist:
        return JsonResponse({'error': 'Movimiento no encontrado'}, status=404)
    
    return JsonResponse({
        'id': movimiento.id,
        'tipo': movimiento.tipo,
        'descripcion': movimiento.descripcion,
        'monto': float(movimiento.monto),
        'fecha': movimiento.fecha.strftime('%d/%m/%Y %H:%M'),
    })


@login_required
@verificar_permiso('caja')
def api_caja_saldo_actual(request):
    """API para obtener el saldo actual de caja"""
    from django.db.models import Sum
    
    ingresos = MovimientoCaja.objects.filter(tipo='Ingreso').aggregate(total=Sum('monto'))['total'] or 0
    egresos = MovimientoCaja.objects.filter(tipo='Egreso').aggregate(total=Sum('monto'))['total'] or 0
    saldo = float(ingresos) - float(egresos)
    
    return JsonResponse({
        'saldo': saldo,
        'ingresos': float(ingresos),
        'egresos': float(egresos),
    })


@csrf_exempt
@require_POST
@login_required
@verificar_permiso('caja')
def api_caja_cierre(request):
    """API para realizar cierre de caja"""
    from django.db.models import Sum
    from django.utils import timezone
    from decimal import Decimal
    from django.db import transaction
    
    caja = CajaDiaria.objects.filter(estado='ABIERTA').first()
    if not caja:
        return JsonResponse({'error': 'No hay ninguna caja abierta para cerrar'}, status=400)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
        monto_real = Decimal(str(data.get('monto_real', 0)))
        
        # Calcular el saldo de SISTEMA desde que se abri贸 la caja
        ingresos = MovimientoCaja.objects.filter(caja_diaria=caja, tipo='Ingreso').aggregate(total=Sum('monto'))['total'] or 0
        egresos = MovimientoCaja.objects.filter(caja_diaria=caja, tipo='Egreso').aggregate(total=Sum('monto'))['total'] or 0
        saldo_sistema = Decimal(str(ingresos)) - Decimal(str(egresos))
        
        diferencia = monto_real - saldo_sistema
        
        with transaction.atomic():
            caja.fecha_cierre = timezone.now()
            caja.monto_cierre_sistema = saldo_sistema
            caja.monto_cierre_real = monto_real
            caja.estado = 'CERRADA'
            caja.observaciones = data.get('observaciones', '')
            caja.save()
            
            # Registrar movimiento de ajuste si hay diferencia (Sobrante/Faltante)
            if abs(diferencia) > 0.01:
                tipo_ajuste = 'Ingreso' if diferencia > 0 else 'Egreso'
                desc_ajuste = f"Ajuste Arqueo Caja #{caja.id} ({'Sobrante' if diferencia > 0 else 'Faltante'})"
                
                mov_ajuste = MovimientoCaja.objects.create(
                    caja_diaria=caja,
                    usuario=request.user,
                    tipo=tipo_ajuste,
                    descripcion=desc_ajuste,
                    monto=abs(diferencia)
                )

                # Generar asiento contable de Arqueo
                try:
                    from .services import AccountingService
                    AccountingService.registrar_arqueo_caja(mov_ajuste, diferencia)
                except Exception as e:
                    print(f"Error generando asiento arqueo: {e}")

        return JsonResponse({
            'ok': True,
            'id': caja.id,
            'saldo_sistema': float(saldo_sistema),
            'monto_real': float(monto_real),
            'diferencia': float(diferencia),
            'ingresos_dia': float(ingresos),
            'egresos_dia': float(egresos),
            'saldo_total': float(monto_real)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@verificar_permiso('caja')
def api_caja_estado(request):
    """API para verificar el estado de la caja actual (si hay una abierta)"""
    caja_abierta = CajaDiaria.objects.filter(estado='ABIERTA').first()
    if caja_abierta:
        return JsonResponse({
            'abierta': True,
            'id': caja_abierta.id,
            'fecha_apertura': caja_abierta.fecha_apertura.strftime('%d/%m/%Y %H:%M'),
            'monto_apertura': float(caja_abierta.monto_apertura),
            'usuario': caja_abierta.usuario.username
        })
    return JsonResponse({'abierta': False})


@csrf_exempt
@require_POST
@login_required
@verificar_permiso('caja')
def api_caja_apertura(request):
    """API para abrir una nueva sesi贸n de caja"""
    if CajaDiaria.objects.filter(estado='ABIERTA').exists():
        return JsonResponse({'error': 'Ya existe una caja abierta'}, status=400)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
        # Fix: Handle None explicitly if forwarded as null from frontend
        val_monto = data.get('monto')
        if val_monto is None:
            val_monto = 0
            
        monto = Decimal(str(val_monto))
        
        if monto < 0:
            return JsonResponse({'error': 'El monto no puede ser negativo'}, status=400)
            
        with transaction.atomic():
            caja = CajaDiaria.objects.create(
                usuario=request.user,
                monto_apertura=monto,
                estado='ABIERTA'
            )
            # Tambi茅n creamos un movimiento de apertura
            MovimientoCaja.objects.create(
                caja_diaria=caja,
                usuario=request.user,
                tipo='Ingreso',
                descripcion=f'Apertura de Caja #{caja.id}',
                monto=monto
            )
        
        return JsonResponse({'ok': True, 'id': caja.id})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# =======================================
#  API PLAN DE CUENTAS
# =======================================

@login_required
@verificar_permiso('contabilidad')
def api_plan_cuentas_lista(request):
    """API para listar el plan de cuentas en formato jer谩rquico"""
    from administrar.models import PlanCuenta
    
    def construir_arbol(cuentas, padre_id=None):
        """Construye el 谩rbol jer谩rquico de cuentas"""
        resultado = []
        cuentas_nivel = cuentas.filter(padre_id=padre_id)
        
        for cuenta in cuentas_nivel:
            cuenta_dict = {
                'id': cuenta.id,
                'codigo': cuenta.codigo,
                'nombre': cuenta.nombre,
                'tipo': cuenta.tipo,
                'imputable': cuenta.imputable,
                'nivel': cuenta.nivel,
                'padre_id': cuenta.padre_id,
                'hijos': construir_arbol(cuentas, cuenta.id)
            }
            resultado.append(cuenta_dict)
        
        return resultado
    
    try:
        # Obtener todas las cuentas ordenadas por c贸digo
        cuentas = PlanCuenta.objects.all().order_by('codigo')
        
        # Construir 谩rbol jer谩rquico
        arbol = construir_arbol(cuentas)
        
        return JsonResponse({
            'success': True,
            'cuentas': arbol
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_POST
@csrf_exempt
@require_POST
@login_required
@verificar_permiso('contabilidad')
def api_plan_cuentas_crear(request):
    """API para crear una nueva cuenta en el plan de cuentas"""
    from administrar.models import PlanCuenta
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inv谩lido'}, status=400)
    
    # Validaciones
    codigo = data.get('codigo', '').strip()
    nombre = data.get('nombre', '').strip()
    tipo = data.get('tipo', '')
    imputable = data.get('imputable', True)
    nivel = int(data.get('nivel', 1))
    padre_id = data.get('padre_id')
    
    if not codigo:
        return JsonResponse({'error': 'El c贸digo es requerido'}, status=400)
    
    if not nombre:
        return JsonResponse({'error': 'El nombre es requerido'}, status=400)
    
    if tipo not in ['ACTIVO', 'PASIVO', 'PN', 'R_POS', 'R_NEG']:
        return JsonResponse({'error': 'Tipo de cuenta inv谩lido'}, status=400)
    
    # Verificar que el c贸digo sea 煤nico
    if PlanCuenta.objects.filter(codigo=codigo).exists():
        return JsonResponse({'error': 'Ya existe una cuenta con ese c贸digo'}, status=400)
    
    # Verificar que el padre exista si se especifica
    padre = None
    if padre_id:
        try:
            padre = PlanCuenta.objects.get(id=padre_id)
        except PlanCuenta.DoesNotExist:
            return JsonResponse({'error': 'La cuenta padre no existe'}, status=400)
    
    try:
        cuenta = PlanCuenta.objects.create(
            codigo=codigo,
            nombre=nombre,
            tipo=tipo,
            imputable=imputable,
            nivel=nivel,
            padre=padre
        )
        
        return JsonResponse({
            'ok': True,
            'cuenta': {
                'id': cuenta.id,
                'codigo': cuenta.codigo,
                'nombre': cuenta.nombre,
                'tipo': cuenta.tipo,
                'imputable': cuenta.imputable,
                'nivel': cuenta.nivel,
                'padre_id': cuenta.padre_id
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
@csrf_exempt
@require_POST
@login_required
@verificar_permiso('contabilidad')
def api_plan_cuentas_editar(request, id):
    """API para editar una cuenta del plan de cuentas"""
    from administrar.models import PlanCuenta
    
    try:
        cuenta = PlanCuenta.objects.get(id=id)
    except PlanCuenta.DoesNotExist:
        return JsonResponse({'error': 'Cuenta no encontrada'}, status=404)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inv谩lido'}, status=400)
    
    # Validaciones
    codigo = data.get('codigo', '').strip()
    nombre = data.get('nombre', '').strip()
    tipo = data.get('tipo', '')
    imputable = data.get('imputable', True)
    nivel = int(data.get('nivel', 1))
    padre_id = data.get('padre_id')
    
    if not codigo:
        return JsonResponse({'error': 'El c贸digo es requerido'}, status=400)
    
    if not nombre:
        return JsonResponse({'error': 'El nombre es requerido'}, status=400)
    
    if tipo not in ['ACTIVO', 'PASIVO', 'PN', 'R_POS', 'R_NEG']:
        return JsonResponse({'error': 'Tipo de cuenta inv谩lido'}, status=400)
    
    # Verificar que el c贸digo sea 煤nico (excepto para la cuenta actual)
    if PlanCuenta.objects.filter(codigo=codigo).exclude(id=id).exists():
        return JsonResponse({'error': 'Ya existe otra cuenta con ese c贸digo'}, status=400)
    
    # Verificar que el padre exista si se especifica
    padre = None
    if padre_id:
        try:
            padre = PlanCuenta.objects.get(id=padre_id)
            # No permitir que una cuenta sea padre de s铆 misma
            if padre.id == cuenta.id:
                return JsonResponse({'error': 'Una cuenta no puede ser padre de s铆 misma'}, status=400)
        except PlanCuenta.DoesNotExist:
            return JsonResponse({'error': 'La cuenta padre no existe'}, status=400)
    
    try:
        cuenta.codigo = codigo
        cuenta.nombre = nombre
        cuenta.tipo = tipo
        cuenta.imputable = imputable
        cuenta.nivel = nivel
        cuenta.padre = padre
        cuenta.save()
        
        return JsonResponse({
            'ok': True,
            'cuenta': {
                'id': cuenta.id,
                'codigo': cuenta.codigo,
                'nombre': cuenta.nombre,
                'tipo': cuenta.tipo,
                'imputable': cuenta.imputable,
                'nivel': cuenta.nivel,
                'padre_id': cuenta.padre_id
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
@csrf_exempt
@require_POST
@login_required
@verificar_permiso('contabilidad')
def api_plan_cuentas_eliminar(request, id):
    """API para eliminar una cuenta del plan de cuentas"""
    from administrar.models import PlanCuenta
    
    try:
        cuenta = PlanCuenta.objects.get(id=id)
    except PlanCuenta.DoesNotExist:
        return JsonResponse({'error': 'Cuenta no encontrada'}, status=404)
    
    # Verificar que no tenga cuentas hijas
    if cuenta.hijos.exists():
        return JsonResponse({
            'error': 'No se puede eliminar una cuenta que tiene cuentas hijas'
        }, status=400)
    
    # Verificar que no tenga asientos asociados
    if hasattr(cuenta, 'itemasiento_set') and cuenta.itemasiento_set.exists():
        return JsonResponse({
            'error': 'No se puede eliminar una cuenta que tiene asientos asociados'
        }, status=400)
    
    try:
        codigo = cuenta.codigo
        nombre = cuenta.nombre
        cuenta.delete()
        
        return JsonResponse({
            'ok': True,
            'mensaje': f'Cuenta {codigo} - {nombre} eliminada correctamente'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# =======================================
#  CONFIGURACIN DEL SISTEMA
# =======================================

@login_required
@verificar_permiso('reportes')
def reportes(request):
    return render(request, "administrar/reportes.html")

def afip_certificados(request):
    return render(request, "administrar/afip_certificados.html")

@login_required
@verificar_permiso('configuracion')
def afip_integracion(request):
    return render(request, "administrar/afip_integracion.html")

@login_required
@verificar_permiso('configuracion')
def impresora_fiscal(request):
    return render(request, "administrar/impresora_fiscal.html")

@login_required
@verificar_permiso('configuracion')
def sincronizacion(request):
    return render(request, "administrar/sincronizacion.html")

@login_required
@verificar_permiso('configuracion')
def monitor_conexiones(request):
    return render(request, "administrar/monitor_conexiones.html")

@login_required
@verificar_permiso('configuracion')
def mantenimiento(request):
    return render(request, "administrar/mantenimiento.html")

@login_required
@verificar_permiso('configuracion')
def backup(request):
    return render(request, "administrar/backup.html")

@login_required
def backup_download(request):
    """
    Genera y descarga un backup de la base de datos en formato JSON.
    """
    if not request.user.is_staff:
        messages.error(request, "No tienes permisos para realizar esta acci贸n.")
        return redirect("menu")

    try:
        # Buffer en memoria para guardar el dump
        output = io.StringIO()
        
        # Llamar al comando dumpdata
        # Excluimos contenttypes y sessions para evitar problemas al restaurar
        call_command('dumpdata', 'administrar', exclude=['contenttypes', 'auth.permission'], stdout=output)
        
        # Preparar respuesta
        response = HttpResponse(output.getvalue(), content_type='application/json')
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"backup_sistema_{timestamp}.json"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    except Exception as e:
        messages.error(request, f"Error al generar backup: {e}")
        return redirect("backup")

@login_required
@verificar_permiso('configuracion')
def logs(request):
    return render(request, "administrar/logs.html")


# =======================================
#  ADMINISTRACIN DE USUARIOS
# =======================================

@login_required
@verificar_permiso('usuarios')
def usuarios(request):
    return render(request, "administrar/usuarios.html")

@login_required
@verificar_permiso('usuarios')
def admin_usuarios(request):
    users = User.objects.select_related('perfilusuario').filter(is_active=True).exclude(is_superuser=True) # Excluir superadmin para seguridad
    return render(request, "administrar/admin_usuarios.html", {'users': users})

@require_POST
@login_required
@verificar_permiso('usuarios')
def api_usuario_crear(request):
    try:
        data = json.loads(request.body)
        username = data.get('email') # Usamos email como username
        email = data.get('email')
        password = data.get('password')
        first_name = data.get('first_name')
        
        if User.objects.filter(username=username).exists():
            return JsonResponse({'ok': False, 'error': 'El usuario ya existe'})
            
        with transaction.atomic():
            user = User.objects.create_user(username=username, email=email, password=password, first_name=first_name)
            
            # Crear perfil
            perfil = PerfilUsuario.objects.create(
                user=user,
                acceso_ventas=data.get('acceso_ventas', False),
                acceso_compras=data.get('acceso_compras', False),
                acceso_productos=data.get('acceso_productos', False),
                acceso_clientes=data.get('acceso_clientes', False),
                acceso_proveedores=data.get('acceso_proveedores', False),
                acceso_caja=data.get('acceso_caja', False),
                acceso_contabilidad=data.get('acceso_contabilidad', False),
                acceso_configuracion=data.get('acceso_configuracion', False),
                acceso_usuarios=data.get('acceso_usuarios', False),
                acceso_reportes=data.get('acceso_reportes', False)
            )
            
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@require_POST
@login_required
@verificar_permiso('usuarios')
def api_usuario_editar(request):
    try:
        data = json.loads(request.body)
        user_id = data.get('id')
        user = User.objects.get(id=user_id)
        
        with transaction.atomic():
            user.first_name = data.get('first_name')
            user.email = data.get('email')
            user.username = data.get('email') # Mantener sync
            
            if data.get('password'):
                user.set_password(data.get('password'))
                
            user.save()
            
            # Actualizar perfil
            perfil, created = PerfilUsuario.objects.get_or_create(user=user)
            perfil.acceso_ventas = data.get('acceso_ventas', False)
            perfil.acceso_compras = data.get('acceso_compras', False)
            perfil.acceso_productos = data.get('acceso_productos', False)
            perfil.acceso_clientes = data.get('acceso_clientes', False)
            perfil.acceso_proveedores = data.get('acceso_proveedores', False)
            perfil.acceso_caja = data.get('acceso_caja', False)
            perfil.acceso_contabilidad = data.get('acceso_contabilidad', False)
            perfil.acceso_configuracion = data.get('acceso_configuracion', False)
            perfil.acceso_usuarios = data.get('acceso_usuarios', False)
            perfil.acceso_reportes = data.get('acceso_reportes', False)
            perfil.save()
            
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@require_POST
@login_required
@verificar_permiso('usuarios')
def api_usuario_eliminar(request, id):
    try:
        user = User.objects.get(id=id)
        if user.is_superuser:
             return JsonResponse({'ok': False, 'error': 'No se puede eliminar al superusuario'})
        
        user.is_active = False # Soft delete
        user.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@login_required
@verificar_permiso('usuarios')
def api_usuario_detalle(request, id):
    try:
        user = User.objects.get(id=id)
        perfil = user.perfilusuario
        
        data = {
            'id': user.id,
            'first_name': user.first_name,
            'email': user.email,
            'acceso_ventas': perfil.acceso_ventas,
            'acceso_compras': perfil.acceso_compras,
            'acceso_productos': perfil.acceso_productos,
            'acceso_clientes': perfil.acceso_clientes,
            'acceso_proveedores': perfil.acceso_proveedores,
            'acceso_caja': perfil.acceso_caja,
            'acceso_contabilidad': perfil.acceso_contabilidad,
            'acceso_configuracion': perfil.acceso_configuracion,
            'acceso_usuarios': perfil.acceso_usuarios,
            'acceso_reportes': perfil.acceso_reportes,
        }
        return JsonResponse({'ok': True, 'data': data})
    except Exception as e:
         return JsonResponse({'ok': False, 'error': str(e)})

@login_required
@verificar_permiso('usuarios')
def admin_personalizado(request):
    return render(request, "administrar/admin_personalizado.html")

# ==========================================
#  MI PERFIL (Self-Service)
# ==========================================

@login_required
def mi_perfil(request):
    return render(request, "administrar/mi_perfil.html")

@require_POST
@login_required
def api_mi_perfil_password(request):
    try:
        data = json.loads(request.body)
        current_pass = data.get('current_password')
        new_pass = data.get('new_password')
        confirm_pass = data.get('confirm_password')

        if not current_pass or not new_pass:
            return JsonResponse({'ok': False, 'error': 'Faltan datos'})

        if new_pass != confirm_pass:
            return JsonResponse({'ok': False, 'error': 'Las contrase帽as nuevas no coinciden'})

        if not request.user.check_password(current_pass):
            return JsonResponse({'ok': False, 'error': 'La contrase帽a actual es incorrecta'})

        request.user.set_password(new_pass)
        request.user.save()
        
        # Mantener sesi贸n activa tras cambio de contrase帽a
        update_session_auth_hash(request, request.user)

        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@require_POST
@login_required
def api_mi_perfil_imagen(request):
    try:
        if 'imagen' not in request.FILES:
            return JsonResponse({'ok': False, 'error': 'No se recibi贸 ninguna imagen'})
            
        imagen = request.FILES['imagen']
        
        # Validar tama帽o (ej: max 2MB)
        if imagen.size > 2 * 1024 * 1024:
            return JsonResponse({'ok': False, 'error': 'La imagen es demasiado grande (m谩x 2MB)'})
            
        perfil = request.user.perfilusuario
        perfil.imagen = imagen
        perfil.save()
        
        return JsonResponse({'ok': True, 'url': perfil.imagen.url})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@login_required
def api_mi_perfil_info(request):
    """Retorna informaci贸n del perfil y permisos del usuario actual"""
    try:
        perfil, created = PerfilUsuario.objects.get_or_create(user=request.user)
        data = {
            'id': request.user.id,
            'username': request.user.username,
            'email': request.user.email,
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'is_staff': request.user.is_staff,
            'rol': 'Administrador' if request.user.is_staff else 'Vendedor',
            'permisos': {
                'ventas': perfil.acceso_ventas,
                'compras': perfil.acceso_compras,
                'productos': perfil.acceso_productos,
                'clientes': perfil.acceso_clientes,
                'proveedores': perfil.acceso_proveedores,
                'caja': perfil.acceso_caja,
                'contabilidad': perfil.acceso_contabilidad,
                'configuracion': perfil.acceso_configuracion,
                'usuarios': perfil.acceso_usuarios,
                'reportes': perfil.acceso_reportes,
                'pedidos': perfil.acceso_pedidos,
                'bancos': perfil.acceso_bancos,
                'ctacte': perfil.acceso_ctacte,
                'remitos': perfil.acceso_remitos,
            },
            'imagen_url': perfil.imagen.url if perfil.imagen else None
        }
        return JsonResponse({'ok': True, 'data': data})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@login_required
def seguridad(request):
    return render(request, "administrar/seguridad.html")


# =======================================
#  SOPORTE Y DOCUMENTACIN
# =======================================

@login_required
def ayuda_diagnostico(request):
    return render(request, "administrar/ayuda_diagnostico.html")

@login_required
def acerca(request):
    return render(request, "administrar/acerca.html")

@login_required
def ayuda(request):
    return render(request, "administrar/ayuda.html")


# =======================================
#  PRODUCTOS  CRUD HTML
# =======================================

@login_required
@verificar_permiso('productos')
def productos_lista(request):
    productos = Producto.objects.all().order_by("descripcion")
    rubros = Rubro.objects.all().order_by("nombre")
    marcas = Marca.objects.all().order_by("nombre")
    proveedores = Proveedor.objects.all().order_by("nombre")

    return render(request, "administrar/productos.html", {
        "productos": productos,
        "rubros": rubros,
        "marcas": marcas,
        "proveedores": proveedores
    })


@login_required
@verificar_permiso('productos')
def actualizar_precios(request):
    """Vista para actualizaci贸n masiva de precios"""
    rubros = Rubro.objects.all().order_by("nombre")
    marcas = Marca.objects.all().order_by("nombre")
    
    return render(request, "administrar/actualizar_precios.html", {
        "rubros": rubros,
        "marcas": marcas
    })


@require_POST
@csrf_protect
def api_actualizar_precios_masivo(request):
    """API para actualizar precios de forma masiva"""
    import json
    from decimal import Decimal
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inv谩lido"}, status=400)
    
    tipo_actualizacion = data.get("tipo_actualizacion")  # "MONTO" o "PORCENTAJE"
    valor = data.get("valor", 0)
    rubros_ids = data.get("rubros", [])  # Lista de IDs de rubros
    marcas_ids = data.get("marcas", [])  # Lista de IDs de marcas
    productos_ids = data.get("productos", [])  # Lista de IDs de productos espec铆ficos
    campos = data.get("campos", [])  # Lista de campos a actualizar
    
    # Validaciones
    if not tipo_actualizacion or tipo_actualizacion not in ["MONTO", "PORCENTAJE"]:
        return JsonResponse({"ok": False, "error": "Tipo de actualizaci贸n inv谩lido"}, status=400)
    
    if not campos:
        return JsonResponse({"ok": False, "error": "Debe seleccionar al menos un campo para actualizar"}, status=400)
    
    try:
        valor = Decimal(str(valor))
    except:
        return JsonResponse({"ok": False, "error": "Valor inv谩lido"}, status=400)
    
    # Filtrar productos
    productos = Producto.objects.all()
    
    # Si hay productos espec铆ficos seleccionados, usar solo esos
    if productos_ids:
        productos = productos.filter(id__in=productos_ids)
    else:
        # Si no hay productos espec铆ficos, aplicar filtros de rubro/marca
        if rubros_ids:
            productos = productos.filter(rubro_id__in=rubros_ids)
        
        if marcas_ids:
            productos = productos.filter(marca_id__in=marcas_ids)
    
    # Actualizar productos
    productos_actualizados = 0
    
    try:
        with transaction.atomic():
            for producto in productos:
                for campo in campos:
                    # Obtener valor actual
                    valor_actual = getattr(producto, campo, 0)
                    if valor_actual is None:
                        valor_actual = 0
                    
                    valor_actual = Decimal(str(valor_actual))
                    
                    # Calcular nuevo valor
                    if tipo_actualizacion == "MONTO":
                        nuevo_valor = valor_actual + valor
                    else:  # PORCENTAJE
                        nuevo_valor = valor_actual * (1 + valor / 100)
                    
                    # Asegurar que no sea negativo
                    nuevo_valor = max(nuevo_valor, Decimal('0'))
                    
                    # Actualizar campo
                    setattr(producto, campo, nuevo_valor)
                
                producto.save()
                productos_actualizados += 1
        
        return JsonResponse({
            "ok": True,
            "productos_actualizados": productos_actualizados,
            "mensaje": f"Se actualizaron {productos_actualizados} productos correctamente"
        })
    
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


@login_required
@verificar_permiso('productos')
def producto_nuevo(request):
    if request.method == "POST":
        marca = Marca.objects.get(id=request.POST["marca"]) if request.POST.get("marca") else None
        rubro = Rubro.objects.get(id=request.POST["rubro"]) if request.POST.get("rubro") else None
        proveedor = Proveedor.objects.get(id=request.POST["proveedor"]) if request.POST.get("proveedor") else None

        Producto.objects.create(
            codigo=request.POST["codigo"],
            descripcion=request.POST["descripcion"],
            descripcion_larga=request.POST.get("descripcion_larga", ""),
            marca=marca,
            rubro=rubro,
            proveedor=proveedor,
            tipo_bulto=request.POST.get("tipo_bulto", "UN"),
            stock_inicial=request.POST.get("stock_inicial", 0),
            stock=request.POST.get("stock", 0),
            stock_minimo=request.POST.get("stock_minimo", 0),
            stock_maximo=request.POST.get("stock_maximo", 0),
            costo=request.POST.get("costo", 0),
            precio_efectivo=request.POST.get("precio_efectivo", 0),
            precio_tarjeta=request.POST.get("precio_tarjeta", 0),
            precio_ctacte=request.POST.get("precio_ctacte", 0),
            precio_lista4=request.POST.get("precio_lista4", 0),
            imagen=request.FILES.get("imagen"),
        )
        return redirect("productos")

    return redirect("productos")


@login_required
@verificar_permiso('productos')
def producto_editar(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == "POST":
        producto.codigo = request.POST["codigo"]
        producto.descripcion = request.POST["descripcion"]
        producto.descripcion_larga = request.POST.get("descripcion_larga", "")

        producto.marca = Marca.objects.get(id=request.POST["marca"]) if request.POST.get("marca") else None
        producto.rubro = Rubro.objects.get(id=request.POST["rubro"]) if request.POST.get("rubro") else None
        producto.proveedor = Proveedor.objects.get(id=request.POST["proveedor"]) if request.POST.get("proveedor") else None

        producto.tipo_bulto = request.POST.get("tipo_bulto", "UN")
        producto.stock_inicial = request.POST.get("stock_inicial", 0)
        producto.stock = request.POST.get("stock", 0)
        producto.stock_minimo = request.POST.get("stock_minimo", 0)
        producto.stock_maximo = request.POST.get("stock_maximo", 0)

        producto.costo = request.POST.get("costo", 0)
        producto.precio_efectivo = request.POST.get("precio_efectivo", 0)
        producto.precio_tarjeta = request.POST.get("precio_tarjeta", 0)
        producto.precio_ctacte = request.POST.get("precio_ctacte", 0)
        producto.precio_lista4 = request.POST.get("precio_lista4", 0)

        if "imagen" in request.FILES:
            producto.imagen = request.FILES["imagen"]

        producto.save()
        return redirect("productos")

    return redirect("productos")


@login_required
@verificar_permiso('productos')
def producto_eliminar(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.delete()
    return redirect("productos")


# =======================================
#  PROVEEDORES  CRUD HTML
# =======================================

@login_required
@verificar_permiso('proveedores')
def proveedor_nuevo(request):
    if request.method == "POST":
        provincia = Provincia.objects.get(id=request.POST["provincia"]) if request.POST.get("provincia") else None
        localidad = Localidad.objects.get(id=request.POST["localidad"]) if request.POST.get("localidad") else None

        Proveedor.objects.create(
            nombre=request.POST["nombre"],
            cuit=request.POST.get("cuit", ""),
            telefono=request.POST.get("telefono", ""),
            email=request.POST.get("email", ""),
            direccion=request.POST.get("direccion", ""),
            provincia=provincia,
            localidad=localidad,
            notas=request.POST.get("notas", ""),
        )
        return redirect("proveedores")

    return redirect("proveedores")


@login_required
@verificar_permiso('proveedores')
def proveedor_editar(request, id):
    prov = get_object_or_404(Proveedor, id=id)

    if request.method == "POST":
        provincia = Provincia.objects.get(id=request.POST["provincia"]) if request.POST.get("provincia") else None
        localidad = Localidad.objects.get(id=request.POST["localidad"]) if request.POST.get("localidad") else None

        prov.nombre = request.POST["nombre"]
        prov.cuit = request.POST.get("cuit", "")
        prov.telefono = request.POST.get("telefono", "")
        prov.email = request.POST.get("email", "")
        prov.direccion = request.POST.get("direccion", "")
        prov.provincia = provincia
        prov.localidad = localidad
        prov.notas = request.POST.get("notas", "")
        prov.save()

        return redirect("proveedores")

    return redirect("proveedores")


@login_required
def proveedor_eliminar(request, id):
    prov = get_object_or_404(Proveedor, id=id)
    prov.delete()
    return redirect("proveedores")


# -----------------------------------------
# LISTA COMPLETA DE PROVEEDORES (API)
# -----------------------------------------
def api_proveedores_lista(request):
    proveedores = Proveedor.objects.all().order_by("nombre")

    data = [{
        "id": p.id,
        "nombre": p.nombre,
        "cuit": p.cuit,
        "telefono": p.telefono,
        "email": p.email,
    } for p in proveedores]

    return JsonResponse(data, safe=False)


@csrf_exempt
def api_proveedores_editar(request, id):
    if request.method != "POST":
        return JsonResponse({"error": "M茅todo no permitido"}, status=405)

    try:
        prov = Proveedor.objects.get(id=id)
    except Proveedor.DoesNotExist:
        return JsonResponse({"error": "Proveedor no encontrado"}, status=404)

    data = request.POST

    try:
        prov.nombre = data.get("nombre", prov.nombre)
        prov.cuit = data.get("cuit", prov.cuit)
        prov.telefono = data.get("telefono", prov.telefono)
        prov.email = data.get("email", prov.email)
        prov.direccion = data.get("direccion", prov.direccion)

        prov.provincia_id = data.get("provincia") or None
        prov.localidad_id = data.get("localidad") or None

        prov.condicion_fiscal = data.get("condicion_fiscal", prov.condicion_fiscal)
        prov.cbu = data.get("cbu", prov.cbu)
        prov.alias = data.get("alias", prov.alias)

        prov.notas = data.get("notas", prov.notas)

        prov.save()

        return JsonResponse({"ok": True})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@csrf_exempt
def api_proveedores_eliminar(request, id):
    if request.method != "POST":
        return JsonResponse({"error": "M茅todo no permitido"}, status=405)

    try:
        prov = Proveedor.objects.get(id=id)
        prov.delete()
        return JsonResponse({"ok": True})

    except Proveedor.DoesNotExist:
        return JsonResponse({"error": "Proveedor no encontrado"}, status=404)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# ======================================================
# API  Buscar productos para pedidos (compatible 100%)
# ======================================================
def api_buscar_productos(request):
    q = request.GET.get("q", "").strip()

    if len(q) < 1:
        return JsonResponse([], safe=False)

    productos = Producto.objects.filter(
        Q(codigo__icontains=q) |
        Q(descripcion__icontains=q)
    )[:20]

    data = []

    for p in productos:
        data.append({
            "id": p.id,
            "codigo": p.codigo,
            "descripcion": p.descripcion,
            "stock": p.stock,
            "precios": {
                "1": float(p.precio_efectivo),
                "2": float(p.precio_ctacte),
                "3": float(p.precio_tarjeta),
                "4": float(p.precio_lista4) if p.precio_lista4 is not None else float(p.precio_efectivo),
            }
        })

    return JsonResponse(data, safe=False)


# -----------------------------------------
# BUSCAR proveedores (por texto)
# -----------------------------------------
def api_proveedores_buscar(request):
    q = request.GET.get("q", "")

    proveedores = Proveedor.objects.filter(nombre__icontains=q)[:20]

    data = [{
        "id": p.id,
        "nombre": p.nombre,
        "cuit": p.cuit,
        "telefono": p.telefono,
    } for p in proveedores]

    return JsonResponse(data, safe=False)


def api_proveedores_detalle(request, id):
    try:
        p = Proveedor.objects.get(id=id)
    except Proveedor.DoesNotExist:
        return JsonResponse({"error": "Proveedor no encontrado"}, status=404)

    return JsonResponse({
        "id": p.id,
        "nombre": p.nombre,
        "cuit": p.cuit,
        "telefono": p.telefono,
        "email": p.email,
        "direccion": p.direccion,
        "provincia": p.provincia_id,
        "localidad": p.localidad_id,
        "condicion_fiscal": p.condicion_fiscal,
        "cbu": p.cbu,
        "alias": p.alias,
        "notas": p.notas,
    })


@csrf_exempt
def api_proveedores_nuevo(request):
    if request.method != "POST":
        return JsonResponse({"error": "M茅todo no permitido"}, status=405)

    data = request.POST

    try:
        prov = Proveedor.objects.create(
            nombre=data.get("nombre"),
            cuit=data.get("cuit", ""),
            telefono=data.get("telefono", ""),
            email=data.get("email", ""),
            direccion=data.get("direccion", ""),
            provincia_id=data.get("provincia") or None,
            localidad_id=data.get("localidad") or None,
            condicion_fiscal=data.get("condicion_fiscal", "CF"),
            cbu=data.get("cbu", ""),
            alias=data.get("alias", ""),
            notas=data.get("notas", ""),
        )

        return JsonResponse({"ok": True, "id": prov.id})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# ===============================================================
#                     CRUD MARCAS (AJAX)
# ===============================================================

@login_required
def marcas_lista(request):
    marcas = Marca.objects.all().order_by("nombre")
    return render(request, "administrar/marcas.html", {"marcas": marcas})


def api_marcas_listar(request):
    data = list(Marca.objects.values("id", "nombre", "descripcion"))
    return JsonResponse({"data": data})


def api_marcas_detalle(request, id):
    m = get_object_or_404(Marca, id=id)
    return JsonResponse({
        "id": m.id,
        "nombre": m.nombre,
        "descripcion": m.descripcion or ""
    })


@require_POST
@csrf_protect
def api_marcas_guardar(request):
    import json
    try:
        data = json.loads(request.body.decode("utf-8"))
    except:
        return JsonResponse({"error": "JSON inv谩lido"}, status=400)

    marca_id = data.get("id")
    nombre = (data.get("nombre") or "").strip()
    descripcion = (data.get("descripcion") or "").strip()

    errors = {}

    if not nombre:
        errors["nombre"] = ["El nombre es obligatorio."]

    qs = Marca.objects.filter(nombre__iexact=nombre)
    if marca_id:
        qs = qs.exclude(id=marca_id)

    if qs.exists():
        errors["nombre"] = ["Ya existe una marca con ese nombre."]

    if errors:
        return JsonResponse({"errors": errors}, status=400)

    if marca_id:
        m = get_object_or_404(Marca, id=marca_id)
        m.nombre = nombre
        m.descripcion = descripcion
        m.save()
    else:
        m = Marca.objects.create(nombre=nombre, descripcion=descripcion)

    return JsonResponse({"ok": True, "id": m.id})

def api_marcas_detalle(request, id):
    m = get_object_or_404(Marca, id=id)
    return JsonResponse({
        "id": m.id,
        "nombre": m.nombre,
        "descripcion": m.descripcion or ""
    })


@require_POST
@csrf_protect
def api_marcas_eliminar(request, id):
    m = get_object_or_404(Marca, id=id)
    try:
        m.delete()
        return JsonResponse({"ok": True})
    except:
        return JsonResponse({"error": "No se puede eliminar esta marca."}, status=400)



# =======================================
# Ь COMPRAS  ORDEN + RECEPCIN
# =======================================

def api_compras_listar(request):
    ordenes = OrdenCompra.objects.select_related("proveedor").all().order_by("-fecha")
    compras = Compra.objects.select_related("proveedor", "orden_compra").all().order_by("-fecha")

    data_ordenes = [{
        "id": o.id,
        "proveedor": o.proveedor.nombre,
        "fecha": o.fecha.strftime("%d/%m/%Y %H:%M"),
        "estado": o.estado,
        "total_estimado": o.total_estimado,
    } for o in ordenes]

    data_compras = [{
        "id": c.id,
        "proveedor": c.proveedor.nombre,
        "fecha": c.fecha.strftime("%d/%m/%Y %H:%M"),
        "orden_origen": f"OC {c.orden_compra.id}" if c.orden_compra else "-",
        "orden_compra_id": c.orden_compra.id if c.orden_compra else None,
        "total": c.total,
        "estado": c.estado,
    } for c in compras]

    return JsonResponse({"ordenes": data_ordenes, "compras": data_compras})


def api_orden_compra_detalle(request, id):
    """API para obtener el detalle de una orden de compra"""
    try:
        oc = OrdenCompra.objects.select_related("proveedor").prefetch_related("detalles__producto").get(id=id)
    except OrdenCompra.DoesNotExist:
        return JsonResponse({"error": "Orden de compra no encontrada"}, status=404)

    items = [{
        "id": d.id,
        "producto_id": d.producto.id,
        "producto_codigo": d.producto.codigo,
        "producto_descripcion": d.producto.descripcion,
        "cantidad": float(d.cantidad),
        "precio": float(d.precio),
        "subtotal": float(d.subtotal),
    } for d in oc.detalles.all()]

    return JsonResponse({
        "id": oc.id,
        "fecha": oc.fecha.strftime("%d/%m/%Y %H:%M"),
        "proveedor": oc.proveedor.nombre,
        "proveedor_cuit": oc.proveedor.cuit,
        "proveedor_telefono": oc.proveedor.telefono,
        "proveedor_email": oc.proveedor.email,
        "estado": oc.estado,
        "observaciones": oc.observaciones or "",
        "total": float(oc.total_estimado),
        "items": items,
    })


@require_POST
@csrf_protect
def api_orden_compra_guardar(request):
    """API para guardar una nueva orden de compra.
       Si recibe 'recepcionar': true, la procesa inmediatamente (impacto Stock/Contabilidad).
    """
    import json
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inv谩lido"}, status=400)

    proveedor_id = data.get("proveedor")
    observaciones = data.get("observaciones", "").strip()
    items = data.get("items", [])
    recepcionar = data.get("recepcionar", False)
    medio_pago = data.get("medio_pago", "CONTADO")
    datos_cheque = data.get("datos_cheque", None)

    if not proveedor_id:
        return JsonResponse({"ok": False, "error": "Debe seleccionar un proveedor"}, status=400)

    if not items:
        return JsonResponse({"ok": False, "error": "Debe agregar al menos un producto"}, status=400)

    try:
        proveedor = Proveedor.objects.get(id=proveedor_id)
    except Proveedor.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Proveedor no encontrado"}, status=404)

    try:
        with transaction.atomic():
            # Crear la orden de compra
            orden = OrdenCompra.objects.create(
                proveedor=proveedor,
                estado="PENDIENTE",
                observaciones=observaciones,
                total_estimado=0
            )

            total = 0
            # Crear los detalles
            for item in items:
                producto_id = item.get("producto_id")
                cantidad = item.get("cantidad", 0)
                precio = item.get("precio", 0)

                try:
                    producto = Producto.objects.get(id=producto_id)
                except Producto.DoesNotExist:
                     raise Exception(f"Producto {producto_id} no encontrado")

                subtotal = cantidad * precio
                total += subtotal

                DetalleOrdenCompra.objects.create(
                    orden=orden,
                    producto=producto,
                    cantidad=cantidad,
                    precio=precio,
                    subtotal=subtotal
                )

            # Actualizar el total
            orden.total_estimado = total
            orden.save()

            # Si se solicita la recepci贸n inmediata
            if recepcionar:
                recepcionar_orden_interna(orden, medio_pago, datos_cheque)

            return JsonResponse({"ok": True, "orden_id": orden.id})

    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


def recepcionar_orden_interna(oc, medio_pago, datos_cheque=None):
    """L贸gica interna para recepcionar una OC (Stock, Asientos, Pagos)"""
    # 1. Crear Compra
    compra = Compra.objects.create(
        proveedor=oc.proveedor,
        orden_compra=oc,
        total=oc.total_estimado,
        estado="REGISTRADA",
        observaciones=f"Recepci贸n autom谩tica de OC {oc.id}",
    )

    # 2. Detalles y Stock
    for det in oc.detalles.all():
        DetalleCompra.objects.create(
            compra=compra,
            producto=det.producto,
            cantidad=det.cantidad,
            precio=det.precio,
            subtotal=det.subtotal,
        )

        prod = det.producto
        prod.stock += det.cantidad
        prod.save()

        MovimientoStock.objects.create(
            producto=prod,
            tipo="IN",
            cantidad=det.cantidad,
            referencia=f"Compra {compra.id}",
            observaciones=f"Recepci贸n OC {oc.id}",
        )

    # 3. Actualizar estado OC
    oc.estado = "RECIBIDA"
    oc.save()
    
    # 4. Registrar movimiento financiero
    proveedor = oc.proveedor
    if medio_pago == 'CTACTE':
            # Compra a credito -> Aumenta Deuda Proveedor (HABER)
            saldo_actual = proveedor.saldo_actual
            nuevo_saldo = saldo_actual + oc.total_estimado
            
            MovimientoCuentaCorrienteProveedor.objects.create(
                proveedor=proveedor,
                tipo='HABER',
                descripcion=f"Compra (OC #{oc.id}) - Cta. Cte.",
                monto=oc.total_estimado,
                saldo=nuevo_saldo
            )
            
            proveedor.saldo_actual = nuevo_saldo
            proveedor.save()

    elif medio_pago == 'CHEQUE':
        # Pago con Cheque Propio
        from datetime import date
        from .models import Cheque
        # Si vienen datos del cheque, creamos el Cheque Propio
        cheque = None
        if datos_cheque:
            try:
                # datos_cheque = { banco: '...', numero: '...', fechaVto: 'YYYY-MM-DD' }
                cheque = Cheque.objects.create(
                    tipo='PROPIO',
                    estado='CARTERA', # O 'ENTREGADO'? Generalmente nace y se entrega.
                    banco=datos_cheque.get('banco', 'Banco'),
                    numero=datos_cheque.get('numero', ''),
                    fecha_emision=date.today(),
                    fecha_pago=datos_cheque.get('fechaVto') or date.today(),
                    monto=oc.total_estimado,
                    destinatario=oc.proveedor.nombre, # A quien se lo damos
                    observaciones=f"Pago Compra #{compra.id}"
                )
                # Actualizamos estado a ENTREGADO porque ya se lo dimos al proveedor
                cheque.estado = 'ENTREGADO' 
                cheque.save()
            except Exception as e:
                print(f"Error creando cheque propio: {e}")

    else:
        # Contado -> Egresa dinero de Caja
        caja = CajaDiaria.objects.filter(estado='ABIERTA').first()
        MovimientoCaja.objects.create(
            caja_diaria=caja,
            tipo="Egreso",
            descripcion=f"Compra #{compra.id} - {oc.proveedor.nombre}",
            monto=oc.total_estimado,
        )

    # 5. Generar Asiento Contable
    try:
        from .services import AccountingService
        AccountingService.registrar_compra(compra)
        
        if medio_pago == 'CTACTE':
            pass # Ya se registr贸 la deuda en Cta Cte, y el asiento de compra genera Proveedores (H). 
                 # No hace falta asiento de pago.
        elif medio_pago == 'CHEQUE':
            if 'cheque' in locals() and cheque:
                AccountingService.registrar_pago_compra_cheque_propio(compra, cheque)
        else:
             # Contado Efectivo
             AccountingService.registrar_pago_compra_contado(compra)

    except Exception as e:
        print(f"Error generando asiento de compra {compra.id}: {e}")


@require_POST
@csrf_protect
def api_orden_compra_recibir(request, id):
    oc = get_object_or_404(OrdenCompra, id=id)

    if oc.estado not in ["PENDIENTE", "APROBADA"]:
        return JsonResponse({"error": "La orden no est谩 en estado pendiente o aprobada"}, status=400)

    try:
        import json
        try:
            data = json.loads(request.body.decode('utf-8'))
            medio_pago = data.get('medio_pago', 'CONTADO')
        except:
            medio_pago = 'CONTADO'

        with transaction.atomic():
            recepcionar_orden_interna(oc, medio_pago, None) # TODO: Soportar cheque en recepci贸n manual diferida? Por ahora no.

        return JsonResponse({"ok": True})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@require_POST
@csrf_protect
def api_orden_compra_cancelar(request, id):
    oc = get_object_or_404(OrdenCompra, id=id)
    
    if oc.estado not in ["PENDIENTE", "BORRADOR", "APROBADA"]:
        return JsonResponse({"error": "No se puede cancelar esta orden"}, status=400)

    oc.estado = "CANCELADA"
    oc.save()
    return JsonResponse({"ok": True})


# =======================================
#  CLIENTES  API
# =======================================

def get_lista_precio_display(self):
    mapping = {
        "1": "Efectivo / Contado",
        "2": "Cuenta Corriente",
        "3": "Tarjeta",
        "4": "Mayorista",
    }
    return mapping.get(self.lista_precio, "Efectivo / Contado")


@login_required
def api_clientes_buscar(request):
    """API para buscar/listar clientes (usado por Clientes, Ventas y Cta Cte)"""
    import json
    from django.core.paginator import Paginator

    # Parametros
    q = request.GET.get("q", "").strip() or request.GET.get("busqueda", "").strip()
    page_number = request.GET.get("page", 1)
    per_page = request.GET.get("per_page", 10)
    estado_deuda = request.GET.get("estado_deuda", "todos")
    
    # QueryBase
    qs = Cliente.objects.all().order_by('nombre')

    # Filtros de Texto
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) |
            Q(dni__icontains=q) | # Add DNI search
            Q(cuit__icontains=q) |
            Q(telefono__icontains=q)
        )
    
    # Filtro de Estado de Deuda (Cta Cte)
    if estado_deuda == "con_deuda":
        qs = qs.filter(saldo_actual__gt=0)
    elif estado_deuda == "al_dia":
        qs = qs.filter(saldo_actual=0)
    elif estado_deuda == "saldo_favor":
        qs = qs.filter(saldo_actual__lt=0)

    # Paginacion
    paginator = Paginator(qs, per_page)
    try:
        page_obj = paginator.page(page_number)
    except:
        page_obj = paginator.page(1)

    datos = []
    for c in page_obj:
        datos.append({
            "id": c.id,
            "nombre": c.nombre,
            "apellido": c.apellido if hasattr(c, 'apellido') else "",
            "cuit": c.cuit or "",
            "dni": c.dni if hasattr(c, 'dni') else "",
            "condicion_fiscal": c.condicion_fiscal,
            "condicion_fiscal_display": c.get_condicion_fiscal_display(),
            "lista_precio": c.lista_precio or "EFECTIVO",
            "telefono": c.telefono or "",
            "direccion": c.domicilio or "", # Unify address
            "activo": c.activo,
            "saldo_actual": c.saldo_actual, # Frontend expects this for CtaCte
            "saldo": c.saldo_actual, # Legacy support
        })

    # Respuesta
    # Si viene con 'q' corto y sin paginacion expl铆cita, mantener formato legacy {data: []} para autocompletes viejos
    is_autocomplete = request.GET.get("page") is None and q and len(q) >= 2
    if is_autocomplete:
         return JsonResponse({"data": datos})

    # Respuesta Est谩ndar Paginada
    return JsonResponse({
        "clientes": datos,
        "total": paginator.count,
        "total_pages": paginator.num_pages,
        "current_page": page_obj.number
    })


def es_cuit_valido(cuit: str) -> bool:
    if not cuit:
        return False

    cuit = re.sub(r"[^0-9]", "", cuit)

    if len(cuit) != 11:
        return False

    try:
        nums = list(map(int, cuit))
    except ValueError:
        return False

    coef = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    suma = sum([a * b for a, b in zip(coef, nums[:10])])
    resto = suma % 11
    dv = 11 - resto
    if dv == 11:
        dv = 0
    elif dv == 10:
        dv = 9

    return dv == nums[10]


def api_buscar_clientes(request):
    q = request.GET.get("q", "")

    clientes = Cliente.objects.filter(
        Q(nombre__icontains=q) |
        Q(cuit__icontains=q)
    )[:20]

    data = []
    for c in clientes:
        data.append({
            "id": c.id,
            "nombre": c.nombre,
            "cuit": c.cuit,
            "condicion_fiscal": c.get_condicion_fiscal_display(),
            "lista_precios": c.lista_precio,
            "telefono": c.telefono,
            "email": c.email,
        })

    return JsonResponse(data, safe=False)


def api_cliente_detalle(request, id):
    try:
        c = Cliente.objects.get(id=id)
    except Cliente.DoesNotExist:
        return JsonResponse({"error": "Cliente no encontrado"}, status=404)

    return JsonResponse({
        "id": c.id,
        "nombre": c.nombre,
        "tipo_cliente": c.tipo_cliente,
        "cuit": c.cuit or "",
        "condicion_fiscal": c.condicion_fiscal,
        "domicilio": c.domicilio or "",
        "provincia": c.provincia_id,
        "localidad": c.localidad_id,
        "telefono": c.telefono or "",
        "email": c.email or "",
        "lista_precio": c.lista_precio,
        "tipo_factura_preferida": c.tipo_factura_preferida,
        "descuento_predeterminado": float(c.descuento_predeterminado or 0),
        "tiene_ctacte": c.tiene_ctacte,
        "limite_credito": float(c.limite_credito or 0),
        "permitir_superar_limite": c.permitir_superar_limite,
        "contacto_nombre": c.contacto_nombre or "",
        "contacto_telefono": c.contacto_telefono or "",
        "contacto_email": c.contacto_email or "",
        "rubro_cliente": c.rubro_cliente or "",
        "canal": c.canal or "",
        "origen": c.origen or "",
        "notas": c.notas or "",
        "activo": c.activo,
    })


@csrf_exempt
def api_cliente_nuevo(request):
    if request.method != "POST":
        return JsonResponse({"error": "M茅todo inv谩lido"}, status=400)

    data = request.POST
    errors = {}

    nombre = data.get("nombre", "").strip()
    if not nombre:
        errors["nombre"] = "El nombre es obligatorio."

    email = data.get("email", "").strip()
    telefono = data.get("telefono", "").strip()
    cuit = data.get("cuit", "").strip()

    if cuit and not re.match(r"^\d{2}-?\d{8}-?\d$", cuit):
        errors["cuit"] = "El CUIT debe tener formato v谩lido (20-12345678-3)."

    if email and not re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,4}$', email):
        errors["email"] = "Email inv谩lido."

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    c = Cliente.objects.create(
        nombre=nombre,
        tipo_cliente=data.get("tipo_cliente", "P"),
        cuit=cuit,
        condicion_fiscal=data.get("condicion_fiscal", "CF"),
        domicilio=data.get("domicilio", "").strip(),
        provincia_id=data.get("provincia") or None,
        localidad_id=data.get("localidad") or None,
        telefono=telefono,
        email=email,
        lista_precio=data.get("lista_precio", "1"),
        tipo_factura_preferida=data.get("tipo_factura_preferida", "B"),
        descuento_predeterminado=data.get("descuento_predeterminado") or 0,
        tiene_ctacte=data.get("tiene_ctacte") == "on",
        limite_credito=data.get("limite_credito") or 0,
        permitir_superar_limite=data.get("permitir_superar_limite") == "on",
        contacto_nombre=data.get("contacto_nombre", "").strip(),
        contacto_telefono=data.get("contacto_telefono", "").strip(),
        contacto_email=data.get("contacto_email", "").strip(),
        rubro_cliente=data.get("rubro_cliente", "").strip(),
        canal=data.get("canal", ""),
        origen=data.get("origen", ""),
        notas=data.get("notas", "").strip(),
        activo=data.get("activo") == "on",
    )

    return JsonResponse({
        "ok": True,
        "cliente": {
            "id": c.id,
            "nombre": c.nombre,
            "cuit": c.cuit,
            "telefono": c.telefono,
            "condicion_fiscal_display": c.get_condicion_fiscal_display(),
            "lista_precio_display": c.get_lista_precio_display(),
            "tiene_ctacte": c.tiene_ctacte,
        }
    })


@csrf_exempt
def api_cliente_editar(request, id):
    if request.method != "POST":
        return JsonResponse({"error": "M茅todo inv谩lido"}, status=400)

    try:
        c = Cliente.objects.get(id=id)
    except Cliente.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Cliente no encontrado."}, status=404)

    data = request.POST
    errors = {}

    nombre = data.get("nombre", "").strip()
    if not nombre:
        errors["nombre"] = "El nombre es obligatorio."

    email = data.get("email", "").strip()
    cuit = data.get("cuit", "").strip()
    telefono = data.get("telefono", "").strip()

    if cuit and not re.match(r"^\d{2}-?\d{8}-?\d$", cuit):
        errors["cuit"] = "El CUIT debe tener un formato v谩lido (20-12345678-3)."

    if email and not re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,4}$', email):
        errors["email"] = "Email inv谩lido."

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    c.nombre = nombre
    c.tipo_cliente = data.get("tipo_cliente", "P")
    c.cuit = cuit
    c.condicion_fiscal = data.get("condicion_fiscal", "CF")
    c.domicilio = data.get("domicilio", "").strip()
    c.provincia_id = data.get("provincia") or None
    c.localidad_id = data.get("localidad") or None
    c.telefono = telefono
    c.email = email
    c.lista_precio = data.get("lista_precio", "1")
    c.tipo_factura_preferida = data.get("tipo_factura_preferida", "B")
    c.descuento_predeterminado = data.get("descuento_predeterminado") or 0
    c.tiene_ctacte = data.get("tiene_ctacte") == "on"
    c.limite_credito = data.get("limite_credito") or 0
    c.permitir_superar_limite = data.get("permitir_superar_limite") == "on"
    c.contacto_nombre = data.get("contacto_nombre", "").strip()
    c.contacto_telefono = data.get("contacto_telefono", "").strip()
    c.contacto_email = data.get("contacto_email", "").strip()
    c.rubro_cliente = data.get("rubro_cliente", "").strip()
    c.canal = data.get("canal", "")
    c.origen = data.get("origen", "")
    c.notas = data.get("notas", "").strip()
    c.activo = data.get("activo") == "on"

    c.save()

    return JsonResponse({
        "ok": True,
        "cliente": {
            "id": c.id,
            "nombre": c.nombre,
            "cuit": c.cuit,
            "telefono": c.telefono,
            "condicion_fiscal_display": c.get_condicion_fiscal_display(),
            "lista_precio_display": f"Lista {c.lista_precio}",
            "tiene_ctacte": c.tiene_ctacte,
        }
    })


@csrf_exempt
def api_cliente_eliminar(request, id):
    try:
        c = Cliente.objects.get(id=id)
        c.delete()
        return JsonResponse({"ok": True})
    except Cliente.DoesNotExist:
        return JsonResponse({"error": "Cliente no encontrado"}, status=404)


# =======================================
#  PROVINCIAS / LOCALIDADES  API
# =======================================

def api_provincias(request):
    provincias = list(Provincia.objects.all().values("id", "nombre"))
    return JsonResponse(provincias, safe=False)


def api_localidades(request, provincia_id):
    localidades = list(
        Localidad.objects.filter(provincia_id=provincia_id).values("id", "nombre")
    )
    return JsonResponse(localidades, safe=False)


# =======================================
#  PRODUCTOS  API GENERAL (OFICIAL)
# =======================================

def api_productos_lista(request):
    """API para listar productos con paginaci贸n y filtros"""
    try:
        # Par谩metros de paginaci贸n
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        # Filtros
        busqueda = request.GET.get('busqueda', '').strip()
        rubro_id = request.GET.get('rubro', '').strip()
        marca_id = request.GET.get('marca', '').strip()
        stock_filter = request.GET.get('stock', 'todos').strip()
        
        # Query base
        productos = Producto.objects.select_related("marca", "rubro").all()
        
        # Aplicar filtros
        if busqueda:
            productos = productos.filter(
                Q(codigo__icontains=busqueda) |
                Q(descripcion__icontains=busqueda)
            )
        
        if rubro_id:
            productos = productos.filter(rubro_id=rubro_id)
            
        if marca_id:
            productos = productos.filter(marca_id=marca_id)
            
        if stock_filter == 'con_stock':
            productos = productos.filter(stock__gt=0)
        elif stock_filter == 'sin_stock':
            productos = productos.filter(stock__lte=0)
        elif stock_filter in ['stock_bajo', 'bajo']:
            # Consideramos stock bajo <= 10, o <= stock_minimo
            from django.db.models import F
            productos = productos.filter(Q(stock__lte=10) | Q(stock__lte=F('stock_minimo')))
        
        # Ordenar
        productos = productos.order_by('descripcion')
        
        # Contar total
        total = productos.count()
        total_pages = (total + per_page - 1) // per_page
        
        # Paginaci贸n
        start = (page - 1) * per_page
        end = start + per_page
        productos_paginados = productos[start:end]
        
        data = []
        for p in productos_paginados:
            data.append({
                "id": p.id,
                "codigo": p.codigo,
                "descripcion": p.descripcion,
                "marca": p.marca.nombre if p.marca else "",
                "rubro": p.rubro.nombre if p.rubro else "",
                "precio_efectivo": float(p.precio_efectivo),
                "costo": float(p.costo) if p.costo else 0,
                "stock": p.stock,
            })

        return JsonResponse({
            'productos': data,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages
        })
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=500)





def api_producto_info(request, id, lista):
    try:
        p = Producto.objects.only(
            "id", "codigo", "descripcion", "stock",
            "precio_efectivo", "precio_ctacte", "precio_tarjeta", "precio_lista4"
        ).get(pk=id)
    except Producto.DoesNotExist:
        return JsonResponse({"error": "Producto no encontrado"}, status=404)

    precios = {
        "1": p.precio_efectivo,
        "2": p.precio_ctacte,
        "3": p.precio_tarjeta,
        "4": p.precio_lista4,
    }
    precio = precios.get(lista, p.precio_efectivo)

    return JsonResponse({
        "id": p.id,
        "codigo": p.codigo,
        "descripcion": p.descripcion,
        "stock": p.stock,
        "precio_seleccionado": float(precio),
        "precios": {
            "efectivo": float(p.precio_efectivo),
            "ctacte": float(p.precio_ctacte),
            "tarjeta": float(p.precio_tarjeta),
            "mayorista": float(p.precio_lista4 or 0),
        }
    })

@login_required
def presupuesto_editar(request, id):
    """Vista para editar un presupuesto existente"""
    presupuesto = get_object_or_404(Presupuesto, pk=id)
    
    # Validar estado
    if presupuesto.estado != 'PENDIENTE':
        # Solo se pueden editar pendientes (o decidir politica)
        # Podriamos redirigir con un mensaje
        pass 

    # Serializar items
    items_data = []
    for det in presupuesto.detalles.all():
        items_data.append({
            "id": det.producto.id if det.producto else None,
            "descripcion": det.descripcion_producto,
            "cantidad": float(det.cantidad),
            "precio": float(det.precio_unitario),
            "subtotal": float(det.subtotal)
        })

    # Datos iniciales para el JS
    context = {
        "presupuesto": presupuesto,
        "items_json": json.dumps(items_data),
        "cliente_json": json.dumps({
            "id": presupuesto.cliente.id if presupuesto.cliente else None,
            "nombre": presupuesto.cliente.nombre if presupuesto.cliente else "",
            "cuit": presupuesto.cliente.cuit if presupuesto.cliente else "",
            "condicion_fiscal": presupuesto.cliente.condicion_fiscal if presupuesto.cliente else "",
            "lista_precio": presupuesto.cliente.lista_precio if presupuesto.cliente else "1"
        })
    }
    return render(request, "administrar/presupuesto_nuevo.html", context)


def api_productos_detalle(request, id):
    """
    DETALLE CANNICO DE PRODUCTO
    Usado por productos.js para cargar el modal EDITAR.
    """
    try:
        p = Producto.objects.get(id=id)
    except Producto.DoesNotExist:
        return JsonResponse({"error": "Producto no encontrado"}, status=404)

    return JsonResponse({
        "id": p.id,
        "codigo": p.codigo,
        "descripcion": p.descripcion,
        "descripcion_larga": p.descripcion_larga,

        "marca": p.marca_id,
        "rubro": p.rubro_id,
        "proveedor": p.proveedor_id,

        "tipo_bulto": p.tipo_bulto,
        "stock_inicial": p.stock_inicial,
        "stock": p.stock,
        "stock_minimo": p.stock_minimo,
        "stock_maximo": p.stock_maximo,

        "costo": float(p.costo),
        "precio_efectivo": float(p.precio_efectivo),
        "precio_tarjeta": float(p.precio_tarjeta),
        "precio_ctacte": float(p.precio_ctacte),
        "precio_lista4": float(p.precio_lista4 or 0),

        "imagen_url": p.imagen.url if p.imagen else ""
    })


@csrf_exempt
def api_productos_nuevo(request):
    if request.method != "POST":
        return JsonResponse({"error": "M茅todo no permitido"}, status=405)

    errors = {}

    codigo = request.POST.get("codigo", "").strip()

    if not codigo:
        errors["codigo"] = ["El c贸digo es obligatorio."]
    else:
        if Producto.objects.filter(codigo__iexact=codigo).exists():
            errors["codigo"] = ["Este c贸digo ya existe. Eleg铆 otro."]

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    try:
        p = Producto(
            codigo=codigo,
            descripcion=request.POST.get("descripcion", "").strip(),
            descripcion_larga=request.POST.get("descripcion_larga", "").strip(),

            tipo_bulto=request.POST.get("tipo_bulto") or "UN",
            stock_inicial=int(request.POST.get("stock_inicial", 0) or 0),
            stock=int(request.POST.get("stock", 0) or 0),
            stock_minimo=int(request.POST.get("stock_minimo", 0) or 0),
            stock_maximo=int(request.POST.get("stock_maximo", 0) or 0),

            costo=float(request.POST.get("costo", 0) or 0),
            precio_efectivo=float(request.POST.get("precio_efectivo", 0) or 0),
            precio_tarjeta=float(request.POST.get("precio_tarjeta", 0) or 0),
            precio_ctacte=float(request.POST.get("precio_ctacte", 0) or 0),
            precio_lista4=float(request.POST.get("precio_lista4", 0) or 0),

            marca_id=request.POST.get("marca") or None,
            rubro_id=request.POST.get("rubro") or None,
            proveedor_id=request.POST.get("proveedor") or None,
        )

        if "imagen" in request.FILES:
            p.imagen = request.FILES["imagen"]

        p.save()
        return JsonResponse({"ok": True})

    except Exception as e:
        print("ERROR NUEVO PRODUCTO:", str(e))
        return JsonResponse({
            "ok": False,
            "error": "Ocurri贸 un error inesperado al guardar el producto."
        }, status=500)


@csrf_exempt
def api_productos_editar(request, id):
    if request.method != "POST":
        return JsonResponse({"error": "M茅todo no permitido"}, status=405)

    try:
        p = Producto.objects.get(pk=id)
    except Producto.DoesNotExist:
        return JsonResponse({
            "ok": False,
            "error": "Producto no encontrado."
        }, status=404)

    data = request.POST
    errors = {}

    codigo_nuevo = data.get("codigo", "").strip()

    if not codigo_nuevo:
        errors["codigo"] = ["El c贸digo es obligatorio."]
    else:
        if Producto.objects.filter(codigo__iexact=codigo_nuevo).exclude(id=id).exists():
            errors["codigo"] = ["Ya existe otro producto con ese c贸digo."]

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    def num(x):
        try:
            if x in ["", None]:
                return 0
            return float(x)
        except Exception:
            return 0

    p.codigo = codigo_nuevo
    p.descripcion = data.get("descripcion", "").strip()
    p.descripcion_larga = data.get("descripcion_larga", "").strip()

    p.tipo_bulto = data.get("tipo_bulto", p.tipo_bulto)

    p.stock_inicial = num(data.get("stock_inicial"))
    p.stock = num(data.get("stock"))
    p.stock_minimo = num(data.get("stock_minimo"))
    p.stock_maximo = num(data.get("stock_maximo"))

    p.costo = num(data.get("costo"))
    p.precio_efectivo = num(data.get("precio_efectivo"))
    p.precio_tarjeta = num(data.get("precio_tarjeta"))
    p.precio_ctacte = num(data.get("precio_ctacte"))
    p.precio_lista4 = num(data.get("precio_lista4"))

    p.marca_id = data.get("marca") or None
    p.rubro_id = data.get("rubro") or None
    p.proveedor_id = data.get("proveedor") or None

    if "imagen" in request.FILES:
        p.imagen = request.FILES["imagen"]

    try:
        p.save()
        return JsonResponse({"ok": True})

    except Exception as e:
        print("ERROR EDITAR PRODUCTO:", e)
        return JsonResponse({
            "ok": False,
            "error": "Ocurri贸 un error inesperado al guardar el producto."
        }, status=500)


@csrf_exempt
def api_productos_eliminar(request, id):
    try:
        Producto.objects.get(id=id).delete()
    except Producto.DoesNotExist:
        return JsonResponse({"error": "Producto no encontrado"}, status=404)

    return JsonResponse({"success": True})


@login_required
def productos_lista(request):
    """
    Vista principal para la p谩gina de productos.
    Renderiza el template con los datos necesarios para los filtros.
    """
    return render(request, "administrar/productos.html")




# ===============================================================
#                     CRUD MARCAS (AJAX)
# ===============================================================

@login_required
def marcas_lista(request):
    return render(request, "administrar/marcas.html")


def api_marcas_listar(request):
    marcas = list(Marca.objects.all().values("id", "nombre", "descripcion").order_by("nombre"))
    return JsonResponse({"data": marcas})


def api_marcas_detalle(request, id):
    m = get_object_or_404(Marca, id=id)
    return JsonResponse({
        "id": m.id,
        "nombre": m.nombre,
        "descripcion": m.descripcion or ""
    })


@require_POST
@csrf_protect
def api_marcas_guardar(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"error": "JSON inv谩lido."}, status=400)

    marca_id = data.get("id")
    nombre = (data.get("nombre") or "").strip()
    descripcion = (data.get("descripcion") or "").strip()

    errors = {}

    if not nombre:
        errors["nombre"] = ["El nombre es obligatorio."]

    if nombre:
        qs = Marca.objects.filter(nombre__iexact=nombre)
        if marca_id:
            qs = qs.exclude(id=marca_id)
        if qs.exists():
            errors["nombre"] = ["Ya existe una marca con ese nombre."]

    if errors:
        return JsonResponse({"errors": errors}, status=400)

    if marca_id:
        m = get_object_or_404(Marca, id=marca_id)
        m.nombre = nombre
        m.descripcion = descripcion
        m.save()
    else:
        m = Marca.objects.create(nombre=nombre, descripcion=descripcion)

    return JsonResponse({"ok": True, "id": m.id})


@require_POST
@csrf_protect
def api_marcas_eliminar(request, id):
    m = get_object_or_404(Marca, id=id)
    try:
        m.delete()
        return JsonResponse({"ok": True})
    except Exception:
        return JsonResponse({"error": "No se puede eliminar (puede estar en uso)."}, status=400)



def api_proveedores(request):
    proveedores = list(Proveedor.objects.all().values("id", "nombre"))
    return JsonResponse(proveedores, safe=False)


# =========================================
#  VERIFICAR CDIGO DE PRODUCTO
# =========================================
def api_productos_verificar_codigo(request):
    codigo = request.GET.get("codigo", "").strip()
    id_actual = request.GET.get("id", None)

    if not codigo:
        return JsonResponse({"existe": False})

    qs = Producto.objects.filter(codigo__iexact=codigo)

    if id_actual:
        qs = qs.exclude(id=id_actual)

    return JsonResponse({"existe": qs.exists()})


# =======================================
#  VENTAS  LGICA
# =======================================

@login_required
def buscar_producto(request):
    texto = request.GET.get("q", "").strip()

    if texto == "":
        return JsonResponse({"ok": False, "msg": "Vac铆o"}, status=400)

    productos = Producto.objects.filter(
        Q(codigo__icontains=texto) |
        Q(descripcion__icontains=texto) |
        Q(marca__nombre__icontains=texto)
    ).select_related("marca")[:5]

    resultados = []

    for p in productos:
        resultados.append({
            "id": p.id,
            "nombre": p.descripcion,
            "precio": float(p.precio_efectivo),
            "stock": p.stock,
            "marca": p.marca.nombre if p.marca else "",
            "codigo": p.codigo,
        })

    return JsonResponse({"ok": True, "resultados": resultados})


@login_required
def buscar_productos(request):
    q = request.GET.get("q", "").strip()

    productos = Producto.objects.filter(
        Q(codigo__icontains=q) |
        Q(descripcion__icontains=q) |
        Q(marca__nombre__icontains=q) |
        Q(rubro__nombre__icontains=q)
    ).values(
        "id",
        "codigo",
        "descripcion",
        "precio_efectivo",
        "precio_tarjeta",
        "precio_ctacte",
        "stock"
    )[:10]

    return JsonResponse(list(productos), safe=False)


# =======================================
#  VENTAS - API
# =======================================

def api_ventas_listar(request):
    """API para listar todas las ventas"""
    ventas = Venta.objects.select_related("cliente").all().order_by("-fecha")
    
    data = []
    for v in ventas:
        data.append({
            "id": v.id,
            "fecha": v.fecha.strftime("%d/%m/%Y %H:%M"),
            "cliente": v.cliente.nombre if v.cliente else "Sin cliente",
            "tipo_comprobante": v.tipo_comprobante,
            "total": float(v.total),
            "estado": v.estado,
        })
    
    return JsonResponse({"ok": True, "data": data})

@login_required
def api_venta_detalle(request, id):
    """API para obtener detalles de una venta espec铆fica"""
    try:
        venta = Venta.objects.select_related('cliente', 'pedido_origen').get(pk=id)
        
        detalles = []
        for det in venta.detalles.select_related('producto').all():
            detalles.append({
                'id': det.id,
                'producto_id': det.producto.id,
                'producto_codigo': det.producto.codigo,
                'producto_descripcion': det.producto.descripcion,
                'cantidad': float(det.cantidad),
                'precio_unitario': float(det.precio_unitario),
                'subtotal': float(det.subtotal)
            })

        data = {
            'id': venta.id,
            'fecha': venta.fecha.strftime('%d/%m/%Y %H:%M'),
            'cliente_nombre': venta.cliente.nombre,
            'cliente_telefono': venta.cliente.telefono,
            'cliente_email': venta.cliente.email,
            'tipo_comprobante': venta.tipo_comprobante,
            'total': float(venta.total),
            'estado': venta.estado,
            'medio_pago': venta.medio_pago,
            'cae': venta.cae,
            'pedido_origen_id': venta.pedido_origen.id if venta.pedido_origen else None,
            'detalles': detalles
        }
        return JsonResponse(data)
    except Venta.DoesNotExist:
        return JsonResponse({'error': 'Venta no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_POST
@csrf_protect
def api_venta_guardar(request):
    """API para guardar una nueva venta"""
    import json
    from decimal import Decimal
    try:
        data = json.loads(request.body.decode("utf-8"))
    except:
        return JsonResponse({"error": "JSON inv谩lido"}, status=400)

    try:
        with open("debug_remito.log", "a") as f:
            f.write(f"Recibido: {data}\n")
    except:
        pass

    cliente_id = data.get("cliente_id")
    items = data.get("items", [])
    total_general = Decimal(str(data.get("total_general", 0)))
    medio_pago = data.get("medio_pago", "EFECTIVO")
    
    if not items:
        return JsonResponse({"error": "Debe agregar al menos un producto"}, status=400)

    try:
        # Obtener o crear cliente
        if cliente_id and cliente_id != "":
            cliente = Cliente.objects.get(id=cliente_id)
        else:
            cliente = Cliente.objects.filter(nombre="Consumidor Final").first()
            if not cliente:
                cliente = Cliente.objects.create(
                    nombre="Consumidor Final",
                    tipo_cliente="P",
                    condicion_fiscal="CF",
                )

        with transaction.atomic():
            venta = Venta.objects.create(
                cliente=cliente,
                tipo_comprobante="B",
                total=total_general,
                estado="Emitida",
                medio_pago=medio_pago
            )

            for item in items:
                producto = Producto.objects.get(id=item["id"])
                cantidad = Decimal(str(item["cantidad"]))
                
                # VALIDACIN DE STOCK
                if producto.stock < cantidad:
                    raise ValueError(f"Stock insuficiente para {producto.descripcion}. Stock actual: {producto.stock}")

                precio = Decimal(str(item["precio"]))
                subtotal = Decimal(str(item["subtotal"]))

                DetalleVenta.objects.create(
                    venta=venta,
                    producto=producto,
                    cantidad=cantidad,
                    precio_unitario=precio,
                    subtotal=subtotal,
                )

                producto.stock -= cantidad
                producto.save()

            # Registrar movimiento segun medio de pago
            cheque_creado = None  # Para uso posterior en contabilidad
            
            if medio_pago == "CTACTE":
                # Venta a cuenta corriente - No impacta caja, solo registra deuda del cliente
                nuevo_saldo = cliente.saldo_actual + Decimal(total_general)
                
                MovimientoCuentaCorriente.objects.create(
                    cliente=cliente,
                    tipo='DEBE',
                    descripcion=f"Venta #{venta.id} (Cta. Cte.)",
                    monto=total_general,
                    saldo=nuevo_saldo,
                    venta=venta
                )
                
                cliente.saldo_actual = nuevo_saldo
                cliente.save()
                
            elif medio_pago == "CHEQUE":
                # Venta con cheque - Crear cheque de tercero en cartera
                # El cheque NO impacta caja hasta que se deposite
                cheque_data = data.get("cheque", {})
                
                from .models import Cheque
                from datetime import datetime
                
                # Parsear fecha de pago si viene como string
                fecha_pago = cheque_data.get("fecha_pago")
                if isinstance(fecha_pago, str):
                    try:
                        fecha_pago = datetime.strptime(fecha_pago, "%Y-%m-%d").date()
                    except:
                        fecha_pago = venta.fecha.date()
                elif not fecha_pago:
                    fecha_pago = venta.fecha.date()
                
                cheque_creado = Cheque.objects.create(
                    tipo='TERCERO',
                    numero=cheque_data.get("numero", f"VTA-{venta.id}"),
                    banco=cheque_data.get("banco", "Sin especificar"),
                    firmante=cheque_data.get("firmante", cliente.nombre),
                    cuit_firmante=cheque_data.get("cuit", cliente.cuit or ""),
                    monto=Decimal(str(total_general)),
                    fecha_emision=venta.fecha.date(),
                    fecha_pago=fecha_pago,
                    estado='CARTERA',
                    cliente=cliente,
                    observaciones=f"Cobro Venta #{venta.id}"
                )
                # NO registramos movimiento de caja para cheques (van a cartera)
                # NO registramos movimiento de caja para cheques (van a cartera)
                
            else:
                # EFECTIVO o TARJETA - Impactan en Caja inmediatamente
                caja = CajaDiaria.objects.filter(estado='ABIERTA').first()
                MovimientoCaja.objects.create(
                    caja_diaria=caja,
                    usuario=request.user if request.user.is_authenticated else None,
                    tipo="Ingreso",
                    descripcion=f"Venta #{venta.id} - {medio_pago}",
                    monto=total_general,
                )

            # Generar Remito Autom谩tico (Si la empresa lo tiene habilitado)
            # Ignoramos lo que venga del frontend, mandamos la configuraci贸n de la empresa
            empresa_config = Empresa.objects.first()
            if empresa_config and empresa_config.habilita_remitos:
                # Buscar 煤ltimo n煤mero de remito para mantener correlatividad (opcional, o dejar que id lo maneje)
                remito = Remito.objects.create(
                    cliente=cliente,
                    venta_asociada=venta,
                    fecha=venta.fecha,
                    direccion_entrega=cliente.domicilio or "Retiro en Local",
                    estado='ENTREGADO',
                    punto_venta=empresa_config.punto_venta
                )
                
                for det in venta.detalles.all():
                    DetalleRemito.objects.create(
                        remito=remito,
                        producto=det.producto,
                        cantidad=det.cantidad
                    )

        # Generar Asientos Contables (fuera del transaction para que la venta ya exista)
        try:
            from .services import AccountingService
            
            # 1. Asiento de Venta (com煤n a todos los medios de pago)
            # Debe: Deudores por Ventas | Haber: Ventas + IVA + CMV
            AccountingService.registrar_venta(venta)

            # 2. Asiento de Cobro (seg煤n medio de pago)
            if medio_pago == "EFECTIVO":
                # Debe: Caja | Haber: Deudores por Ventas
                AccountingService.registrar_cobro_venta_contado(venta)
                
            elif medio_pago == "TARJETA":
                # Debe: Tarjetas a Cobrar + Comisiones | Haber: Deudores por Ventas
                AccountingService.registrar_cobro_venta_tarjeta(venta)
                
            elif medio_pago == "CHEQUE" and cheque_creado:
                # Debe: Valores a Depositar | Haber: Deudores por Ventas
                AccountingService.registrar_cobro_venta_cheque(venta, cheque_creado)
                
            # CTACTE: No genera asiento de cobro inmediato (el cliente debe, se cobra despu茅s)
                
        except Exception as e:
            print(f"Error generando asiento de venta {venta.id}: {e}")

        return JsonResponse({"ok": True, "venta_id": venta.id})

    except ValueError as e:
        return JsonResponse({"error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"error": f"Error al guardar la venta: {str(e)}"}, status=500)





@login_required
def api_productos_buscar(request):
    """API para buscar productos (autocomplete)"""
    query = request.GET.get("q", "").strip()
    
    if len(query) < 2:
        return JsonResponse({"data": []})
    
    productos = Producto.objects.filter(
        Q(codigo__icontains=query) |
        Q(descripcion__icontains=query) |
        Q(marca__nombre__icontains=query) |
        Q(rubro__nombre__icontains=query)
    ).select_related("marca", "rubro")[:15]
    
    data = []
    for p in productos:
        data.append({
            "id": p.id,
            "codigo": p.codigo,
            "descripcion": p.descripcion,
            "marca": p.marca.nombre if p.marca else "",
            "rubro": p.rubro.nombre if p.rubro else "",
            "precio_efectivo": float(p.precio_efectivo),
            "precio_tarjeta": float(p.precio_tarjeta),
            "costo": float(p.costo) if p.costo else 0,
            "stock": p.stock,
        })
    
    return JsonResponse({"data": data})



#==========================================
# Api. NUEVO PEDIDO
# =========================================
@require_POST
def api_pedido_nuevo(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inv谩lido."}, status=400)

    cliente_id = data.get("cliente_id")
    items = data.get("items", [])
    observaciones = data.get("observaciones", "").strip()

    if not cliente_id:
        return JsonResponse({"ok": False, "error": "Debe seleccionar un cliente."}, status=400)

    if not items:
        return JsonResponse({"ok": False, "error": "El pedido no tiene 铆tems."}, status=400)

    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Cliente no encontrado."}, status=400)

    with transaction.atomic():
        pedido = Pedido.objects.create(
            cliente=cliente,
            estado="PENDIENTE",
            total=0,
            observaciones=observaciones,
        )

        total = Decimal("0.00")

        for it in items:
            try:
                prod_id = int(it.get("id"))
                cantidad = Decimal(str(it.get("cantidad", "0")))
                precio = Decimal(str(it.get("precio", "0")))
            except Exception:
                transaction.set_rollback(True)
                return JsonResponse({"ok": False, "error": "Datos de 铆tems inv谩lidos."}, status=400)

            if cantidad <= 0 or precio < 0:
                transaction.set_rollback(True)
                return JsonResponse({"ok": False, "error": "Cantidad o precio inv谩lidos."}, status=400)

            try:
                prod = Producto.objects.get(pk=prod_id)
            except Producto.DoesNotExist:
                transaction.set_rollback(True)
                return JsonResponse({"ok": False, "error": f"Producto ID {prod_id} no existe."}, status=400)

            subtotal = cantidad * precio
            total += subtotal

            DetallePedido.objects.create(
                pedido=pedido,
                producto=prod,
                cantidad=cantidad,
                precio_unitario=precio,
                subtotal=subtotal,
            )

        pedido.total = total
        pedido.save()

    return JsonResponse({"ok": True, "pedido_id": pedido.id})


# =========================================
# Api. CAMBIAR ESTADO PEDIDOS
# =========================================
@require_POST
def api_pedido_cambiar_estado(request, pedido_id):
    nuevo_estado = request.POST.get("estado")

    if nuevo_estado not in dict(Pedido.ESTADO_PEDIDO):
        return JsonResponse({"ok": False, "error": "Estado inv谩lido."}, status=400)

    try:
        pedido = Pedido.objects.get(pk=pedido_id)
    except Pedido.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Pedido no encontrado."}, status=404)

    if pedido.estado == "FACTURADO" and nuevo_estado != "FACTURADO":
        return JsonResponse({"ok": False, "error": "El pedido ya est谩 facturado."}, status=400)

    pedido.estado = nuevo_estado
    pedido.save()

    return JsonResponse({"ok": True, "estado": pedido.get_estado_display()})


# =========================================
# Api. CAMBIAR DE PEDIDO A VENTA
# =========================================
@require_POST
def api_pedido_facturar(request, pedido_id):
    try:
        pedido = Pedido.objects.select_related("cliente").prefetch_related("detalles").get(pk=pedido_id)
    except Pedido.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Pedido no encontrado."}, status=404)

    if pedido.estado == "FACTURADO" and pedido.venta_id:
        return JsonResponse({"ok": False, "error": "El pedido ya est谩 facturado."}, status=400)

    with transaction.atomic():
        venta = Venta.objects.create(
            cliente=pedido.cliente,
            tipo_comprobante="B",
            total=0,
            estado="Emitida",
        )

        total = Decimal("0.00")

        for det in pedido.detalles.all():
            DetalleVenta.objects.create(
                venta=venta,
                producto=det.producto,
                cantidad=det.cantidad,
                precio_unitario=det.precio_unitario,
                subtotal=det.subtotal,
            )

            total += det.subtotal

            prod = det.producto
            prod.stock = prod.stock - det.cantidad
            prod.save()

            MovimientoStock.objects.create(
                producto=prod,
                tipo="OUT",
                cantidad=det.cantidad,
                referencia=f"Venta {venta.id}",
                observaciones=f"Desde pedido {pedido.id}",
            )

        venta.total = total
        venta.save()

        pedido.estado = "FACTURADO"
        pedido.venta = venta
        pedido.save()

    return JsonResponse({"ok": True, "venta_id": venta.id})


# =========================================
# PANTALLAS DE PEDIDOS
# =========================================

@login_required
def pedidos_listado(request):
    estado = request.GET.get("estado", "")
    cliente_id = request.GET.get("cliente", "")
    fecha_desde = request.GET.get("desde", "")
    fecha_hasta = request.GET.get("hasta", "")

    pedidos = Pedido.objects.select_related("cliente").all().order_by("-fecha")

    if estado:
        pedidos = pedidos.filter(estado=estado)

    if cliente_id:
        pedidos = pedidos.filter(cliente_id=cliente_id)

    if fecha_desde:
        pedidos = pedidos.filter(fecha__date__gte=fecha_desde)

    if fecha_hasta:
        pedidos = pedidos.filter(fecha__date__lte=fecha_hasta)

    clientes = Cliente.objects.filter(activo=True).order_by("nombre")

    context = {
        "pedidos": pedidos,
        "clientes": clientes,
        "estado_actual": estado,
        "cliente_actual": cliente_id,
        "desde": fecha_desde,
        "hasta": fecha_hasta,
    }
    return render(request, "administrar/pedidos.html", context)


@login_required
def pedido_nuevo(request):
    return render(request, "administrar/pedido_nuevo.html", {})


# ===============================================================
#                     CRUD RUBROS (AJAX)
# ===============================================================
# -------- LISTA HTML --------
@login_required
def rubros_lista(request):
    rubros = Rubro.objects.all().order_by("nombre")
    return render(request, "administrar/rubros.html", {"rubros": rubros})


# -------- API: DETALLE --------
def api_rubros_detalle(request, id):
    r = get_object_or_404(Rubro, id=id)
    return JsonResponse({
        "id": r.id,
        "nombre": r.nombre,
        "descripcion": r.descripcion or ""
    })


# -------- API: GUARDAR (Nuevo / Editar) JSON --------
@require_POST
@csrf_protect
def api_rubros_guardar(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"error": "JSON inv谩lido."}, status=400)

    rubro_id = data.get("id")
    nombre = (data.get("nombre") or "").strip()
    descripcion = (data.get("descripcion") or "").strip()

    errors = {}

    if not nombre:
        errors["nombre"] = ["El nombre es obligatorio."]

    if nombre:
        qs = Rubro.objects.filter(nombre__iexact=nombre)
        if rubro_id:
            qs = qs.exclude(id=rubro_id)
        if qs.exists():
            errors["nombre"] = ["Ya existe un rubro con ese nombre."]

    if errors:
        return JsonResponse({"errors": errors}, status=400)

    if rubro_id:
        r = get_object_or_404(Rubro, id=rubro_id)
        r.nombre = nombre
        r.descripcion = descripcion
        r.save()
    else:
        r = Rubro.objects.create(nombre=nombre, descripcion=descripcion)

    return JsonResponse({"ok": True, "id": r.id})


# -------- API: ELIMINAR --------
@require_POST
@csrf_protect
def api_rubros_eliminar(request, id):
    r = get_object_or_404(Rubro, id=id)
    try:
        r.delete()
        return JsonResponse({"ok": True})
    except Exception:
        return JsonResponse({"error": "No se puede eliminar (puede estar en uso)."}, status=400)


# -------- API: VERIFICAR NOMBRE (Opcional) --------
def api_rubros_verificar_nombre(request):
    nombre = (request.GET.get("nombre") or "").strip()
    rubro_id = request.GET.get("id")

    existe = False
    if nombre:
        qs = Rubro.objects.filter(nombre__iexact=nombre)
        if rubro_id:
            qs = qs.exclude(id=rubro_id)
        existe = qs.exists()

    return JsonResponse({"existe": existe})



# ===============================================================
#                     CRUD UNIDADES (AJAX)
# ===============================================================

@login_required
def unidades_lista(request):
    return render(request, "administrar/unidades.html")


def api_unidades_listar(request):
    unidades = list(Unidad.objects.all().values("id", "nombre", "descripcion").order_by("nombre"))
    return JsonResponse({"data": unidades})


def api_unidades_detalle(request, id):
    u = get_object_or_404(Unidad, id=id)
    return JsonResponse({
        "id": u.id,
        "nombre": u.nombre,
        "descripcion": u.descripcion or ""
    })


@require_POST
@csrf_protect
def api_unidades_guardar(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"error": "JSON inv谩lido."}, status=400)

    unidad_id   = data.get("id")
    nombre      = (data.get("nombre") or "").strip()
    descripcion = (data.get("descripcion") or "").strip()

    errors = {}

    if not nombre:
        errors["nombre"] = ["El nombre es obligatorio."]

    if nombre:
        qs = Unidad.objects.filter(nombre__iexact=nombre)
        if unidad_id:
            qs = qs.exclude(id=unidad_id)
        if qs.exists():
            errors["nombre"] = ["Ya existe una unidad con ese nombre."]

    if errors:
        return JsonResponse({"errors": errors}, status=400)

    if unidad_id:
        u = get_object_or_404(Unidad, id=unidad_id)
        u.nombre = nombre
        u.descripcion = descripcion
        u.save()
    else:
        u = Unidad.objects.create(nombre=nombre, descripcion=descripcion)

    return JsonResponse({"ok": True, "id": u.id})


@require_POST
@csrf_protect
def api_unidades_eliminar(request, id):
    u = get_object_or_404(Unidad, id=id)
    try:
        u.delete()
        return JsonResponse({"ok": True})
    except Exception:
        return JsonResponse({"error": "No se puede eliminar (puede estar en uso)."}, status=400)


# ===============================================================
#                     CRUD LOCALIDADES (AJAX)
# ===============================================================

@login_required
def localidades_lista(request):
    return render(request, "administrar/localidades.html")


def api_localidades_listar(request):
    data = list(Localidad.objects.values("id", "nombre", "codigo_postal"))
    return JsonResponse({"data": data})


def api_localidades_detalle(request, id):
    l = get_object_or_404(Localidad, id=id)
    return JsonResponse({
        "id": l.id,
        "nombre": l.nombre,
        "codigo_postal": l.codigo_postal
    })


@require_POST
@csrf_protect
def api_localidades_guardar(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"error": "JSON inv谩lido."}, status=400)

    localidad_id = data.get("id")
    nombre = (data.get("nombre") or "").strip()
    codigo_postal = (data.get("codigo_postal") or "").strip()

    errors = {}

    if not nombre:
        errors["nombre"] = ["El nombre es obligatorio."]

    if nombre:
        qs = Localidad.objects.filter(nombre__iexact=nombre)
        if localidad_id:
            qs = qs.exclude(id=localidad_id)
        if qs.exists():
            errors["nombre"] = ["Ya existe una localidad con ese nombre."]



    if errors:
        return JsonResponse({"errors": errors}, status=400)

    if localidad_id:
        l = get_object_or_404(Localidad, id=localidad_id)
        l.nombre = nombre
        l.codigo_postal = codigo_postal
        l.save()
    else:
        l = Localidad.objects.create(
            nombre=nombre,
            codigo_postal=codigo_postal
        )

    return JsonResponse({"ok": True, "id": l.id})


@require_POST
@csrf_protect
def api_localidades_eliminar(request, id):
    l = get_object_or_404(Localidad, id=id)
    try:
        l.delete()
        return JsonResponse({"ok": True})
    except Exception:
        return JsonResponse({"error": "No se puede eliminar."}, status=400)



# ==========================================
# RUBROS
# ==========================================


@login_required
def rubros_lista(request):
    """Vista principal para la p谩gina de rubros"""
    return render(request, "administrar/rubros.html")


def api_rubros_listar(request):
    """API para listar todos los rubros"""
    rubros = Rubro.objects.all().order_by('nombre')
    data = []
    for r in rubros:
        data.append({
            'id': r.id,
            'nombre': r.nombre,
            'descripcion': r.descripcion or ''
        })
    return JsonResponse({'ok': True, 'data': data})


def api_rubros_detalle(request, id):
    """API para obtener detalles de un rubro"""
    rubro = get_object_or_404(Rubro, id=id)
    return JsonResponse({
        'id': rubro.id,
        'nombre': rubro.nombre,
        'descripcion': rubro.descripcion or ''
    })


@require_POST
@csrf_protect
def api_rubros_guardar(request):
    """API para crear o editar un rubro"""
    rubro_id = request.POST.get('id', '').strip()
    nombre = request.POST.get('nombre', '').strip()
    descripcion = request.POST.get('descripcion', '').strip()

    if not nombre:
        return JsonResponse({'ok': False, 'error': 'El nombre es obligatorio'}, status=400)

    try:
        if rubro_id:
            # Editar
            rubro = get_object_or_404(Rubro, id=rubro_id)
            rubro.nombre = nombre
            rubro.descripcion = descripcion
            rubro.save()
        else:
            # Crear nuevo
            rubro = Rubro.objects.create(
                nombre=nombre,
                descripcion=descripcion
            )
        
        return JsonResponse({'ok': True, 'id': rubro.id})
    
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@require_POST
@csrf_protect
def api_rubros_eliminar(request, id):
    """API para eliminar un rubro"""
    rubro = get_object_or_404(Rubro, id=id)
    try:
        rubro.delete()
        return JsonResponse({'ok': True, 'success': True})
    except Exception as e:
        return JsonResponse({'error': f'No se puede eliminar: {str(e)}'}, status=400)


def api_rubros_verificar_nombre(request):
    """API para verificar si un nombre de rubro ya existe"""
    nombre = request.GET.get('nombre', '').strip()
    id_actual = request.GET.get('id', None)
    
    if not nombre:
        return JsonResponse({'existe': False})
    
    qs = Rubro.objects.filter(nombre__iexact=nombre)
    
    if id_actual:
        qs = qs.exclude(id=id_actual)
    
    return JsonResponse({'existe': qs.exists()})


# =======================================
#  FACTURACIN A4 Y EDITOR DE PLANTILLAS
# =======================================

@login_required
def invoice_template_edit(request, template_id=None):
    if template_id:
        template = get_object_or_404(InvoiceTemplate, pk=template_id)
    else:
        # Si no se pasa ID, intentamos buscar la activa o la primera, o creamos una nueva
        template = InvoiceTemplate.objects.filter(active=True).first()
        if not template:
            template = InvoiceTemplate.objects.first()

    if request.method == 'POST':
        form = InvoiceTemplateForm(request.POST, request.FILES, instance=template)
        if form.is_valid():
            saved_template = form.save()
            messages.success(request, 'Plantilla guardada correctamente.')
            return redirect('invoice_template_edit_id', template_id=saved_template.id)
    else:
        form = InvoiceTemplateForm(instance=template)

    return render(request, 'administrar/invoice_template_edit.html', {
        'form': form,
        'template': template
    })

@login_required
def invoice_print(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)
    
    # Obtener el modelo solicitado
    model = request.GET.get('model', 'modern')
    # Lista de modelos v谩lidos (cat谩logo de 10)
    modelos_validos = ['modern', 'minimal', 'classic', 'elegant', 'tech', 'industrial', 'eco', 'compact', 'luxury', 'bold']
    if model not in modelos_validos:
        model = 'modern'
        
    # Construir el nombre de la plantilla
    template_name = f'administrar/comprobantes/inv_{model}.html'
    
    # Obtener la plantilla activa (fallback para datos extra si es necesario)
    template_config = InvoiceTemplate.objects.filter(active=True).first()
    if not template_config:
        template_config = InvoiceTemplate.objects.first()
    
    # Contexto para la factura
    context = {
        'venta': venta,
        'cliente': venta.cliente,
        'detalles': venta.detalles.all(),
        'template': template_config,
        'empresa': Empresa.objects.first(),
        'fecha_actual': venta.fecha,
    }
    
    try:
        return render(request, template_name, context)
    except:
        # Fallback a modern si la espec铆fica no existe
        return render(request, 'administrar/comprobantes/inv_modern.html', context)

from django.views.decorators.clickjacking import xframe_options_exempt

@xframe_options_exempt
@login_required
def invoice_print_preview(request):
    """Vista para previsualizar la plantilla con datos de ejemplo"""
    template_id = request.GET.get('template_id')
    if template_id:
        template = get_object_or_404(InvoiceTemplate, pk=template_id)
    else:
        template = InvoiceTemplate() # Plantilla vac铆a/default
        
        # Si vienen datos del formulario en POST (para previsualizaci贸n en tiempo real sin guardar)
        # Esto requerir铆a un manejo m谩s complejo con JS enviando el form al iframe.
        # Por ahora, asumimos que se guarda para ver cambios o se usa JS para inyectar CSS/HTML.
        # Una mejora ser铆a permitir POST a esta vista para renderizar sin guardar.

    # Datos de ejemplo
    class MockObj:
        pass
    
    cliente = MockObj()
    cliente.nombre = "Cliente Ejemplo S.A."
    cliente.cuit = "30-12345678-9"
    cliente.domicilio = "Av. Siempre Viva 123"
    cliente.localidad = MockObj()
    cliente.localidad.nombre = "Springfield"
    cliente.condicion_fiscal_display = "Responsable Inscripto"
    
    venta = MockObj()
    venta.id = 12345
    venta.fecha = datetime.date.today()
    venta.total = 15000.00
    venta.observaciones = "Entrega en horario comercial."
    
    detalles = []
    for i in range(1, 4):
        d = MockObj()
        d.producto = MockObj()
        d.producto.codigo = f"PROD-00{i}"
        d.producto.descripcion = f"Producto de Ejemplo {i}"
        d.cantidad = i * 2
        d.precio_unitario = 1000.00 * i
        d.subtotal = d.cantidad * d.precio_unitario
        detalles.append(d)

    empresa = Empresa.objects.first()
    if not empresa:
        empresa = MockObj()
        empresa.nombre = "Mi Empresa S.R.L."
        empresa.cuit = "30-00000000-1"
        empresa.direccion = "Calle Falsa 123"
        empresa.telefono = "555-1234"
        empresa.email = "info@miempresa.com"
        empresa.condicion_fiscal = "RI"

    context = {
        'venta': venta,
        'cliente': cliente,
        'detalles': detalles,
        'template': template,
        'empresa': empresa,
        'fecha_actual': datetime.date.today(),
        'is_preview': True
    }
    
    return render(request, 'administrar/invoice_a4.html', context)

# =======================================
#  FACTURACIN A4 Y EDITOR DE PLANTILLAS
# =======================================

PREDEFINED_TEMPLATES = [
    {
        'title': 'Plantilla Moderna',
        'description': 'Dise帽o actual con gradientes vibrantes y tipograf铆a moderna.',
        'css': '''
            :root {
                --primary-color: #667eea;
                --secondary-color: #764ba2;
            }
            .invoice-title h1 {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
                font-size: 32px;
            }
            header {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 30px;
                border-radius: 15px 15px 0 0;
            }
            .totals-table .total-row td {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
            }
        '''
    },
    {
        'title': 'Plantilla Cl谩sica',
        'description': 'Dise帽o profesional y tradicional con tipograf铆a serif.',
        'css': '''
            :root {
                --primary-color: #2c3e50;
                --secondary-color: #95a5a6;
            }
            .invoice-title h1 {
                font-family: Georgia, serif;
                color: #2c3e50;
                border-bottom: 3px solid #2c3e50;
                padding-bottom: 10px;
            }
            header {
                border-bottom: 3px solid #2c3e50;
            }
        '''
    },
    {
        'title': 'Plantilla Minimalista',
        'description': 'Dise帽o limpio y simple con mucho espacio en blanco.',
        'css': '''
            :root {
                --primary-color: #2d3748;
                --border-color: #e2e8f0;
            }
            body {
                font-weight: 300;
            }
            .invoice-title h1 {
                font-weight: 300;
                letter-spacing: 8px;
                color: #2d3748;
            }
            table {
                border: none;
            }
            th, td {
                border: none;
                border-bottom: 1px solid #e2e8f0;
            }
        '''
    },
    {
        'title': 'Plantilla Corporativa',
        'description': 'Dise帽o formal con azul corporativo.',
        'css': '''
            :root {
                --primary-color: #1e3a8a;
                --secondary-color: #3b82f6;
            }
            header {
                background: #1e3a8a;
                color: white;
                padding: 30px;
            }
            .invoice-title h1 {
                color: white;
            }
            .totals-table .total-row td {
                background: #1e3a8a;
                color: white;
            }
        '''
    },
    {
        'title': 'Plantilla Creativa',
        'description': 'Dise帽o vibrante y 煤nico con colores llamativos.',
        'css': '''
            :root {
                --primary-color: #f97316;
                --secondary-color: #84cc16;
            }
            .invoice-title h1 {
                background: linear-gradient(135deg, #f97316 0%, #84cc16 100%);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
                transform: rotate(-2deg);
                display: inline-block;
            }
            header {
                background: linear-gradient(135deg, #f97316 0%, #84cc16 100%);
                color: white;
                padding: 30px;
                border-radius: 20px;
            }
        '''
    },
    {
        'title': 'Plantilla Elegante',
        'description': 'Dise帽o sofisticado con detalles dorados.',
        'css': '''
            :root {
                --primary-color: #1a1a1a;
                --secondary-color: #d4af37;
            }
            header {
                background: #1a1a1a;
                color: #d4af37;
                padding: 30px;
                border: 2px solid #d4af37;
            }
            .invoice-title h1 {
                font-family: Georgia, serif;
                color: #d4af37;
                border: 2px solid #d4af37;
                padding: 20px;
                display: inline-block;
            }
            .totals-table .total-row td {
                background: #1a1a1a;
                color: #d4af37;
                border: 1px solid #d4af37;
            }
        '''
    }
]

@login_required
def invoice_templates_gallery(request):
    """Vista para mostrar la galer铆a de plantillas predefinidas"""
    
    # Procesar plantillas para la vista
    templates_view = []
    
    for pt in PREDEFINED_TEMPLATES:
        # Copia para no modificar la original
        tmpl_data = pt.copy()
        
        # Buscar si ya existe en DB
        db_tmpl = InvoiceTemplate.objects.filter(title=pt['title']).first()
        
        if db_tmpl:
            tmpl_data['id'] = db_tmpl.id
            tmpl_data['active'] = db_tmpl.active
            # Si existe, usamos el CSS de la DB por si fue editado? 
            # Por ahora mostramos el original como "preview" de lo que es, 
            # pero el ID permite activarla.
        else:
            # Si no existe, pasamos el t铆tulo como ID temporal para que el activate lo cree
            tmpl_data['id'] = pt['title'] # Usamos el t铆tulo como ID temporal
            tmpl_data['active'] = False
            
        templates_view.append(tmpl_data)

    return render(request, 'administrar/invoice_templates_gallery.html', {
        'templates': templates_view
    })

@csrf_exempt
def invoice_template_activate(request, template_name):
    """Vista para activar una plantilla. template_name puede ser un ID (int) o un T铆tulo (str)"""
    if request.method == 'POST':
        template = None
        
        # 1. Intentar buscar por ID num茅rico (si ya existe en DB)
        if str(template_name).isdigit():
            try:
                template = InvoiceTemplate.objects.get(pk=template_name)
            except InvoiceTemplate.DoesNotExist:
                pass
        
        # 2. Si no se encontr贸 por ID, buscar por t铆tulo en las predefinidas
        if not template:
            # Buscar en la lista de predefinidas
            predefined = next((t for t in PREDEFINED_TEMPLATES if t['title'] == template_name), None)
            
            if predefined:
                # Verificar si ya existe en DB por t铆tulo (por si acaso)
                template = InvoiceTemplate.objects.filter(title=template_name).first()
                
                if not template:
                    # CREARLA si no existe
                    template = InvoiceTemplate.objects.create(
                        title=predefined['title'],
                        css=predefined['css'],
                        active=False # Se activar谩 abajo
                    )
        
        if template:
            # Desactivar todas
            InvoiceTemplate.objects.all().update(active=False)
            # Activar la seleccionada
            template.active = True
            template.save()
            return JsonResponse({'success': True, 'message': 'Plantilla activada correctamente'})
        else:
            return JsonResponse({'success': False, 'message': 'Plantilla no encontrada'}, status=404)
    
    return JsonResponse({'success': False, 'message': 'M茅todo no permitido'}, status=405)


# =======================================
# Ь PEDIDOS - API MODERNA
# =======================================

@login_required
def pedidos(request):
    """Vista principal de pedidos con interfaz moderna"""
    return render(request, 'administrar/pedidos.html')


def api_pedidos_lista(request):
    """API para listar pedidos con paginaci贸n y filtros"""
    # Par谩metros de paginaci贸n
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    
    # Filtros
    busqueda = request.GET.get('busqueda', '').strip()
    estado = request.GET.get('estado', '').strip()
    cliente_id = request.GET.get('cliente', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    
    # Query base
    pedidos = Pedido.objects.select_related('cliente', 'venta').prefetch_related('detalles__producto').all()
    
    # Aplicar filtros
    if busqueda:
        pedidos = pedidos.filter(
            Q(id__icontains=busqueda) |
            Q(cliente__nombre__icontains=busqueda) |
            Q(observaciones__icontains=busqueda)
        )
    
    if estado:
        # Manejar m煤ltiples estados separados por comas (ej: "PENDIENTE,PREPARACION,LISTO")
        if ',' in estado:
            estados_lista = [e.strip() for e in estado.split(',')]
            pedidos = pedidos.filter(estado__in=estados_lista)
        else:
            pedidos = pedidos.filter(estado=estado)
    
    if cliente_id:
        pedidos = pedidos.filter(cliente_id=cliente_id)
    
    if fecha_desde:
        pedidos = pedidos.filter(fecha__date__gte=fecha_desde)
    
    if fecha_hasta:
        pedidos = pedidos.filter(fecha__date__lte=fecha_hasta)
    
    # Ordenar
    pedidos = pedidos.order_by('-fecha')
    
    # Contar total
    total = pedidos.count()
    
    # Paginaci贸n
    start = (page - 1) * per_page
    end = start + per_page
    pedidos_page = pedidos[start:end]
    
    # Serializar
    data = []
    for p in pedidos_page:
        # Contar items
        num_items = p.detalles.count()
        
        data.append({
            'id': p.id,
            'fecha': p.fecha.strftime('%d/%m/%Y %H:%M'),
            'cliente_id': p.cliente.id,
            'cliente_nombre': p.cliente.nombre,
            'total': float(p.total),
            'estado': p.estado,
            'estado_display': p.get_estado_display(),
            'observaciones': p.observaciones or '',
            'venta_id': p.venta_id,
            'num_items': num_items,
        })
    
    return JsonResponse({
        'pedidos': data,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page
    })


def api_pedido_detalle(request, id):
    """API para obtener detalle de un pedido"""
    try:
        pedido = Pedido.objects.select_related('cliente', 'venta').prefetch_related('detalles__producto').get(id=id)
    except Pedido.DoesNotExist:
        return JsonResponse({'error': 'Pedido no encontrado'}, status=404)
    
    # Serializar detalles
    detalles = []
    for d in pedido.detalles.all():
        detalles.append({
            'id': d.id,
            'producto_id': d.producto.id,
            'producto_codigo': d.producto.codigo,
            'producto_descripcion': d.producto.descripcion,
            'cantidad': float(d.cantidad),
            'precio_unitario': float(d.precio_unitario),
            'subtotal': float(d.subtotal),
        })
    
    return JsonResponse({
        'id': pedido.id,
        'fecha': pedido.fecha.strftime('%d/%m/%Y %H:%M'),
        'cliente_id': pedido.cliente.id,
        'cliente_nombre': pedido.cliente.nombre,
        'cliente_cuit': pedido.cliente.cuit or '',
        'cliente_telefono': pedido.cliente.telefono or '',
        'cliente_email': pedido.cliente.email or '',
        'estado': pedido.estado,
        'estado_display': pedido.get_estado_display(),
        'total': float(pedido.total),
        'observaciones': pedido.observaciones or '',
        'venta_id': pedido.venta_id,
        'detalles': detalles,
    })

@login_required
def pedido_print(request, pedido_id):
    """Vista para imprimir un pedido (Presupuesto)"""
    from .models import Pedido, Empresa
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    
    model = request.GET.get('model', 'modern')
    template_name = f'administrar/comprobantes/ped_{model}.html'
    
    context = {
        'pedido': pedido,
        'empresa': Empresa.objects.first(),
    }
    
    try:
        return render(request, template_name, context)
    except:
        return render(request, 'administrar/comprobantes/ped_modern.html', context)


@csrf_exempt
@require_POST
def api_pedido_crear(request):
    """API para crear un nuevo pedido"""
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inv谩lido'}, status=400)
    
    # Validar datos
    cliente_id = data.get('cliente_id')
    observaciones = data.get('observaciones', '')
    detalles = data.get('detalles', [])
    
    if not cliente_id:
        return JsonResponse({'error': 'Cliente es requerido'}, status=400)
    
    if not detalles or len(detalles) == 0:
        return JsonResponse({'error': 'Debe agregar al menos un producto'}, status=400)
    
    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)
    
    # Crear pedido
    try:
        with transaction.atomic():
            pedido = Pedido.objects.create(
                cliente=cliente,
                observaciones=observaciones,
                total=0
            )
            
            total = Decimal('0.00')
            
            # Crear detalles
            for det in detalles:
                producto_id = det.get('producto_id')
                cantidad = Decimal(str(det.get('cantidad', 0)))
                precio_unitario = Decimal(str(det.get('precio_unitario', 0)))
                
                if not producto_id or cantidad <= 0:
                    continue
                
                try:
                    producto = Producto.objects.get(id=producto_id)
                except Producto.DoesNotExist:
                    continue
                
                subtotal = cantidad * precio_unitario
                
                DetallePedido.objects.create(
                    pedido=pedido,
                    producto=producto,
                    cantidad=cantidad,
                    precio_unitario=precio_unitario,
                    subtotal=subtotal
                )
                
                total += subtotal
            
            # Actualizar total
            pedido.total = total
            pedido.save()
            
            return JsonResponse({
                'ok': True,
                'id': pedido.id,
                'message': 'Pedido creado exitosamente'
            })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
def api_pedido_editar(request, id):
    """API para editar un pedido existente"""
    try:
        pedido = Pedido.objects.get(id=id)
    except Pedido.DoesNotExist:
        return JsonResponse({'error': 'Pedido no encontrado'}, status=404)
    
    # No permitir editar pedidos facturados
    if pedido.venta_id:
        return JsonResponse({'error': 'No se puede editar un pedido ya facturado'}, status=400)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inv谩lido'}, status=400)
    
    # Validar datos
    cliente_id = data.get('cliente_id')
    observaciones = data.get('observaciones', '')
    detalles = data.get('detalles', [])
    
    if not cliente_id:
        return JsonResponse({'error': 'Cliente es requerido'}, status=400)
    
    if not detalles or len(detalles) == 0:
        return JsonResponse({'error': 'Debe agregar al menos un producto'}, status=400)
    
    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)
    
    # Actualizar pedido
    try:
        with transaction.atomic():
            pedido.cliente = cliente
            pedido.observaciones = observaciones
            
            # Eliminar detalles anteriores
            pedido.detalles.all().delete()
            
            total = Decimal('0.00')
            
            # Crear nuevos detalles
            for det in detalles:
                producto_id = det.get('producto_id')
                cantidad = Decimal(str(det.get('cantidad', 0)))
                precio_unitario = Decimal(str(det.get('precio_unitario', 0)))
                
                if not producto_id or cantidad <= 0:
                    continue
                
                try:
                    producto = Producto.objects.get(id=producto_id)
                except Producto.DoesNotExist:
                    continue
                
                subtotal = cantidad * precio_unitario
                
                DetallePedido.objects.create(
                    pedido=pedido,
                    producto=producto,
                    cantidad=cantidad,
                    precio_unitario=precio_unitario,
                    subtotal=subtotal
                )
                
                total += subtotal
            
            # Actualizar total
            pedido.total = total
            pedido.save()
            
            return JsonResponse({
                'ok': True,
                'message': 'Pedido actualizado exitosamente'
            })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
def api_pedido_eliminar(request, id):
    """API para eliminar un pedido"""
    try:
        pedido = Pedido.objects.get(id=id)
    except Pedido.DoesNotExist:
        return JsonResponse({'error': 'Pedido no encontrado'}, status=404)
    
    # No permitir eliminar pedidos facturados
    if pedido.venta_id:
        return JsonResponse({'error': 'No se puede eliminar un pedido ya facturado'}, status=400)
    
    try:
        pedido.delete()
        return JsonResponse({'ok': True, 'message': 'Pedido eliminado exitosamente'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
def api_pedido_cambiar_estado(request, id):
    """API para cambiar el estado de un pedido"""
    try:
        pedido = Pedido.objects.get(id=id)
    except Pedido.DoesNotExist:
        return JsonResponse({'error': 'Pedido no encontrado'}, status=404)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inv谩lido'}, status=400)
    
    nuevo_estado = data.get('estado', '').strip()
    
    # Validar estado
    estados_validos = [e[0] for e in Pedido.ESTADO_PEDIDO]
    if nuevo_estado not in estados_validos:
        return JsonResponse({'error': 'Estado inv谩lido'}, status=400)

    # Si el nuevo estado es FACTURADO y no tiene venta, realizar proceso de facturaci贸n
    if nuevo_estado == 'FACTURADO' and not pedido.venta_id:
        try:
            with transaction.atomic():
                # Determinar tipo de comprobante
                tipo_cbte = 'A' if pedido.cliente.condicion_fiscal == 'RI' else 'B'

                # Crear venta
                venta = Venta.objects.create(
                    cliente=pedido.cliente,
                    total=pedido.total,
                    estado='Emitida',
                    tipo_comprobante=tipo_cbte,
                    fecha=datetime.datetime.now()
                )
                
                # Crear detalles y descontar stock
                detalles_pedido = DetallePedido.objects.filter(pedido=pedido)
                for det in detalles_pedido:
                    DetalleVenta.objects.create(
                        venta=venta,
                        producto=det.producto,
                        cantidad=det.cantidad,
                        precio_unitario=det.precio_unitario,
                        subtotal=det.subtotal
                    )
                    
                    # Descontar stock
                    producto = det.producto
                    producto.stock -= det.cantidad
                    producto.save()
                
                # Vincular pedido con venta
                pedido.venta = venta
                pedido.estado = 'FACTURADO'
                pedido.save()
                
                return JsonResponse({
                    'ok': True, 
                    'message': f'Pedido facturado correctamente. Venta #{venta.id}',
                    'venta_id': venta.id
                })
        except Exception as e:
            return JsonResponse({'error': f'Error al facturar: {str(e)}'}, status=500)
    
    pedido.estado = nuevo_estado
    pedido.save()
    
    return JsonResponse({
        'ok': True,
        'message': 'Estado actualizado exitosamente',
        'estado': pedido.estado,
        'estado_display': pedido.get_estado_display()
    })

# FACTURAR PEDIDO (CONVERTIR A VENTA)
# =======================================================
@csrf_exempt
@require_POST
def api_pedido_facturar(request, id):
    try:
        with transaction.atomic():
            pedido = Pedido.objects.select_related('cliente').get(pk=id)
            
            # Validar que no est茅 ya facturado
            if pedido.venta_id:
                return JsonResponse({'ok': False, 'error': 'El pedido ya fue facturado'})
            
            # Determinar tipo de comprobante
            tipo_cbte = 'A' if pedido.cliente.condicion_fiscal == 'RI' else 'B'

            # Crear venta
            # Asumimos medio de pago CTACTE por defecto o lo que corresponda
            # Podr铆amos recibirlo por par谩metro si fuera necesario
            venta = Venta.objects.create(
                cliente=pedido.cliente,
                total=pedido.total,
                estado='Emitida',
                tipo_comprobante=tipo_cbte, 
                fecha=datetime.datetime.now(),
                pedido_origen=pedido  # Vincular con el pedido que gener贸 esta venta
            )
            
            # Crear detalles y descontar stock
            detalles_pedido = DetallePedido.objects.filter(pedido=pedido)
            for det in detalles_pedido:
                DetalleVenta.objects.create(
                    venta=venta,
                    producto=det.producto,
                    cantidad=det.cantidad,
                    precio_unitario=det.precio_unitario,
                    subtotal=det.subtotal
                )
                
                # Descontar stock
                producto = det.producto
                producto.stock -= det.cantidad
                producto.save()
            
            # Vincular pedido con venta
            pedido.venta = venta
            pedido.estado = 'FACTURADO'
            pedido.save()
            
            # ==========================================
            # GENERACIN AUTOMTICA DE ASIENTO CONTABLE
            # ==========================================
            try:
                from administrar.models import EjercicioContable, PlanCuenta, Asiento, ItemAsiento
                from django.db.models import Max
                
                # Buscar ejercicio vigente
                ejercicio = EjercicioContable.objects.filter(
                    cerrado=False,
                    fecha_inicio__lte=venta.fecha,
                    fecha_fin__gte=venta.fecha
                ).first()
                
                # Si no encuentra por fecha exacta, buscar el 煤ltimo abierto (fallback para pruebas)
                if not ejercicio:
                    ejercicio = EjercicioContable.objects.filter(cerrado=False).last()

                if ejercicio:
                    # Buscar cuentas (Hardcodeadas seg煤n script de generaci贸n)
                    cta_ventas = PlanCuenta.objects.filter(codigo='4.1.01').first() # Ventas (Corregido a Nivel 3)
                    cta_deudores = PlanCuenta.objects.filter(codigo='1.1.02.001').first() # Deudores por Ventas
                    cta_iva = PlanCuenta.objects.filter(codigo='2.1.02.001').first() # IVA D茅bito Fiscal
                    
                    if cta_ventas and cta_deudores:
                        # Calcular IVA simplificado (asumiendo que el total incluye IVA)
                        monto_iva = Decimal('0')
                        monto_neto = venta.total
                        
                        if venta.tipo_comprobante == 'A':
                            monto_neto = venta.total / Decimal('1.21')
                            monto_iva = venta.total - monto_neto
                        
                        # Obtener correlativo
                        numero = (Asiento.objects.filter(ejercicio=ejercicio).aggregate(m=Max('numero'))['m'] or 0) + 1
                        
                        asiento = Asiento.objects.create(
                            numero=numero,
                            fecha=venta.fecha,
                            descripcion=f"Venta Fac-{venta.tipo_comprobante} {venta.id} - {pedido.cliente.nombre}",
                            ejercicio=ejercicio,
                            origen='VENTAS'
                        )
                        
                        # DEBE: Deudores por Venta (Total)
                        ItemAsiento.objects.create(
                            asiento=asiento,
                            cuenta=cta_deudores,
                            debe=venta.total,
                            haber=0,
                            descripcion=f"Factura {venta.tipo_comprobante}-{venta.id}"
                        )
                        
                        # HABER: Ventas (Neto)
                        ItemAsiento.objects.create(
                            asiento=asiento,
                            cuenta=cta_ventas,
                            debe=0,
                            haber=monto_neto,
                            descripcion=f"Venta Art铆culos"
                        )
                        
                        # HABER: IVA (si aplica y > 0)
                        if monto_iva > 0 and cta_iva:
                            ItemAsiento.objects.create(
                                asiento=asiento,
                                cuenta=cta_iva,
                                debe=0,
                                haber=monto_iva,
                                descripcion=f"IVA D茅bito Fiscal"
                            )
                            
            except Exception as e:
                print(f"Error generando asiento contable para venta {venta.id}: {str(e)}")
            
            return JsonResponse({
                'ok': True,
                'message': f'Pedido facturado correctamente. Venta #{venta.id}',
                'venta_id': venta.id
            })

    except Pedido.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Pedido no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


# =======================================
#  API EJERCICIOS CONTABLES
# =======================================

def api_ejercicios_listar(request):
    """API para listar ejercicios contables con estad铆sticas"""
    from administrar.models import EjercicioContable, Asiento
    from django.db.models import Count, Max

    try:
        # Annotate with stats
        ejercicios = EjercicioContable.objects.annotate(
            cantidad_asientos=Count('asiento'),
            ultimo_movimiento=Max('asiento__fecha')
        ).order_by('-fecha_inicio')

        data = []
        for e in ejercicios:
            data.append({
                'id': e.id,
                'descripcion': e.descripcion,
                'fecha_inicio': str(e.fecha_inicio) if e.fecha_inicio else None,
                'fecha_fin': str(e.fecha_fin) if e.fecha_fin else None,
                'cerrado': e.cerrado,
                'cantidad_asientos': e.cantidad_asientos,
                'ultimo_movimiento': e.ultimo_movimiento.strftime('%d/%m/%Y') if e.ultimo_movimiento else 'Sin movimientos'
            })
            
        return JsonResponse({
            'success': True,
            'ejercicios': data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_POST
def api_ejercicios_crear(request):
    """API para crear un nuevo ejercicio contable"""
    from administrar.models import EjercicioContable
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inv谩lido'}, status=400)
    
    descripcion = data.get('descripcion', '').strip()
    fecha_inicio = data.get('fecha_inicio', '')
    fecha_fin = data.get('fecha_fin', '')
    
    if not descripcion:
        return JsonResponse({'error': 'La descripci贸n es requerida'}, status=400)
    
    try:
        ejercicio = EjercicioContable.objects.create(
            descripcion=descripcion,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin
        )
        
        return JsonResponse({
            'ok': True,
            'ejercicio': {
                'id': ejercicio.id,
                'descripcion': ejercicio.descripcion,
                'fecha_inicio': str(ejercicio.fecha_inicio) if ejercicio.fecha_inicio else None,
                'fecha_fin': str(ejercicio.fecha_fin) if ejercicio.fecha_fin else None,
                'cerrado': ejercicio.cerrado
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
def api_ejercicios_editar(request, id):
    """API para editar un ejercicio contable"""
    from administrar.models import EjercicioContable
    
    try:
        ejercicio = EjercicioContable.objects.get(id=id)
    except EjercicioContable.DoesNotExist:
        return JsonResponse({'error': 'Ejercicio no encontrado'}, status=404)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inv谩lido'}, status=400)
    
    descripcion = data.get('descripcion', '').strip()
    fecha_inicio = data.get('fecha_inicio', '')
    fecha_fin = data.get('fecha_fin', '')
    cerrado = data.get('cerrado', False)
    
    if not descripcion:
        return JsonResponse({'error': 'La descripci贸n es requerida'}, status=400)
    
    try:
        ejercicio.descripcion = descripcion
        ejercicio.fecha_inicio = fecha_inicio
        ejercicio.fecha_fin = fecha_fin
        ejercicio.cerrado = cerrado
        ejercicio.save()
        
        return JsonResponse({
            'ok': True,
            'ejercicio': {
                'id': ejercicio.id,
                'descripcion': ejercicio.descripcion,
                'fecha_inicio': str(ejercicio.fecha_inicio) if ejercicio.fecha_inicio else None,
                'fecha_fin': str(ejercicio.fecha_fin) if ejercicio.fecha_fin else None,
                'cerrado': ejercicio.cerrado
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
def api_ejercicios_eliminar(request, id):
    """API para eliminar un ejercicio contable"""
    from administrar.models import EjercicioContable
    
    try:
        ejercicio = EjercicioContable.objects.get(id=id)
    except EjercicioContable.DoesNotExist:
        return JsonResponse({'error': 'Ejercicio no encontrado'}, status=404)
    
    # Verificar si tiene asientos asociados
    if hasattr(ejercicio, 'asiento_set') and ejercicio.asiento_set.exists():
        return JsonResponse({
            'error': 'No se puede eliminar un ejercicio que tiene asientos contables asociados'
        }, status=400)
    
    try:
        ejercicio.delete()
        return JsonResponse({
            'ok': True,
            'mensaje': 'Ejercicio eliminado correctamente'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ==========================================
# API ASIENTOS CONTABLES
# ==========================================

@require_http_methods(["GET"])
def api_asientos_listar(request):
    """API para listar todos los asientos contables"""
    from administrar.models import Asiento, ItemAsiento
    
    try:
        asientos = Asiento.objects.all().order_by('-fecha', '-numero')
        data = []
        
        for asiento in asientos:
            # Calcular totales
            items = ItemAsiento.objects.filter(asiento=asiento)
            total_debe = sum(item.debe for item in items)
            total_haber = sum(item.haber for item in items)
            
            data.append({
                'id': asiento.id,
                'numero': asiento.numero,
                'fecha': str(asiento.fecha) if asiento.fecha else None,
                'descripcion': asiento.descripcion,
                'ejercicio_id': asiento.ejercicio_id,
                'origen': asiento.origen,
                'total_debe': float(total_debe),
                'total_haber': float(total_haber)
            })
        
        return JsonResponse({
            'success': True,
            'asientos': data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def api_asientos_crear(request):
    """API para crear un nuevo asiento contable"""
    from administrar.models import Asiento, ItemAsiento, EjercicioContable
    import json
    from decimal import Decimal
    
    try:
        data = json.loads(request.body)
        
        # Validar ejercicio
        try:
            ejercicio = EjercicioContable.objects.get(id=data['ejercicio_id'])
            if ejercicio.cerrado:
                return JsonResponse({
                    'ok': False,
                    'error': 'No se pueden crear asientos en un ejercicio cerrado'
                }, status=400)
        except EjercicioContable.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Ejercicio no encontrado'}, status=404)
        
        # Validar que el asiento est茅 balanceado
        movimientos = data.get('movimientos', [])
        if len(movimientos) < 2:
            return JsonResponse({
                'ok': False,
                'error': 'Debe haber al menos 2 movimientos'
            }, status=400)
        
        total_debe = sum(Decimal(str(m.get('debe', 0))) for m in movimientos)
        total_haber = sum(Decimal(str(m.get('haber', 0))) for m in movimientos)
        
        if abs(total_debe - total_haber) > Decimal('0.01'):
            return JsonResponse({
                'ok': False,
                'error': f'El asiento est谩 descuadrado. Debe: {total_debe}, Haber: {total_haber}'
            }, status=400)
        
        # Crear asiento
        asiento = Asiento.objects.create(
            numero=data['numero'],
            fecha=data['fecha'],
            descripcion=data['descripcion'],
            ejercicio_id=data['ejercicio_id'],
            tipo=data.get('tipo', 'MANUAL')
        )
        
        # Crear items
        for mov in movimientos:
            if mov.get('cuenta_id') and (mov.get('debe', 0) > 0 or mov.get('haber', 0) > 0):
                ItemAsiento.objects.create(
                    asiento=asiento,
                    cuenta_id=mov['cuenta_id'],
                    debe=Decimal(str(mov.get('debe', 0))),
                    haber=Decimal(str(mov.get('haber', 0)))
                )
        
        return JsonResponse({
            'ok': True,
            'asiento': {
                'id': asiento.id,
                'numero': asiento.numero,
                'fecha': str(asiento.fecha),
                'descripcion': asiento.descripcion,
                'ejercicio_id': asiento.ejercicio_id,
                'tipo': asiento.tipo
            }
        })
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def api_asientos_detalle(request, id):
    """API para obtener el detalle de un asiento"""
    from administrar.models import Asiento, ItemAsiento
    
    try:
        asiento = Asiento.objects.get(id=id)
        items = ItemAsiento.objects.filter(asiento=asiento).select_related('cuenta')
        
        movimientos = []
        for item in items:
            movimientos.append({
                'id': item.id,
                'cuenta_id': item.cuenta_id,
                'cuenta_codigo': item.cuenta.codigo,
                'cuenta_nombre': item.cuenta.nombre,
                'debe': float(item.debe),
                'haber': float(item.haber)
            })
        
        return JsonResponse({
            'ok': True,
            'asiento': {
                'id': asiento.id,
                'numero': asiento.numero,
                'fecha': str(asiento.fecha),
                'descripcion': asiento.descripcion,
                'ejercicio_id': asiento.ejercicio_id,
                'origen': asiento.origen,
                'movimientos': movimientos
            }
        })
    except Asiento.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Asiento no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def api_asientos_eliminar(request, id):
    """API para eliminar un asiento contable"""
    from administrar.models import Asiento, EjercicioContable
    
    try:
        asiento = Asiento.objects.get(id=id)
        
        # Verificar que el ejercicio no est茅 cerrado
        if asiento.ejercicio.cerrado:
            return JsonResponse({
                'ok': False,
                'error': 'No se pueden eliminar asientos de un ejercicio cerrado'
            }, status=400)
        
        asiento.delete()
        return JsonResponse({
            'ok': True,
            'mensaje': 'Asiento eliminado correctamente'
        })
        
    except Asiento.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Asiento no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


# ==========================================
# API BALANCE DE SUMAS Y SALDOS
# ==========================================

@require_http_methods(["GET"])
def api_balance_generar(request):
    """API para generar el balance de sumas y saldos"""
    from administrar.models import PlanCuenta, Asiento, ItemAsiento, EjercicioContable
    from decimal import Decimal
    from django.db.models import Sum, Q
    
    try:
        ejercicio_id = request.GET.get('ejercicio_id')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        nivel_filtro = request.GET.get('nivel')
        solo_con_movimientos = request.GET.get('solo_con_movimientos', 'true').lower() == 'true'
        
        if not ejercicio_id:
            return JsonResponse({'success': False, 'error': 'Debe especificar un ejercicio'}, status=400)
        
        # Obtener ejercicio
        try:
            ejercicio = EjercicioContable.objects.get(id=ejercicio_id)
        except EjercicioContable.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Ejercicio no encontrado'}, status=404)
        
        # Establecer fechas por defecto del ejercicio si no se especifican
        if not fecha_desde:
            fecha_desde = ejercicio.fecha_inicio
        if not fecha_hasta:
            fecha_hasta = ejercicio.fecha_fin
        
        # Obtener todas las cuentas
        cuentas_query = PlanCuenta.objects.all().order_by('codigo')
        
        # Filtrar por nivel si se especifica
        if nivel_filtro:
            cuentas_query = cuentas_query.filter(nivel__lte=int(nivel_filtro))
        
        cuentas_data = []
        total_debe = Decimal('0')
        total_haber = Decimal('0')
        total_saldo_deudor = Decimal('0')
        total_saldo_acreedor = Decimal('0')
        
        for cuenta in cuentas_query:
            # Obtener movimientos de la cuenta en el per铆odo
            items = ItemAsiento.objects.filter(
                cuenta=cuenta,
                asiento__ejercicio_id=ejercicio_id,
                asiento__fecha__gte=fecha_desde,
                asiento__fecha__lte=fecha_hasta
            )
            
            # Calcular totales
            debe = items.aggregate(total=Sum('debe'))['total'] or Decimal('0')
            haber = items.aggregate(total=Sum('haber'))['total'] or Decimal('0')
            saldo = debe - haber
            
            # Si solo queremos cuentas con movimientos, saltar las que no tienen
            if solo_con_movimientos and debe == 0 and haber == 0:
                continue
            
            cuentas_data.append({
                'id': cuenta.id,
                'codigo': cuenta.codigo,
                'nombre': cuenta.nombre,
                'nivel': cuenta.nivel,
                'tipo': cuenta.tipo,
                'debe': float(debe),
                'haber': float(haber),
                'saldo': float(saldo)
            })
            
            # Acumular totales
            total_debe += debe
            total_haber += haber
            if saldo > 0:
                total_saldo_deudor += saldo
            else:
                total_saldo_acreedor += abs(saldo)
        
        return JsonResponse({
            'success': True,
            'cuentas': cuentas_data,
            'totales': {
                'total_debe': float(total_debe),
                'total_haber': float(total_haber),
                'total_saldo_deudor': float(total_saldo_deudor),
                'total_saldo_acreedor': float(total_saldo_acreedor)
            },
            'ejercicio': {
                'id': ejercicio.id,
                'descripcion': ejercicio.descripcion,
                'fecha_desde': str(fecha_desde),
                'fecha_hasta': str(fecha_hasta)
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def api_balance_exportar(request):
    """API para exportar el balance a Excel"""
    from administrar.models import PlanCuenta, Asiento, ItemAsiento, EjercicioContable
    from decimal import Decimal
    from django.db.models import Sum
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    try:
        ejercicio_id = request.GET.get('ejercicio_id')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        nivel_filtro = request.GET.get('nivel')
        solo_con_movimientos = request.GET.get('solo_con_movimientos', 'true').lower() == 'true'
        
        if not ejercicio_id:
            return JsonResponse({'error': 'Debe especificar un ejercicio'}, status=400)
        
        ejercicio = EjercicioContable.objects.get(id=ejercicio_id)
        
        if not fecha_desde:
            fecha_desde = ejercicio.fecha_inicio
        if not fecha_hasta:
            fecha_hasta = ejercicio.fecha_fin
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance de Sumas y Saldos"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # T铆tulo
        ws.merge_cells('A1:F1')
        ws['A1'] = 'BALANCE DE SUMAS Y SALDOS'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Informaci贸n del ejercicio
        ws['A2'] = f'Ejercicio: {ejercicio.descripcion}'
        ws['A3'] = f'Per铆odo: {fecha_desde} al {fecha_hasta}'
        
        # Encabezados
        headers = ['C贸digo', 'Cuenta', 'Debe', 'Haber', 'Saldo Deudor', 'Saldo Acreedor']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        cuentas_query = PlanCuenta.objects.all().order_by('codigo')
        if nivel_filtro:
            cuentas_query = cuentas_query.filter(nivel__lte=int(nivel_filtro))
        
        row = 6
        total_debe = Decimal('0')
        total_haber = Decimal('0')
        total_saldo_deudor = Decimal('0')
        total_saldo_acreedor = Decimal('0')
        
        for cuenta in cuentas_query:
            items = ItemAsiento.objects.filter(
                cuenta=cuenta,
                asiento__ejercicio_id=ejercicio_id,
                asiento__fecha__gte=fecha_desde,
                asiento__fecha__lte=fecha_hasta
            )
            
            debe = items.aggregate(total=Sum('debe'))['total'] or Decimal('0')
            haber = items.aggregate(total=Sum('haber'))['total'] or Decimal('0')
            saldo = debe - haber
            
            if solo_con_movimientos and debe == 0 and haber == 0:
                continue
            
            ws.cell(row=row, column=1).value = cuenta.codigo
            ws.cell(row=row, column=2).value = cuenta.nombre
            ws.cell(row=row, column=3).value = float(debe)
            ws.cell(row=row, column=4).value = float(haber)
            ws.cell(row=row, column=5).value = float(saldo) if saldo > 0 else 0
            ws.cell(row=row, column=6).value = float(abs(saldo)) if saldo < 0 else 0
            
            # Aplicar formato num茅rico
            for col in range(3, 7):
                ws.cell(row=row, column=col).number_format = '#,##0.00'
            
            total_debe += debe
            total_haber += haber
            if saldo > 0:
                total_saldo_deudor += saldo
            else:
                total_saldo_acreedor += abs(saldo)
            
            row += 1
        
        # Totales
        ws.cell(row=row, column=1).value = 'TOTALES'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3).value = float(total_debe)
        ws.cell(row=row, column=4).value = float(total_haber)
        ws.cell(row=row, column=5).value = float(total_saldo_deudor)
        ws.cell(row=row, column=6).value = float(total_saldo_acreedor)
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).border = border
            if col >= 3:
                ws.cell(row=row, column=col).number_format = '#,##0.00'
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="balance_{ejercicio.descripcion.replace(" ", "_")}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ==========================================
# API LIBRO MAYOR
# ==========================================

@require_http_methods(["GET"])
def api_mayor_consultar(request):
    """API para consultar el libro mayor de una cuenta"""
    from administrar.models import PlanCuenta, ItemAsiento, Asiento
    from decimal import Decimal
    from django.db.models import Q
    
    try:
        cuenta_id = request.GET.get('cuenta_id')
        ejercicio_id = request.GET.get('ejercicio_id')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        
        if not cuenta_id:
            return JsonResponse({'success': False, 'error': 'Debe especificar una cuenta'}, status=400)
        
        # Obtener cuenta
        try:
            cuenta = PlanCuenta.objects.get(id=cuenta_id)
        except PlanCuenta.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Cuenta no encontrada'}, status=404)
        
        # Construir lista de cuentas a consultar (incluyendo hijos si es agrupadora)
        cuentas_ids = [cuenta.id]
        if not cuenta.imputable:
            # Funci贸n auxiliar para obtener descendientes
            def obtener_descendientes(cuenta_padre):
                descendientes = []
                for hijo in cuenta_padre.hijos.all():
                    descendientes.append(hijo.id)
                    descendientes.extend(obtener_descendientes(hijo))
                return descendientes
            
            cuentas_ids.extend(obtener_descendientes(cuenta))
        
        # Construir query de movimientos
        query = Q(cuenta_id__in=cuentas_ids)
        
        if ejercicio_id:
            query &= Q(asiento__ejercicio_id=ejercicio_id)
        if fecha_desde:
            query &= Q(asiento__fecha__gte=fecha_desde)
        if fecha_hasta:
            query &= Q(asiento__fecha__lte=fecha_hasta)
        
        # Obtener movimientos ordenados por fecha
        items = ItemAsiento.objects.filter(query).select_related('asiento').order_by('asiento__fecha', 'asiento__numero')
        
        # Calcular saldo inicial (movimientos anteriores a fecha_desde)
        saldo_inicial = Decimal('0')
        if fecha_desde:
            items_anteriores = ItemAsiento.objects.filter(
                cuenta_id__in=cuentas_ids,
                asiento__fecha__lt=fecha_desde
            )
            if ejercicio_id:
                items_anteriores = items_anteriores.filter(asiento__ejercicio_id=ejercicio_id)
            
            for item in items_anteriores:
                saldo_inicial += item.debe - item.haber
        
        # Procesar movimientos
        movimientos = []
        saldo_acumulado = saldo_inicial
        total_debe = Decimal('0')
        total_haber = Decimal('0')
        
        for item in items:
            saldo_acumulado += item.debe - item.haber
            total_debe += item.debe
            total_haber += item.haber
            
            movimientos.append({
                'id': item.id,
                'fecha': str(item.asiento.fecha),
                'asiento_id': item.asiento.id,
                'asiento_numero': item.asiento.numero,
                'descripcion': item.asiento.descripcion,
                'debe': float(item.debe),
                'haber': float(item.haber),
                'saldo': float(saldo_acumulado)
            })
        
        return JsonResponse({
            'success': True,
            'cuenta': {
                'id': cuenta.id,
                'codigo': cuenta.codigo,
                'nombre': cuenta.nombre,
                'tipo': cuenta.tipo,
                'nivel': cuenta.nivel
            },
            'resumen': {
                'saldo_inicial': float(saldo_inicial),
                'total_debe': float(total_debe),
                'total_haber': float(total_haber),
                'saldo_final': float(saldo_acumulado)
            },
            'movimientos': movimientos
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def api_mayor_exportar(request):
    """API para exportar el libro mayor a Excel"""
    from administrar.models import PlanCuenta, ItemAsiento
    from decimal import Decimal
    from django.db.models import Q
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    try:
        cuenta_id = request.GET.get('cuenta_id')
        ejercicio_id = request.GET.get('ejercicio_id')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        
        if not cuenta_id:
            return JsonResponse({'error': 'Debe especificar una cuenta'}, status=400)
        
        cuenta = PlanCuenta.objects.get(id=cuenta_id)
        
        # Construir query
        query = Q(cuenta=cuenta)
        if ejercicio_id:
            query &= Q(asiento__ejercicio_id=ejercicio_id)
        if fecha_desde:
            query &= Q(asiento__fecha__gte=fecha_desde)
        if fecha_hasta:
            query &= Q(asiento__fecha__lte=fecha_hasta)
        
        items = ItemAsiento.objects.filter(query).select_related('asiento').order_by('asiento__fecha', 'asiento__numero')
        
        # Calcular saldo inicial
        saldo_inicial = Decimal('0')
        if fecha_desde:
            items_anteriores = ItemAsiento.objects.filter(cuenta=cuenta, asiento__fecha__lt=fecha_desde)
            if ejercicio_id:
                items_anteriores = items_anteriores.filter(asiento__ejercicio_id=ejercicio_id)
            for item in items_anteriores:
                saldo_inicial += item.debe - item.haber
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Libro Mayor"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # T铆tulo
        ws.merge_cells('A1:F1')
        ws['A1'] = 'LIBRO MAYOR'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Informaci贸n de la cuenta
        ws['A2'] = f'Cuenta: {cuenta.codigo} - {cuenta.nombre}'
        ws['A2'].font = Font(bold=True)
        ws['A3'] = f'Tipo: {cuenta.tipo} | Nivel: {cuenta.nivel}'
        
        if fecha_desde and fecha_hasta:
            ws['A4'] = f'Per铆odo: {fecha_desde} al {fecha_hasta}'
        
        # Saldo inicial
        ws['A6'] = 'Saldo Inicial:'
        ws['A6'].font = Font(bold=True)
        ws['B6'] = float(saldo_inicial)
        ws['B6'].number_format = '#,##0.00'
        
        # Encabezados
        headers = ['Fecha', 'Asiento', 'Descripci贸n', 'Debe', 'Haber', 'Saldo']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        row = 9
        saldo_acumulado = saldo_inicial
        total_debe = Decimal('0')
        total_haber = Decimal('0')
        
        for item in items:
            saldo_acumulado += item.debe - item.haber
            total_debe += item.debe
            total_haber += item.haber
            
            ws.cell(row=row, column=1).value = str(item.asiento.fecha)
            ws.cell(row=row, column=2).value = item.asiento.numero
            ws.cell(row=row, column=3).value = item.asiento.descripcion
            ws.cell(row=row, column=4).value = float(item.debe)
            ws.cell(row=row, column=5).value = float(item.haber)
            ws.cell(row=row, column=6).value = float(saldo_acumulado)
            
            # Formato num茅rico
            for col in [4, 5, 6]:
                ws.cell(row=row, column=col).number_format = '#,##0.00'
            
            row += 1
        
        # Totales
        ws.cell(row=row, column=1).value = 'TOTALES'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4).value = float(total_debe)
        ws.cell(row=row, column=5).value = float(total_haber)
        ws.cell(row=row, column=6).value = float(saldo_acumulado)
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).border = border
            if col >= 4:
                ws.cell(row=row, column=col).number_format = '#,##0.00'
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="mayor_{cuenta.codigo.replace(".", "_")}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ==========================================
# API REPORTES CONTABLES
# ==========================================

@require_http_methods(["GET"])
def api_reporte_libro_diario(request):
    """API para generar reporte del Libro Diario"""
    from administrar.models import Asiento, ItemAsiento, EjercicioContable
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    try:
        ejercicio_id = request.GET.get('ejercicio_id')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        
        if not ejercicio_id:
            return JsonResponse({'error': 'Debe especificar un ejercicio'}, status=400)
        
        ejercicio = EjercicioContable.objects.get(id=ejercicio_id)
        
        # Obtener asientos
        asientos = Asiento.objects.filter(ejercicio_id=ejercicio_id)
        if fecha_desde:
            asientos = asientos.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            asientos = asientos.filter(fecha__lte=fecha_hasta)
        asientos = asientos.order_by('fecha', 'numero')
        
        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Libro Diario"
        
        # T铆tulo
        ws.merge_cells('A1:E1')
        ws['A1'] = 'LIBRO DIARIO'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f'Ejercicio: {ejercicio.descripcion}'
        ws['A3'] = f'Per铆odo: {fecha_desde or ejercicio.fecha_inicio} al {fecha_hasta or ejercicio.fecha_fin}'
        
        # Encabezados
        headers = ['Fecha', 'Asiento', 'Cuenta', 'Debe', 'Haber']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos
        row = 6
        for asiento in asientos:
            items = ItemAsiento.objects.filter(asiento=asiento).select_related('cuenta')
            
            for idx, item in enumerate(items):
                if idx == 0:
                    ws.cell(row=row, column=1).value = str(asiento.fecha)
                    ws.cell(row=row, column=2).value = asiento.numero
                
                ws.cell(row=row, column=3).value = f"{item.cuenta.codigo} - {item.cuenta.nombre}"
                ws.cell(row=row, column=4).value = float(item.debe)
                ws.cell(row=row, column=5).value = float(item.haber)
                
                ws.cell(row=row, column=4).number_format = '#,##0.00'
                ws.cell(row=row, column=5).number_format = '#,##0.00'
                
                row += 1
            row += 1
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="libro_diario_{ejercicio.descripcion.replace(" ", "_")}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def api_reporte_estado_resultados(request):
    """API para generar Estado de Resultados"""
    from administrar.models import PlanCuenta, ItemAsiento, EjercicioContable
    from decimal import Decimal
    from django.db.models import Sum
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    
    try:
        ejercicio_id = request.GET.get('ejercicio_id')
        if not ejercicio_id:
            return JsonResponse({'error': 'Debe especificar un ejercicio'}, status=400)
        
        ejercicio = EjercicioContable.objects.get(id=ejercicio_id)
        
        cuentas_ingresos = PlanCuenta.objects.filter(tipo='INGRESO')
        cuentas_egresos = PlanCuenta.objects.filter(tipo='EGRESO')
        
        total_ingresos = Decimal('0')
        total_egresos = Decimal('0')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estado de Resultados"
        
        ws.merge_cells('A1:C1')
        ws['A1'] = 'ESTADO DE RESULTADOS'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f'Ejercicio: {ejercicio.descripcion}'
        
        row = 4
        ws.cell(row=row, column=1).value = 'INGRESOS'
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for cuenta in cuentas_ingresos:
            items = ItemAsiento.objects.filter(cuenta=cuenta, asiento__ejercicio_id=ejercicio_id)
            haber = items.aggregate(total=Sum('haber'))['total'] or Decimal('0')
            debe = items.aggregate(total=Sum('debe'))['total'] or Decimal('0')
            saldo = haber - debe
            
            if saldo != 0:
                ws.cell(row=row, column=1).value = f"{cuenta.codigo} - {cuenta.nombre}"
                ws.cell(row=row, column=2).value = float(saldo)
                ws.cell(row=row, column=2).number_format = '#,##0.00'
                total_ingresos += saldo
                row += 1
        
        ws.cell(row=row, column=1).value = 'TOTAL INGRESOS'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = float(total_ingresos)
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 2
        
        ws.cell(row=row, column=1).value = 'EGRESOS'
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for cuenta in cuentas_egresos:
            items = ItemAsiento.objects.filter(cuenta=cuenta, asiento__ejercicio_id=ejercicio_id)
            debe = items.aggregate(total=Sum('debe'))['total'] or Decimal('0')
            haber = items.aggregate(total=Sum('haber'))['total'] or Decimal('0')
            saldo = debe - haber
            
            if saldo != 0:
                ws.cell(row=row, column=1).value = f"{cuenta.codigo} - {cuenta.nombre}"
                ws.cell(row=row, column=2).value = float(saldo)
                ws.cell(row=row, column=2).number_format = '#,##0.00'
                total_egresos += saldo
                row += 1
        
        ws.cell(row=row, column=1).value = 'TOTAL EGRESOS'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = float(total_egresos)
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 2
        
        resultado = total_ingresos - total_egresos
        ws.cell(row=row, column=1).value = 'RESULTADO DEL EJERCICIO'
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).value = float(resultado)
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="estado_resultados_{ejercicio.descripcion.replace(" ", "_")}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def api_reporte_balance_general(request):
    """API para generar Balance General (Estado de Situaci贸n Patrimonial)"""
    from administrar.models import PlanCuenta, ItemAsiento, EjercicioContable
    from decimal import Decimal
    from django.db.models import Sum
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    try:
        ejercicio_id = request.GET.get('ejercicio_id')
        if not ejercicio_id:
            return JsonResponse({'error': 'Debe especificar un ejercicio'}, status=400)
        
        ejercicio = EjercicioContable.objects.get(id=ejercicio_id)
        
        # 1. Calcular Resultado del Ejercicio
        ingresos = ItemAsiento.objects.filter(
            cuenta__tipo='R_POS', 
            asiento__ejercicio_id=ejercicio_id
        ).aggregate(
            total=Sum('haber') - Sum('debe')
        )['total'] or Decimal('0')
        
        egresos = ItemAsiento.objects.filter(
            cuenta__tipo='R_NEG', 
            asiento__ejercicio_id=ejercicio_id
        ).aggregate(
            total=Sum('debe') - Sum('haber')
        )['total'] or Decimal('0')
        
        resultado_ejercicio = ingresos - egresos
        
        # 2. Obtener cuentas patrimoniales
        cuentas_activo = PlanCuenta.objects.filter(tipo='ACTIVO').order_by('codigo')
        cuentas_pasivo = PlanCuenta.objects.filter(tipo='PASIVO').order_by('codigo')
        cuentas_pn = PlanCuenta.objects.filter(tipo='PN').order_by('codigo')
        
        # 3. Generar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance General"
        
        # Estilos
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=16)
        section_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # T铆tulo
        ws.merge_cells('A1:C1')
        ws['A1'] = 'BALANCE GENERAL'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A2'] = f'Ejercicio: {ejercicio.descripcion}'
        
        row = 4
        
        # --- ACTIVO ---
        ws.cell(row=row, column=1).value = 'ACTIVO'
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = section_fill
        row += 1
        
        total_activo = Decimal('0')
        
        for cuenta in cuentas_activo:
            saldo = ItemAsiento.objects.filter(
                cuenta=cuenta, 
                asiento__ejercicio_id=ejercicio_id
            ).aggregate(
                total=Sum('debe') - Sum('haber')
            )['total'] or Decimal('0')
            
            # Mostrar solo si tiene saldo o es imputable con movimiento
            if saldo != 0:
                ws.cell(row=row, column=1).value = f"{cuenta.codigo} - {cuenta.nombre}"
                ws.cell(row=row, column=2).value = float(saldo)
                ws.cell(row=row, column=2).number_format = '#,##0.00'
                total_activo += saldo
                row += 1
                
        ws.cell(row=row, column=1).value = 'TOTAL ACTIVO'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = float(total_activo)
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 2
        
        # --- PASIVO ---
        ws.cell(row=row, column=1).value = 'PASIVO'
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = section_fill
        row += 1
        
        total_pasivo = Decimal('0')
        
        for cuenta in cuentas_pasivo:
            saldo = ItemAsiento.objects.filter(
                cuenta=cuenta, 
                asiento__ejercicio_id=ejercicio_id
            ).aggregate(
                total=Sum('haber') - Sum('debe')
            )['total'] or Decimal('0')
            
            if saldo != 0:
                ws.cell(row=row, column=1).value = f"{cuenta.codigo} - {cuenta.nombre}"
                ws.cell(row=row, column=2).value = float(saldo)
                ws.cell(row=row, column=2).number_format = '#,##0.00'
                total_pasivo += saldo
                row += 1
                
        ws.cell(row=row, column=1).value = 'TOTAL PASIVO'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = float(total_pasivo)
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 2
        
        # --- PATRIMONIO NETO ---
        ws.cell(row=row, column=1).value = 'PATRIMONIO NETO'
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = section_fill
        row += 1
        
        total_pn = Decimal('0')
        
        for cuenta in cuentas_pn:
            saldo = ItemAsiento.objects.filter(
                cuenta=cuenta, 
                asiento__ejercicio_id=ejercicio_id
            ).aggregate(
                total=Sum('haber') - Sum('debe')
            )['total'] or Decimal('0')
            
            if saldo != 0:
                ws.cell(row=row, column=1).value = f"{cuenta.codigo} - {cuenta.nombre}"
                ws.cell(row=row, column=2).value = float(saldo)
                ws.cell(row=row, column=2).number_format = '#,##0.00'
                total_pn += saldo
                row += 1
        
        # Agregar Resultado del Ejercicio
        ws.cell(row=row, column=1).value = "Resultado del Ejercicio"
        ws.cell(row=row, column=1).font = Font(italic=True)
        ws.cell(row=row, column=2).value = float(resultado_ejercicio)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        total_pn += resultado_ejercicio
        row += 1
        
        ws.cell(row=row, column=1).value = 'TOTAL PATRIMONIO NETO'
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = float(total_pn)
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 2
        
        # --- TOTAL PASIVO + PN ---
        ws.cell(row=row, column=1).value = 'TOTAL PASIVO + PATRIMONIO NETO'
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws.cell(row=row, column=2).value = float(total_pasivo + total_pn)
        ws.cell(row=row, column=2).font = Font(bold=True, size=11)
        ws.cell(row=row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="balance_general_{ejercicio.descripcion.replace(" ", "_")}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
@require_http_methods(["GET"])
@login_required
@verificar_permiso('contabilidad')
def api_reporte_resumen_ejercicio(request):
    """API para generar resumen del ejercicio"""
    from administrar.models import PlanCuenta, ItemAsiento, EjercicioContable
    from decimal import Decimal
    from django.db.models import Sum
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    try:
        ejercicio_id = request.GET.get('ejercicio_id')
        if not ejercicio_id:
            return JsonResponse({'error': 'Debe especificar un ejercicio'}, status=400)
        
        ejercicio = EjercicioContable.objects.get(id=ejercicio_id)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen Ejercicio"
        
        ws.merge_cells('A1:D1')
        ws['A1'] = f'RESUMEN DE EJERCICIO: {ejercicio.descripcion}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['C贸digo', 'Cuenta', 'Tipo', 'Saldo Final']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        row = 4
        cuentas = PlanCuenta.objects.all().order_by('codigo')
        
        for cuenta in cuentas:
            # Calcular saldo seg煤n tipo
            debe = ItemAsiento.objects.filter(cuenta=cuenta, asiento__ejercicio_id=ejercicio_id).aggregate(Sum('debe'))['debe__sum'] or 0
            haber = ItemAsiento.objects.filter(cuenta=cuenta, asiento__ejercicio_id=ejercicio_id).aggregate(Sum('haber'))['haber__sum'] or 0
            
            if cuenta.tipo in ['ACTIVO', 'R_NEG']:
                saldo = debe - haber
            else:
                saldo = haber - debe
                
            if saldo != 0:
                ws.cell(row=row, column=1).value = cuenta.codigo
                ws.cell(row=row, column=2).value = cuenta.nombre
                ws.cell(row=row, column=3).value = cuenta.get_tipo_display()
                ws.cell(row=row, column=4).value = float(saldo)
                ws.cell(row=row, column=4).number_format = '#,##0.00'
                row += 1
                
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="resumen_{ejercicio.descripcion.replace(" ", "_")}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


#====================================
# GESTIN DE USUARIOS
#====================================
@login_required
def usuarios_lista(request):
    return render(request, 'administrar/usuarios.html')

@login_required
@verificar_permiso('usuarios')
def api_usuarios_listar(request):
    from django.contrib.auth.models import User
    from .models import PerfilUsuario
    
    usuarios = User.objects.all().values(
        'id', 'username', 'email', 'first_name', 'last_name', 
        'is_active', 'is_staff', 'last_login'
    )
    
    data = []
    for u in usuarios:
        u['rol'] = 'Administrador' if u['is_staff'] else 'Vendedor'
        u['last_login'] = u['last_login'].strftime('%d/%m/%Y %H:%M') if u['last_login'] else 'Nunca'
        
        # Obtener permisos
        try:
            perfil = PerfilUsuario.objects.get(user_id=u['id'])
            u['permisos'] = {
                'ventas': perfil.acceso_ventas,
                'compras': perfil.acceso_compras,
                'productos': perfil.acceso_productos,
                'clientes': perfil.acceso_clientes,
                'proveedores': perfil.acceso_proveedores,
                'caja': perfil.acceso_caja,
                'contabilidad': perfil.acceso_contabilidad,
                'configuracion': perfil.acceso_configuracion,
                'usuarios': perfil.acceso_usuarios,
                'reportes': perfil.acceso_reportes,
                'pedidos': perfil.acceso_pedidos,
                'bancos': perfil.acceso_bancos,
                'ctacte': perfil.acceso_ctacte,
                'remitos': perfil.acceso_remitos,
            }
        except PerfilUsuario.DoesNotExist:
            u['permisos'] = {}
            
        data.append(u)
    
    return JsonResponse({'ok': True, 'data': data})

@login_required
@verificar_permiso('usuarios')
def api_usuario_detalle(request, id):
    from django.contrib.auth.models import User
    from .models import PerfilUsuario
    
    try:
        usuario = User.objects.get(pk=id)
        
        permisos = {}
        try:
            perfil = PerfilUsuario.objects.get(user=usuario)
            permisos = {
                'ventas': perfil.acceso_ventas,
                'compras': perfil.acceso_compras,
                'productos': perfil.acceso_productos,
                'clientes': perfil.acceso_clientes,
                'proveedores': perfil.acceso_proveedores,
                'caja': perfil.acceso_caja,
                'contabilidad': perfil.acceso_contabilidad,
                'configuracion': perfil.acceso_configuracion,
                'usuarios': perfil.acceso_usuarios,
                'reportes': perfil.acceso_reportes,
                'pedidos': perfil.acceso_pedidos,
                'bancos': perfil.acceso_bancos,
                'ctacte': perfil.acceso_ctacte,
                'remitos': perfil.acceso_remitos,
            }
        except PerfilUsuario.DoesNotExist:
            pass

        data = {
            'id': usuario.id,
            'username': usuario.username,
            'email': usuario.email,
            'first_name': usuario.first_name,
            'last_name': usuario.last_name,
            'is_active': usuario.is_active,
            'is_staff': usuario.is_staff,
            'rol': 'Administrador' if usuario.is_staff else 'Vendedor',
            'permisos': permisos
        }
        return JsonResponse({'ok': True, 'data': data})
    except User.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Usuario no encontrado'})

@require_POST
@require_POST
@login_required
@verificar_permiso('usuarios')
def api_usuario_crear(request):
    from django.contrib.auth.models import User
    from .models import PerfilUsuario
    import json
    
    try:
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
        email = data.get('email', '')
        first_name = data.get('first_name', '')
        last_name = data.get('last_name', '')
        is_active = data.get('is_active', True)
        is_staff = data.get('is_staff', False)
        permisos = data.get('permisos', {})
        
        if not username or not password:
            return JsonResponse({'ok': False, 'error': 'Usuario y contrase帽a son obligatorios'})
        
        if User.objects.filter(username=username).exists():
            return JsonResponse({'ok': False, 'error': 'El nombre de usuario ya existe'})
        
        usuario = User.objects.create_user(
            username=username,
            password=password,
            email=email,
            first_name=first_name,
            last_name=last_name,
            is_active=is_active,
            is_staff=is_staff
        )
        
        # Crear perfil con permisos
        PerfilUsuario.objects.create(
            user=usuario,
            acceso_ventas=permisos.get('ventas', False),
            acceso_compras=permisos.get('compras', False),
            acceso_productos=permisos.get('productos', False),
            acceso_clientes=permisos.get('clientes', False),
            acceso_proveedores=permisos.get('proveedores', False),
            acceso_caja=permisos.get('caja', False),
            acceso_contabilidad=permisos.get('contabilidad', False),
            acceso_configuracion=permisos.get('configuracion', False),
            acceso_usuarios=permisos.get('usuarios', False),
            acceso_reportes=permisos.get('reportes', False),
            acceso_pedidos=permisos.get('pedidos', False),
            acceso_bancos=permisos.get('bancos', False),
            acceso_ctacte=permisos.get('ctacte', False),
            acceso_remitos=permisos.get('remitos', False),
        )
        
        return JsonResponse({'ok': True, 'data': {'id': usuario.id}})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@require_POST
@require_POST
@login_required
@verificar_permiso('usuarios')
def api_usuario_editar(request, id):
    from django.contrib.auth.models import User
    from .models import PerfilUsuario
    import json
    
    try:
        usuario = User.objects.get(pk=id)
        data = json.loads(request.body)
        
        usuario.username = data.get('username', usuario.username)
        usuario.email = data.get('email', usuario.email)
        usuario.first_name = data.get('first_name', usuario.first_name)
        usuario.last_name = data.get('last_name', usuario.last_name)
        usuario.is_active = data.get('is_active', usuario.is_active)
        usuario.is_staff = data.get('is_staff', usuario.is_staff)
        
        if data.get('password'):
            usuario.set_password(data['password'])
        
        usuario.save()
        
        # Actualizar permisos
        permisos = data.get('permisos', {})
        perfil, created = PerfilUsuario.objects.get_or_create(user=usuario)
        
        perfil.acceso_ventas = permisos.get('ventas', perfil.acceso_ventas)
        perfil.acceso_compras = permisos.get('compras', perfil.acceso_compras)
        perfil.acceso_productos = permisos.get('productos', perfil.acceso_productos)
        perfil.acceso_clientes = permisos.get('clientes', perfil.acceso_clientes)
        perfil.acceso_proveedores = permisos.get('proveedores', perfil.acceso_proveedores)
        perfil.acceso_caja = permisos.get('caja', perfil.acceso_caja)
        perfil.acceso_contabilidad = permisos.get('contabilidad', perfil.acceso_contabilidad)
        perfil.acceso_configuracion = permisos.get('configuracion', perfil.acceso_configuracion)
        perfil.acceso_usuarios = permisos.get('usuarios', perfil.acceso_usuarios)
        perfil.acceso_reportes = permisos.get('reportes', perfil.acceso_reportes)
        perfil.acceso_pedidos = permisos.get('pedidos', perfil.acceso_pedidos)
        perfil.acceso_bancos = permisos.get('bancos', perfil.acceso_bancos)
        perfil.acceso_ctacte = permisos.get('ctacte', perfil.acceso_ctacte)
        perfil.acceso_remitos = permisos.get('remitos', perfil.acceso_remitos)
        
        perfil.save()
        
        return JsonResponse({'ok': True})
    except User.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Usuario no encontrado'})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@require_http_methods(['DELETE'])
@require_http_methods(['DELETE'])
@login_required
@verificar_permiso('usuarios')
def api_usuario_eliminar(request, id):
    from django.contrib.auth.models import User
    
    try:
        usuario = User.objects.get(pk=id)
        
        if usuario.id == request.user.id:
            return JsonResponse({'ok': False, 'error': 'No puedes eliminar tu propio usuario'})
        
        usuario.delete()
        return JsonResponse({'ok': True})
    except User.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Usuario no encontrado'})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# =========================================
#  API SEGURIDAD Y ACCESOS
# =========================================

@require_POST
@csrf_protect
@login_required
def api_cambiar_contrasena(request):
    """API para cambiar la contrase帽a del usuario actual"""
    from django.contrib.auth import update_session_auth_hash
    from .middleware import log_activity, get_client_ip
    
    try:
        data = json.loads(request.body.decode('utf-8'))
        contrasena_actual = data.get('contrasena_actual', '')
        nueva_contrasena = data.get('nueva_contrasena', '')
        confirmar_contrasena = data.get('confirmar_contrasena', '')
        
        # Validaciones
        if not request.user.check_password(contrasena_actual):
            return JsonResponse({'ok': False, 'error': 'La contrase帽a actual es incorrecta'}, status=400)
        
        if nueva_contrasena != confirmar_contrasena:
            return JsonResponse({'ok': False, 'error': 'Las contrase帽as no coinciden'}, status=400)
        
        if len(nueva_contrasena) < 6:
            return JsonResponse({'ok': False, 'error': 'La contrase帽a debe tener al menos 6 caracteres'}, status=400)
        
        # Cambiar contrase帽a
        request.user.set_password(nueva_contrasena)
        request.user.save()
        
        # Mantener la sesi贸n activa despu茅s del cambio de contrase帽a
        update_session_auth_hash(request, request.user)
        
        # Registrar en bit谩cora
        log_activity(
            user=request.user,
            action_type='PASSWORD_CHANGE',
            module='SEGURIDAD',
            description=f'Cambio de contrase帽a',
            request=request
        )
        
        return JsonResponse({'ok': True, 'mensaje': 'Contrase帽a actualizada correctamente'})
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@login_required
@verificar_permiso('usuarios')
def api_historial_login(request):
    """API para listar el historial de inicios de sesi贸n"""
    from .models import LoginHistory
    from django.core.paginator import Paginator
    
    try:
        # Filtros
        username = request.GET.get('username', '')
        success = request.GET.get('success', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        # Query base
        queryset = LoginHistory.objects.all()
        
        # Aplicar filtros
        if username:
            queryset = queryset.filter(username__icontains=username)
        
        if success == 'true':
            queryset = queryset.filter(success=True)
        elif success == 'false':
            queryset = queryset.filter(success=False)
        
        # Paginaci贸n
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        # Serializar datos
        data = []
        for login in page_obj:
            data.append({
                'id': login.id,
                'username': login.username,
                'ip_address': login.ip_address,
                'timestamp': login.timestamp.strftime('%d/%m/%Y - %H:%M'),
                'success': login.success,
                'user_agent': login.user_agent[:100] if login.user_agent else ''
            })
        
        return JsonResponse({
            'ok': True,
            'data': data,
            'total': paginator.count,
            'pages': paginator.num_pages,
            'current_page': page
        })
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@login_required
@verificar_permiso('usuarios')
def api_sesiones_activas(request):
    """API para listar sesiones activas"""
    from .models import ActiveSession
    
    try:
        sesiones = ActiveSession.objects.select_related('user').all()
        
        data = []
        for sesion in sesiones:
            data.append({
                'id': sesion.id,
                'user_id': sesion.user.id,
                'username': sesion.user.username,
                'email': sesion.user.email,
                'ip_address': sesion.ip_address,
                'login_time': sesion.login_time.strftime('%d/%m/%Y - %H:%M'),
                'last_activity': sesion.last_activity.strftime('%d/%m/%Y - %H:%M'),
                'session_key': sesion.session_key
            })
        
        return JsonResponse({'ok': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@require_POST
@csrf_protect
@login_required
@verificar_permiso('usuarios')
def api_cerrar_sesion_usuario(request, session_id):
    """API para cerrar una sesi贸n activa de un usuario"""
    from .models import ActiveSession
    from django.contrib.sessions.models import Session
    from .middleware import log_activity
    
    try:
        sesion = ActiveSession.objects.select_related('user').get(id=session_id)
        
        # No permitir cerrar la propia sesi贸n
        if sesion.user.id == request.user.id:
            return JsonResponse({'ok': False, 'error': 'No puedes cerrar tu propia sesi贸n'}, status=400)
        
        # Eliminar la sesi贸n de Django
        try:
            Session.objects.filter(session_key=sesion.session_key).delete()
        except:
            pass
        
        # Registrar en bit谩cora
        log_activity(
            user=request.user,
            action_type='SESSION_CLOSE',
            module='SEGURIDAD',
            description=f'Cerr贸 sesi贸n remota de {sesion.user.username}',
            details={'target_user': sesion.user.username, 'ip': sesion.ip_address},
            request=request
        )
        
        # Eliminar registro de sesi贸n activa
        sesion.delete()
        
        return JsonResponse({'ok': True, 'mensaje': 'Sesi贸n cerrada correctamente'})
        
    except ActiveSession.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Sesi贸n no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@login_required
@verificar_permiso('usuarios')
def api_bitacora_actividades(request):
    """API para listar la bit谩cora de actividades"""
    from .models import ActivityLog
    from django.core.paginator import Paginator
    
    try:
        # Filtros
        username = request.GET.get('username', '')
        module = request.GET.get('module', '')
        action_type = request.GET.get('action_type', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        # Query base
        queryset = ActivityLog.objects.select_related('user').all()
        
        # Aplicar filtros
        if username:
            queryset = queryset.filter(user__username__icontains=username)
        
        if module:
            queryset = queryset.filter(module=module)
        
        if action_type:
            queryset = queryset.filter(action_type=action_type)
        
        # Paginaci贸n
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        # Serializar datos
        data = []
        for log in page_obj:
            data.append({
                'id': log.id,
                'username': log.user.username if log.user else 'Sistema',
                'timestamp': log.timestamp.strftime('%d/%m/%Y - %H:%M'),
                'action_type': log.action_type,
                'action_type_display': log.get_action_type_display(),
                'module': log.module,
                'module_display': log.get_module_display(),
                'description': log.description,
                'ip_address': log.ip_address or '',
                'details': log.details
            })
        
        return JsonResponse({
            'ok': True,
            'data': data,
            'total': paginator.count,
            'pages': paginator.num_pages,
            'current_page': page
        })
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


# ==========================================
# SISTEMA DE BACKUPS
# ==========================================

@login_required
def backups(request):
    """Renderiza la p谩gina de gesti贸n de backups"""
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para acceder a esta secci贸n.')
        return redirect('menu')
    return render(request, 'administrar/backups.html')

@login_required
def api_listar_backups(request):
    """API para listar los backups disponibles"""
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'No autorizado'}, status=403)
    
    from .models import Backup
    
    backups = Backup.objects.all().order_by('-fecha_creacion')
    data = []
    
    for b in backups:
        data.append({
            'id': b.id,
            'nombre': b.nombre,
            'fecha': b.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            'tamanio': b.tamanio_formateado(),
            'ubicacion': b.get_ubicacion_display(),
            'ubicacion_code': b.ubicacion,
            'creado_por': b.creado_por.username if b.creado_por else 'Sistema'
        })
        
    return JsonResponse({'ok': True, 'data': data})

@csrf_exempt
@login_required
def api_crear_backup(request):
    """API para crear un nuevo backup"""
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'No autorizado'}, status=403)
        
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'M茅todo no permitido'}, status=405)
        
    import subprocess
    import os
    import zipfile
    from django.conf import settings
    from .models import Backup
    
    try:
        data = json.loads(request.body)
        ubicacion = data.get('ubicacion', 'LOCAL')
        nombre_personalizado = data.get('nombre', '').strip()
        
        # Configuraci贸n de BD
        db_settings = settings.DATABASES['default']
        db_name = db_settings['NAME']
        db_user = db_settings['USER']
        db_password = db_settings['PASSWORD']
        db_host = db_settings['HOST']
        
        # Directorios
        base_backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        local_backup_dir = os.path.join(base_backup_dir, 'local')
        os.makedirs(local_backup_dir, exist_ok=True)
        
        # Nombre del archivo
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        if nombre_personalizado:
            filename = f"backup_{nombre_personalizado}_{timestamp}.sql"
        else:
            filename = f"backup_{db_name}_{timestamp}.sql"
            
        file_path = os.path.join(local_backup_dir, filename)
        
        # Comando mysqldump
        dump_cmd = [
            'mysqldump',
            f'--host={db_host}',
            f'--user={db_user}',
        ]
        
        if db_password:
            dump_cmd.append(f'--password={db_password}')
            
        dump_cmd.append(db_name)
        
        # Ejecutar backup
        with open(file_path, 'w') as f:
            process = subprocess.run(dump_cmd, stdout=f, stderr=subprocess.PIPE)
            
        if process.returncode != 0:
            error_msg = process.stderr.decode('utf-8')
            return JsonResponse({'ok': False, 'error': f'Error al generar backup: {error_msg}'})
            
        # Comprimir a ZIP
        zip_filename = filename.replace('.sql', '.zip')
        zip_path = os.path.join(local_backup_dir, zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(file_path, arcname=filename)
            
        # Eliminar archivo SQL original
        os.remove(file_path)
        
        # Calcular tama帽o
        file_size = os.path.getsize(zip_path)
        
        # Guardar registro en BD
        backup = Backup.objects.create(
            nombre=nombre_personalizado if nombre_personalizado else f"Backup {timestamp}",
            archivo=zip_path,
            tamanio=file_size,
            ubicacion='LOCAL',
            creado_por=request.user
        )
        
        return JsonResponse({
            'ok': True, 
            'mensaje': 'Backup creado exitosamente',
            'backup': {
                'id': backup.id,
                'nombre': backup.nombre,
                'fecha': backup.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'tamanio': backup.tamanio_formateado()
            }
        })
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@login_required
def api_descargar_backup(request, id):
    """Descargar archivo de backup"""
    if not request.user.is_staff:
        return HttpResponse("No autorizado", status=403)
        
    from .models import Backup
    import os
    
    backup = get_object_or_404(Backup, pk=id)
    
    if os.path.exists(backup.archivo):
        with open(backup.archivo, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/zip")
            response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(backup.archivo)
            return response
    else:
        return HttpResponse("Archivo no encontrado", status=404)

@csrf_exempt
@login_required
def api_restaurar_backup(request, id):
    """Restaurar base de datos desde backup"""
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'No autorizado'}, status=403)
        
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'M茅todo no permitido'}, status=405)
        
    from .models import Backup
    import subprocess
    import os
    import zipfile
    from django.conf import settings
    
    try:
        backup = get_object_or_404(Backup, pk=id)
        
        if not os.path.exists(backup.archivo):
            return JsonResponse({'ok': False, 'error': 'Archivo de backup no encontrado'})
            
        # Configuraci贸n de BD
        db_settings = settings.DATABASES['default']
        db_name = db_settings['NAME']
        db_user = db_settings['USER']
        db_password = db_settings['PASSWORD']
        db_host = db_settings['HOST']
        
        # Directorio temporal
        temp_dir = os.path.join(settings.BASE_DIR, 'backups', 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        # Descomprimir
        sql_file = None
        with zipfile.ZipFile(backup.archivo, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
            for name in zip_ref.namelist():
                if name.endswith('.sql'):
                    sql_file = os.path.join(temp_dir, name)
                    break
        
        if not sql_file:
            return JsonResponse({'ok': False, 'error': 'No se encontr贸 archivo SQL dentro del backup'})
            
        # Comando mysql para restaurar
        restore_cmd = [
            'mysql',
            f'--host={db_host}',
            f'--user={db_user}',
        ]
        
        if db_password:
            restore_cmd.append(f'--password={db_password}')
            
        restore_cmd.append(db_name)
        
        # Ejecutar restauraci贸n
        with open(sql_file, 'r') as f:
            process = subprocess.run(restore_cmd, stdin=f, stderr=subprocess.PIPE)
            
        # Limpiar
        if os.path.exists(sql_file):
            os.remove(sql_file)
            
        if process.returncode != 0:
            error_msg = process.stderr.decode('utf-8')
            return JsonResponse({'ok': False, 'error': f'Error al restaurar: {error_msg}'})
            
        return JsonResponse({'ok': True, 'mensaje': 'Sistema restaurado exitosamente'})
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@csrf_exempt
@login_required
def api_eliminar_backup(request, id):
    """Eliminar backup"""
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'No autorizado'}, status=403)
        
    if request.method != 'DELETE':
        return JsonResponse({'ok': False, 'error': 'M茅todo no permitido'}, status=405)
        
    from .models import Backup
    import os
    
    try:
        backup = get_object_or_404(Backup, pk=id)
        
        if os.path.exists(backup.archivo):
            try:
                os.remove(backup.archivo)
            except Exception as e:
                print(f"Error eliminando archivo: {e}")
                
        backup.delete()
        
        return JsonResponse({'ok': True, 'mensaje': 'Backup eliminado correctamente'})
        
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


# =========================================
#  GESTIN DE CHEQUES
# =========================================
@login_required
def cheques_lista(request):
    return render(request, 'administrar/cheques/cheques.html')




@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_cheques_crear(request):
    try:
        import json
        from .models import Cheque, Cliente, CuentaBancaria, MovimientoBanco
        from .services import AccountingService
        data = json.loads(request.body)
        
        cliente = None
        if data.get('cliente_id'):
            cliente = Cliente.objects.get(id=data['cliente_id'])

        c = Cheque.objects.create(
            numero=data['numero'],
            banco=data['banco'],
            monto=data['monto'],
            tipo=data['tipo'],
            estado=data['estado'],
            fecha_emision=data['fecha_emision'],
            fecha_pago=data['fecha_pago'],
            cliente=cliente,
            firmante=data.get('firmante', ''),
            cuit_firmante=data.get('cuit_firmante', ''),
            destinatario=data.get('destinatario', ''),
            observaciones=data.get('observaciones', '')
        )
        
        # Registrar Alta Contable
        AccountingService.registrar_alta_cheque(c)

        # Chequear Deposito Inicial
        if c.estado == 'DEPOSITADO' and data.get('cuenta_bancaria_id'):
            try:
                cuenta = CuentaBancaria.objects.get(id=data['cuenta_bancaria_id'])
                MovimientoBanco.objects.create(
                    cuenta=cuenta,
                    fecha=c.fecha_pago, # Usamos fecha pago como fecha deposito
                    descripcion=f"Dep贸sito Cheque {c.banco} #{c.numero}",
                    monto=c.monto,
                    referencia_interna=f"Cheque #{c.id}"
                )
                cuenta.saldo_actual = float(cuenta.saldo_actual) + float(c.monto)
                cuenta.save()
                
                # Contabilidad Deposito
                AccountingService.registrar_cambio_estado(c, 'CARTERA', cuenta_destino=cuenta) # Simulamos cambio de cartera a depositado
                
            except Exception as e:
                print(f"Error creando movimiento banco: {e}")

        return JsonResponse({'ok': True, 'id': c.id})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_cheques_editar(request, id):
    try:
        import json
        from .models import Cheque, Cliente, CuentaBancaria, MovimientoBanco
        from .services import AccountingService
        data = json.loads(request.body)
        c = Cheque.objects.get(id=id)
        
        estado_anterior = c.estado

        if data.get('cliente_id'):
            c.cliente = Cliente.objects.get(id=data['cliente_id'])
        else:
            c.cliente = None

        c.numero = data['numero']
        c.banco = data['banco']
        c.monto = data['monto']
        c.tipo = data['tipo']
        c.estado = data['estado']
        c.fecha_emision = data['fecha_emision']
        c.fecha_pago = data['fecha_pago']
        c.firmante = data.get('firmante', '')
        c.cuit_firmante = data.get('cuit_firmante', '')
        c.destinatario = data.get('destinatario', '')
        c.observaciones = data.get('observaciones', '')
        c.save()

        # Chequear cambio a DEPOSITADO
        if estado_anterior != 'DEPOSITADO' and c.estado == 'DEPOSITADO' and data.get('cuenta_bancaria_id'):
            try:
                cuenta = CuentaBancaria.objects.get(id=data['cuenta_bancaria_id'])
                MovimientoBanco.objects.create(
                    cuenta=cuenta,
                    fecha=c.fecha_pago, # fecha deposito
                    descripcion=f"Dep贸sito Cheque {c.banco} #{c.numero}",
                    monto=c.monto,
                    referencia_interna=f"Cheque #{c.id}"
                )
                cuenta.saldo_actual = float(cuenta.saldo_actual) + float(c.monto)
                cuenta.save()
                
                # Contabilidad Deposito
                AccountingService.registrar_cambio_estado(c, estado_anterior, cuenta_destino=cuenta)

            except Exception as e:
                print(f"Error creando movimiento banco: {e}")
        else:
            # Otros cambios de estado
            if estado_anterior != c.estado:
                AccountingService.registrar_cambio_estado(c, estado_anterior)

        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})



@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_cheques_eliminar(request, id):
    from .models import Cheque
    try:
        Cheque.objects.get(id=id).delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@require_http_methods(["GET"])
@login_required
def api_clientes_listar_simple(request):
    from .models import Cliente
    clientes = Cliente.objects.filter(activo=True).values('id', 'nombre').order_by('nombre')
    return JsonResponse({'ok': True, 'clientes': list(clientes)})


# =========================================
#  CUENTAS CORRIENTES (Clientes y Proveedores)
# =========================================

# --- VISTAS HTML ---

@login_required
def cc_clientes_lista(request):
    return render(request, 'administrar/ctacte/clientes/lista.html')

@login_required
def cc_cliente_detalle(request, id):
    return render(request, 'administrar/ctacte/clientes/detalle.html', {'cliente_id': id})

@login_required
def cc_proveedores_lista(request):
    return render(request, 'administrar/ctacte/proveedores/lista.html')

@login_required
def cc_proveedor_detalle(request, id):
    return render(request, 'administrar/ctacte/proveedores/detalle.html', {'proveedor_id': id})


# --- API CLIENTES ---

@login_required
def api_cc_clientes_listar(request):
    from .models import Cliente
    from django.db.models import Q
    
    q = request.GET.get('q', '').strip()
    filtro_saldo = request.GET.get('filtro', 'todos')
    
    clientes_qs = Cliente.objects.filter(activo=True)
    
    if q:
        clientes_qs = clientes_qs.filter(Q(nombre__icontains=q) | Q(cuit__icontains=q))
        
    # Filtro de sado
    if filtro_saldo == 'con_deuda':
        clientes_qs = clientes_qs.filter(saldo_actual__gt=0)
    elif filtro_saldo == 'al_dia':
        clientes_qs = clientes_qs.filter(saldo_actual=0)
    elif filtro_saldo == 'a_favor':
        clientes_qs = clientes_qs.filter(saldo_actual__lt=0)
        
    data = []
    for c in clientes_qs[:50]: # Limitar resultados iniciales
        data.append({
            'id': c.id,
            'nombre': c.nombre,
            'cuit': c.cuit,
            'telefono': c.telefono,
            'limite_credito': float(c.limite_credito),
            'saldo_actual': float(c.saldo_actual),
        })
        
    return JsonResponse({'ok': True, 'clientes': data})


@login_required
def api_cc_cliente_movimientos(request, id):
    from .models import Cliente, MovimientoCuentaCorriente
    
    try:
        cliente = Cliente.objects.get(id=id)
        fecha_desde = request.GET.get('desde')
        fecha_hasta = request.GET.get('hasta')
        
        movs = MovimientoCuentaCorriente.objects.filter(cliente=cliente).order_by('-fecha')
        
        if fecha_desde:
            movs = movs.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta:
            movs = movs.filter(fecha__date__lte=fecha_hasta)
            
        data = []
        for m in movs[:100]:
            data.append({
                'id': m.id,
                'fecha': m.fecha.strftime('%Y-%m-%d %H:%M'),
                'tipo': m.tipo,
                'descripcion': m.descripcion,
                'monto': float(m.monto),
                'saldo': float(m.saldo)
            })
            
        return JsonResponse({
            'ok': True, 
            'cliente': {
                'id': cliente.id,
                'nombre': cliente.nombre,
                'cuit': cliente.cuit,
                'telefono': cliente.telefono,
                'saldo_actual': float(cliente.saldo_actual),
                'limite_credito': float(cliente.limite_credito)
            },
            'movimientos': data
        })
    except Cliente.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Cliente no encontrado'})


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_cc_cliente_nuevo_movimiento(request):
    import json
    from .models import Cliente, MovimientoCuentaCorriente
    from decimal import Decimal
    
    try:
        data = json.loads(request.body)
        cliente_id = data.get('cliente_id')
        tipo = data.get('tipo') # DEBE / HABER
        monto = Decimal(str(data.get('monto')))
        descripcion = data.get('descripcion')
        impactar_caja = data.get('impactar_caja', False)
        
        cliente = Cliente.objects.get(id=cliente_id)
        
        # L贸gica de Saldo Cliente:
        # DEBE (Nota Debito, Venta) -> Aumenta Deuda (Saldo +)
        # HABER (Pago, Nota Credito) -> Disminuye Deuda (Saldo -)
        
        nuevo_saldo = cliente.saldo_actual
        if tipo == 'DEBE':
            nuevo_saldo += monto
            # Si impacta caja siendo DEBE, asumimos que es una "Salida" de dinero (ej. devolucion de saldo a favor)
            if impactar_caja:
                 caja = CajaDiaria.objects.filter(estado='ABIERTA').first()
                 MovimientoCaja.objects.create(
                    caja_diaria=caja,
                    usuario=request.user,
                    tipo="Egreso",
                    descripcion=f"Devoluci贸n/Ajuste - Cliente {cliente.nombre}: {descripcion}",
                    monto=monto
                )
        else:
            nuevo_saldo -= monto
            # Si impacta caja siendo HABER, es un COBRO (Entrada de dinero)
            if impactar_caja:
                caja = CajaDiaria.objects.filter(estado='ABIERTA').first()
                MovimientoCaja.objects.create(
                    caja_diaria=caja,
                    usuario=request.user,
                    tipo="Ingreso",
                    descripcion=f"Cobro - Cliente {cliente.nombre}: {descripcion}",
                    monto=monto
                )
            
        # Crear Movimiento
        MovimientoCuentaCorriente.objects.create(
            cliente=cliente,
            tipo=tipo,
            descripcion=descripcion,
            monto=monto,
            saldo=nuevo_saldo,
            # fecha se setea auto_now_add, si necesitamos fecha manual, habria que permitir editarla o sobreescribirla
        )
        
        # Actualizar Cliente
        cliente.saldo_actual = nuevo_saldo
        cliente.save()
        
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# --- API PROVEEDORES ---

@login_required
def api_cc_proveedores_listar(request):
    from .models import Proveedor
    from django.db.models import Q
    
    q = request.GET.get('q', '').strip()
    filtro_saldo = request.GET.get('filtro', 'todos')
    
    prov_qs = Proveedor.objects.all().order_by('nombre')
    
    if q:
        prov_qs = prov_qs.filter(Q(nombre__icontains=q) | Q(cuit__icontains=q))
        
    if filtro_saldo == 'con_deuda':
        prov_qs = prov_qs.filter(saldo_actual__gt=0)
    elif filtro_saldo == 'al_dia':
        prov_qs = prov_qs.filter(saldo_actual=0)
    elif filtro_saldo == 'a_favor':
        prov_qs = prov_qs.filter(saldo_actual__lt=0)
        
    data = []
    for p in prov_qs[:50]:
        data.append({
            'id': p.id,
            'nombre': p.nombre,
            'cuit': p.cuit,
            'telefono': p.telefono,
            'saldo_actual': float(p.saldo_actual),
        })
        
    return JsonResponse({'ok': True, 'proveedores': data})


@login_required
def api_cc_proveedor_movimientos(request, id):
    from .models import Proveedor, MovimientoCuentaCorrienteProveedor
    
    try:
        proveedor = Proveedor.objects.get(id=id)
        fecha_desde = request.GET.get('desde')
        fecha_hasta = request.GET.get('hasta')
        
        movs = MovimientoCuentaCorrienteProveedor.objects.filter(proveedor=proveedor).order_by('-fecha')
        
        if fecha_desde:
            movs = movs.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta:
            movs = movs.filter(fecha__date__lte=fecha_hasta)
            
        data = []
        for m in movs[:100]:
            data.append({
                'id': m.id,
                'fecha': m.fecha.strftime('%Y-%m-%d %H:%M'),
                'tipo': m.tipo,
                'descripcion': m.descripcion,
                'monto': float(m.monto),
                'saldo': float(m.saldo)
            })
            
        return JsonResponse({
            'ok': True, 
            'proveedor': {
                'id': proveedor.id,
                'nombre': proveedor.nombre,
                'cuit': proveedor.cuit,
                'telefono': proveedor.telefono,
                'saldo_actual': float(proveedor.saldo_actual),
            },
            'movimientos': data
        })
    except Proveedor.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Proveedor no encontrado'})


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_cc_proveedor_nuevo_movimiento(request):
    import json
    from .models import Proveedor, MovimientoCuentaCorrienteProveedor
    from decimal import Decimal
    
    try:
        data = json.loads(request.body)
        proveedor_id = data.get('proveedor_id')
        tipo = data.get('tipo') # DEBE / HABER
        monto = Decimal(str(data.get('monto')))
        descripcion = data.get('descripcion')
        impactar_caja = data.get('impactar_caja', False)
        
        proveedor = Proveedor.objects.get(id=proveedor_id)
        
        # L贸gica de Saldo Proveedor (Pasivo):
        # HABER (Compra, Factura Recibida) -> Aumenta Deuda (Saldo +)
        # DEBE (Pago, Nota Credito Recibida) -> Disminuye Deuda (Saldo -)
        
        nuevo_saldo = proveedor.saldo_actual
        if tipo == 'HABER':
            nuevo_saldo += monto
            # Si impacta caja siendo HABER, asumimos Ingreso de dinero (ej. proveedor nos devuelve plata)
            if impactar_caja:
                 caja = CajaDiaria.objects.filter(estado='ABIERTA').first()
                 MovimientoCaja.objects.create(
                    caja_diaria=caja,
                    usuario=request.user,
                    tipo="Ingreso",
                    descripcion=f"Ajuste/Devoluci贸n - Proveedor {proveedor.nombre}: {descripcion}",
                    monto=monto
                )
        else:
            nuevo_saldo -= monto
            # Si impacta caja siendo DEBE, es un PAGO (Salida de dinero)
            if impactar_caja:
                 caja = CajaDiaria.objects.filter(estado='ABIERTA').first()
                 MovimientoCaja.objects.create(
                    caja_diaria=caja,
                    usuario=request.user,
                    tipo="Egreso",
                    descripcion=f"Pago a Proveedor {proveedor.nombre}: {descripcion}",
                    monto=monto
                )
            
        MovimientoCuentaCorrienteProveedor.objects.create(
            proveedor=proveedor,
            tipo=tipo,
            descripcion=descripcion,
            monto=monto,
            saldo=nuevo_saldo
        )
        
        proveedor.saldo_actual = nuevo_saldo
        proveedor.save()
        
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# =========================================
#  RECIBOS DE COBRO Y PAGO
# =========================================

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_recibo_crear(request):
    """
    Crea un recibo formal de cobro o pago con m煤ltiples formas de pago.
    
    Payload:
    {
        "cliente_id": 1,  // o "proveedor_id": 1
        "fecha": "2025-12-08",
        "observaciones": "Pago parcial Factura #123",
        "items": [
            {"forma_pago": "EFECTIVO", "monto": 10000},
            {"forma_pago": "CHEQUE", "monto": 20000, "cheque_id": 5},
            {"forma_pago": "TRANSFERENCIA", "monto": 5000, "banco": "Santander", "referencia": "OP123456"}
        ]
    }
    """
    import json
    from decimal import Decimal
    from datetime import datetime
    from .models import Cliente, Proveedor, Recibo, ItemRecibo, MovimientoCuentaCorriente, MovimientoCuentaCorrienteProveedor, Cheque, MovimientoCaja
    from .services import AccountingService
    
    try:
        data = json.loads(request.body)
        cliente_id = data.get('cliente_id')
        proveedor_id = data.get('proveedor_id')
        fecha_str = data.get('fecha')
        observaciones = data.get('observaciones', '')
        items_data = data.get('items', [])
        
        if not items_data:
            return JsonResponse({'ok': False, 'error': 'El recibo debe tener al menos una forma de pago'})
        
        # Determinar tipo y entidad
        if cliente_id:
            tipo = 'CLIENTE'
            entidad = Cliente.objects.get(id=cliente_id)
        elif proveedor_id:
            tipo = 'PROVEEDOR'
            entidad = Proveedor.objects.get(id=proveedor_id)
        else:
            return JsonResponse({'ok': False, 'error': 'Debe especificar cliente_id o proveedor_id'})
        
        # Calcular total
        total = sum(Decimal(str(item['monto'])) for item in items_data)
        
        # Obtener siguiente n煤mero de recibo
        ultimo_recibo = Recibo.objects.order_by('-numero').first()
        numero = (ultimo_recibo.numero + 1) if ultimo_recibo else 1
        
        # Parsear fecha
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else datetime.now().date()
        
        # Crear recibo
        recibo = Recibo.objects.create(
            numero=numero,
            tipo=tipo,
            cliente= entidad if tipo == 'CLIENTE' else None,
            proveedor=entidad if tipo == 'PROVEEDOR' else None,
            fecha=fecha,
            total=total,
            observaciones=observaciones
        )
        
        # Crear items del recibo
        for item_data in items_data:
            forma_pago = item_data['forma_pago']
            monto = Decimal(str(item_data['monto']))
            
            item = ItemRecibo.objects.create(
                recibo=recibo,
                forma_pago=forma_pago,
                monto=monto,
                banco=item_data.get('banco', ''),
                referencia=item_data.get('referencia', '')
            )
            
            # Si es pago con cheque, vincular el cheque
            if forma_pago == 'CHEQUE' and item_data.get('cheque_id'):
                cheque = Cheque.objects.get(id=item_data['cheque_id'])
                estado_anterior = cheque.estado
                item.cheque = cheque
                item.save()
                
                # Actualizar estado del cheque
                if tipo == 'CLIENTE':  # Cobro
                    # Al recibir un cheque de un cliente, entra en cartera (no depositado aun)
                    cheque.estado = 'CARTERA'  
                else:  # Pago a proveedor
                    cheque.estado = 'ENTREGADO'
                
                cheque.destinatario = entidad.nombre
                cheque.save()

            # Si es RETENCION, registrar contablemente
            elif forma_pago == 'RETENCION':
                item.retencion_numero = item_data.get('retencion_numero', '')
                item.retencion_tipo = item_data.get('retencion_tipo', '')
                item.save()
            
            # Crear movimiento de caja si es efectivo o transferencia

            if forma_pago in ['EFECTIVO', 'TRANSFERENCIA', 'DEBITO', 'CREDITO']:
                caja_activa = CajaDiaria.objects.filter(estado='ABIERTA').first()
                MovimientoCaja.objects.create(
                    caja_diaria=caja_activa,
                    usuario=request.user,
                    tipo='Ingreso' if tipo == 'CLIENTE' else 'Egreso',
                    descripcion=f"{forma_pago} - Recibo #{recibo.numero_formateado()} - {entidad.nombre}",
                    monto=monto
                )
        
        # Crear movimiento en cuenta corriente
        if tipo == 'CLIENTE':
            nuevo_saldo = entidad.saldo_actual - total  # HABER disminuye deuda
            MovimientoCuentaCorriente.objects.create(
                cliente=entidad,
                tipo='HABER',
                descripcion=f"Recibo de Cobro #{recibo.numero_formateado()}",
                monto=total,
                saldo=nuevo_saldo,
                recibo=recibo
            )
            entidad.saldo_actual = nuevo_saldo
        else:  # PROVEEDOR
            nuevo_saldo = entidad.saldo_actual - total  # DEBE disminuye deuda
            MovimientoCuentaCorrienteProveedor.objects.create(
                proveedor=entidad,
                tipo='DEBE',
                descripcion=f"Recibo de Pago #{recibo.numero_formateado()}",
                monto=total,
                saldo=nuevo_saldo
            )
            entidad.saldo_actual = nuevo_saldo
        
        entidad.save()
        
        #  GENERAR ASIENTO CONTABLE CONSOLIDADO
        try:
            AccountingService.registrar_recibo(recibo)
        except Exception as e:
            print(f"Error generando asiento contable consolidado para recibo {recibo.id}: {e}")

        return JsonResponse({
            'ok': True,
            'recibo_id': recibo.id,
            'numero': recibo.numero_formateado(),
            'total': float(total)
        })
        
    except Cliente.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Cliente no encontrado'})
    except Proveedor.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Proveedor no encontrado'})
    except Cheque.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Cheque no encontrado'})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required
def api_recibos_listar(request):
    """Lista todos los recibos con filtros"""
    from .models import Recibo
    from django.db.models import Q
    
    q = request.GET.get('q', '')
    tipo = request.GET.get('tipo', '')  # CLIENTE o PROVEEDOR
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    recibos = Recibo.objects.select_related('cliente', 'proveedor').order_by('-numero')
    
    if q:
        if tipo == 'CLIENTE':
            recibos = recibos.filter(Q(cliente__nombre__icontains=q) | Q(numero__icontains=q))
        elif tipo == 'PROVEEDOR':
            recibos = recibos.filter(Q(proveedor__nombre__icontains=q) | Q(numero__icontains=q))
        else:
            recibos = recibos.filter(
                Q(cliente__nombre__icontains=q) | 
                Q(proveedor__nombre__icontains=q) | 
                Q(numero__icontains=q)
            )
    
    if tipo:
        recibos = recibos.filter(tipo=tipo)
    
    if desde:
        recibos = recibos.filter(fecha__gte=desde)
    if hasta:
        recibos = recibos.filter(fecha__lte=hasta)
    
    data = []
    for r in recibos[:100]:  # Limitar a 100 resultados
        entidad_nombre = r.cliente.nombre if r.cliente else r.proveedor.nombre
        data.append({
            'id': r.id,
            'numero': r.numero_formateado(),
            'tipo': r.tipo,
            'entidad': entidad_nombre,
            'fecha': r.fecha.strftime('%Y-%m-%d'),
            'total': float(r.total),
            'anulado': r.anulado,
            'observaciones': r.observaciones
        })
    
    return JsonResponse({'ok': True, 'recibos': data})


@login_required
def api_recibo_detalle(request, id):
    """Obtiene el detalle completo de un recibo"""
    from .models import Recibo
    
    try:
        recibo = Recibo.objects.prefetch_related('items').select_related('cliente', 'proveedor').get(id=id)
        
        items = []
        for item in recibo.items.all():
            items.append({
                'forma_pago': item.forma_pago,
                'monto': float(item.monto),
                'banco': item.banco,
                'referencia': item.referencia,
                'cheque': {
                    'numero': item.cheque.numero,
                    'banco': item.cheque.banco
                } if item.cheque else None
            })
        
        entidad = recibo.cliente if recibo.cliente else recibo.proveedor
        
        return JsonResponse({
            'ok': True,
            'recibo': {
                'id': recibo.id,
                'numero': recibo.numero_formateado(),
                'tipo': recibo.tipo,
                'entidad': {
                    'id': entidad.id,
                    'nombre': entidad.nombre,
                    'cuit': entidad.cuit if hasattr(entidad, 'cuit') else ''
                },
                'fecha': recibo.fecha.strftime('%Y-%m-%d'),
                'total': float(recibo.total),
                'observaciones': recibo.observaciones,
                'anulado': recibo.anulado,
                'items': items
            }
        })
    except Recibo.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Recibo no encontrado'})


@login_required
def api_recibo_imprimir(request, id):
    """Vista para imprimir un recibo con selecci贸n de modelo"""
    from .models import Recibo, Empresa
    
    model = request.GET.get('model', 'modern')
    modelos_validos = ['modern', 'minimal', 'classic', 'elegant', 'tech', 'industrial', 'eco', 'compact', 'luxury', 'bold']
    if model not in modelos_validos:
        model = 'modern'
        
    try:
        recibo = Recibo.objects.prefetch_related('items').select_related('cliente', 'proveedor').get(id=id)
        empresa = Empresa.objects.first()
        
        entidad = recibo.cliente if recibo.cliente else recibo.proveedor
        
        context = {
            'recibo': recibo,
            'entidad': entidad,
            'empresa': empresa,
        }
        
        template_name = f'administrar/comprobantes/rec_{model}.html'
        try:
            return render(request, template_name, context)
        except:
            return render(request, 'administrar/comprobantes/rec_modern.html', context)
            
    except Recibo.DoesNotExist:
        return render(request, 'administrar/error.html', {'mensaje': 'Recibo no encontrado'})


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def api_recibo_anular(request, id):
    """Anula un recibo y revierte todos sus efectos"""
    from .models import Recibo, MovimientoCuentaCorriente, MovimientoCuentaCorrienteProveedor
    from datetime import datetime
    
    try:
        recibo = Recibo.objects.get(id=id)
        
        if recibo.anulado:
            return JsonResponse({'ok': False, 'error': 'El recibo ya est谩 anulado'})
        
        # Marcar como anulado
        recibo.anulado = True
        recibo.fecha_anulacion = datetime.now()
        recibo.save()
        
        # Crear movimiento inverso en cuenta corriente
        if recibo.tipo == 'CLIENTE':
            cliente = recibo.cliente
            nuevo_saldo = cliente.saldo_actual + recibo.total  # Revertir: volver a aumentar deuda
            MovimientoCuentaCorriente.objects.create(
                cliente=cliente,
                tipo='DEBE',
                descripcion=f"Anulaci贸n {recibo.numero_formateado()}",
                monto=recibo.total,
                saldo=nuevo_saldo
            )
            cliente.saldo_actual = nuevo_saldo
            cliente.save()
        else:  # PROVEEDOR
            proveedor = recibo.proveedor
            nuevo_saldo = proveedor.saldo_actual + recibo.total  # Revertir: volver a aumentar deuda
            MovimientoCuentaCorrienteProveedor.objects.create(
                proveedor=proveedor,
                tipo='HABER',
                descripcion=f"Anulaci贸n {recibo.numero_formateado()}",
                monto=recibo.total,
                saldo=nuevo_saldo
            )
            proveedor.saldo_actual = nuevo_saldo
            proveedor.save()
        
        # TODO: Revertir estado de cheques y movimientos de caja si es necesario
        
        return JsonResponse({'ok': True, 'mensaje': 'Recibo anulado correctamente'})
        
    except Recibo.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Recibo no encontrado'})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})




# =======================================
#  REPORTES Y ESTADSTICAS
# =======================================


# =======================================
#  REPORTES Y ESTADSTICAS
# =======================================

@login_required
@verificar_permiso('reportes')
def reportes(request):
    """Vista principal de reportes y estad铆sticas"""
    return render(request, "administrar/reportes.html")


@login_required
@verificar_permiso('reportes')
def api_estadisticas_ventas(request):
    """API para obtener estad铆sticas de ventas"""
    from django.db.models import Sum, Count, Avg
    from datetime import datetime, timedelta
    from collections import defaultdict
    
    # Par谩metros de filtro
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    
    # Query base
    ventas = Venta.objects.all()
    
    # Aplicar filtros de fecha
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            ventas = ventas.filter(fecha__date__gte=fecha_desde)
        except ValueError:
            pass
    
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            ventas = ventas.filter(fecha__date__lte=fecha_hasta)
        except ValueError:
            pass
    
    # Si no hay filtros especific set default range (煤ltimos 12 meses)
    if not fecha_desde_str and not fecha_hasta_str:
        hace_12_meses = datetime.now() - timedelta(days=365)
        ventas = ventas.filter(fecha__gte=hace_12_meses)
    
    # Estad铆sticas generales
    total_ventas = ventas.aggregate(total=Sum('total'))['total'] or 0
    cantidad_ventas = ventas.count()
    promedio_venta = ventas.aggregate(promedio=Avg('total'))['promedio'] or 0
    
    # Ventas por mes
    ventas_por_mes = defaultdict(float)
    for venta in ventas:
        mes_key = venta.fecha.strftime('%Y-%m')
        ventas_por_mes[mes_key] += float(venta.total)
    
    # Ordenar por fecha
    ventas_por_mes_lista = [
        {'mes': mes, 'total': total}
        for mes, total in sorted(ventas_por_mes.items())
    ]
    
    # Top 5 clientes
    top_clientes = ventas.values('cliente__nombre').annotate(
        total=Sum('total'),
        cantidad=Count('id')
    ).order_by('-total')[:5]
    
    top_clientes_lista = [
        {
            'nombre': item['cliente__nombre'],
            'total': float(item['total']),
            'cantidad': item['cantidad']
        }
        for item in top_clientes
    ]
    
    # Top 5 productos vendidos
    detalles = DetalleVenta.objects.filter(venta__in=ventas).values(
        'producto__descripcion'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        importe_total=Sum('subtotal')
    ).order_by('-cantidad_total')[:5]
    
    top_productos_lista = [
        {
            'nombre': item['producto__descripcion'],
            'cantidad': float(item['cantidad_total']),
            'importe': float(item['importe_total'])
        }
        for item in detalles
    ]
    
    # Distribuci贸n por m茅todo de pago
    # NOTA: El modelo Venta no tiene campo metodo_pago, por lo que devolvemos lista vac铆a
    # Si en el futuro se agrega este campo, descomentar el c贸digo siguiente:
    # metodos_pago = ventas.values('metodo_pago').annotate(
    #     total=Sum('total'),
    #     cantidad=Count('id')
    # ).order_by('-total')
    
    metodos_pago_lista = []
    # metodos_pago_lista = [
    #     {
    #         'metodo': item['metodo_pago'] or 'No especificado',
    #         'total': float(item['total']),
    #         'cantidad': item['cantidad']
    #     }
    #     for item in metodos_pago
    # ]
    
    return JsonResponse({
        'ok': True,
        'total_ventas': float(total_ventas),
        'cantidad_ventas': cantidad_ventas,
        'promedio_venta': float(promedio_venta),
        'ventas_por_mes': ventas_por_mes_lista,
        'top_clientes': top_clientes_lista,
        'top_productos': top_productos_lista,
        'metodos_pago': metodos_pago_lista,
    })


@login_required
@verificar_permiso('reportes')
def api_estadisticas_compras(request):
    """API para obtener estad铆sticas de compras"""
    from django.db.models import Sum, Count, Avg
    from datetime import datetime, timedelta
    from collections import defaultdict
    
    # Par谩metros de filtro
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    
    # Query base
    compras = Compra.objects.all()
    
    # Aplicar filtros de fecha
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            compras = compras.filter(fecha__date__gte=fecha_desde)
        except ValueError:
            pass
    
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            compras = compras.filter(fecha__date__lte=fecha_hasta)
        except ValueError:
            pass
    
    # Estad铆sticas generales
    total_compras = compras.aggregate(total=Sum('total'))['total'] or 0
    cantidad_compras = compras.count()
    promedio_compra = compras.aggregate(promedio=Avg('total'))['promedio'] or 0
    
    # Compras por mes
    compras_por_mes = defaultdict(float)
    for compra in compras:
        mes_key = compra.fecha.strftime('%Y-%m')
        compras_por_mes[mes_key] += float(compra.total)
    
    compras_por_mes_lista = [
        {'mes': mes, 'total': total}
        for mes, total in sorted(compras_por_mes.items())
    ]
   
    # Top 5 proveedores
    top_proveedores = compras.values('proveedor__nombre').annotate(
        total=Sum('total'),
        cantidad=Count('id')
    ).order_by('-total')[:5]
    
    top_proveedores_lista = [
        {
            'nombre': item['proveedor__nombre'] or 'Sin proveedor',
            'total': float(item['total']),
            'cantidad': item['cantidad']
        }
        for item in top_proveedores
    ]
    
    # Top 5 productos comprados
    detalles = DetalleCompra.objects.filter(compra__in=compras).values(
        'producto__descripcion'
    ).annotate(
        cantidad_total=Sum('cantidad'),
        importe_total=Sum('subtotal')
    ).order_by('-cantidad_total')[:5]
    
    top_productos_lista = [
        {
            'nombre': item['producto__descripcion'],
            'cantidad': float(item['cantidad_total']),
            'importe': float(item['importe_total'])
        }
        for item in detalles
    ]
    
    return JsonResponse({
        'ok': True,
        'total_compras': float(total_compras),
        'cantidad_compras': cantidad_compras,
        'promedio_compra': float(promedio_compra),
        'compras_por_mes': compras_por_mes_lista,
        'top_proveedores': top_proveedores_lista,
        'top_productos': top_productos_lista,
    })


@login_required
@verificar_permiso('reportes')
def api_estadisticas_stock(request):
    """API para obtener estad铆sticas de inventario"""
    from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField
    
    # Productos con stock bajo (menor al stock m铆nimo)
    productos_stock_bajo = Producto.objects.filter(
        stock__lt=F('stock_minimo')
    ).values('codigo', 'descripcion', 'stock', 'stock_minimo')
    
    stock_bajo_lista = [
        {
            'codigo': p['codigo'],
            'descripcion': p['descripcion'],
            'stock_actual': float(p['stock']),
            'stock_minimo': float(p['stock_minimo'])
        }
        for p in productos_stock_bajo[:10]  # Limitar a 10
    ]
    
    # Valorizaci贸n total del inventario (stock * costo)
    total_productos = Producto.objects.count()
    
    # Calcular valorizaci贸n
    valorizacion = Producto.objects.aggregate(
        total=Sum(ExpressionWrapper(
            F('stock') * F('costo'),
            output_field=DecimalField()
        ))
    )['total'] or 0
    
    # Distribuci贸n por rubro
    por_rubro = Producto.objects.values('rubro__nombre').annotate(
        cantidad=Count('id'),
        stock_total=Sum('stock'),
        valor=Sum(ExpressionWrapper(
            F('stock') * F('costo'),
            output_field=DecimalField()
        ))
    ).order_by('-cantidad')[:10]
    
    rubros_lista = [
        {
            'rubro': r['rubro__nombre'] or 'Sin rubro',
            'cantidad_productos': r['cantidad'],
            'stock_total': float(r['stock_total'] or 0),
            'valorizaci贸n': float(r['valor'] or 0)
        }
        for r in por_rubro
    ]
    
    # Distribuci贸n por marca
    por_marca = Producto.objects.values('marca__nombre').annotate(
        cantidad=Count('id'),
        stock_total=Sum('stock')
    ).order_by('-cantidad')[:10]
    
    marcas_lista = [
        {
            'marca': m['marca__nombre'] or 'Sin marca',
            'cantidad_productos': m['cantidad'],
            'stock_total': float(m['stock_total'] or 0)
        }
        for m in por_marca
    ]
    
    return JsonResponse({
        'ok': True,
        'total_productos': total_productos,
        'valorizacion_total': float(valorizacion),
        'productos_stock_bajo': stock_bajo_lista,
        'cantidad_stock_bajo': len(stock_bajo_lista),
        'distribucion_rubros': rubros_lista,
        'distribucion_marcas': marcas_lista,
    })


@login_required
@verificar_permiso('reportes')
def api_estadisticas_caja(request):
    """API para obtener estad铆sticas de caja"""
    from django.db.models import Sum
    from datetime import datetime, timedelta
    from collections import defaultdict
    
    # Par谩metros de filtro
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    
    # Query base
    movimientos = MovimientoCaja.objects.all()
   
    # Aplicar filtros de fecha
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha__date__gte=fecha_desde)
        except ValueError:
            pass
    
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha__date__lte=fecha_hasta)
        except ValueError:
            pass
    
    # Default: 煤ltimos 12 meses
    if not fecha_desde_str and not fecha_hasta_str:
        hace_12_meses = datetime.now() - timedelta(days=365)
        movimientos = movimientos.filter(fecha__gte=hace_12_meses)
    
    # Totales de ingresos y egresos
    ingresos_total = movimientos.filter(tipo='Ingreso').aggregate(total=Sum('monto'))['total'] or 0
    egresos_total = movimientos.filter(tipo='Egreso').aggregate(total=Sum('monto'))['total'] or 0
    balance = float(ingresos_total) - float(egresos_total)
    
    # Ingresos y egresos por mes
    ingresos_por_mes = defaultdict(float)
    egresos_por_mes = defaultdict(float)
    
    for mov in movimientos:
        mes_key = mov.fecha.strftime('%Y-%m')
        if mov.tipo == 'Ingreso':
            ingresos_por_mes[mes_key] += float(mov.monto)
        else:
            egresos_por_mes[mes_key] += float(mov.monto)
    
    # Combinar en una lista ordenada
    todos_meses = set(ingresos_por_mes.keys()) | set(egresos_por_mes.keys())
    movimientos_por_mes_lista = [
        {
            'mes': mes,
            'ingresos': ingresos_por_mes.get(mes, 0),
            'egresos': egresos_por_mes.get(mes, 0),
            'balance': ingresos_por_mes.get(mes, 0) - egresos_por_mes.get(mes, 0)
        }
        for mes in sorted(todos_meses)
    ]
    
    # Promedio diario (si hay filtro de fechas)
    dias_total = 30  # Default
    if fecha_desde_str and fecha_hasta_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            dias_total = (fecha_hasta - fecha_desde).days + 1
        except:
            pass
    
    promedio_diario_ingresos = float(ingresos_total) / dias_total if dias_total > 0 else 0
    promedio_diario_egresos = float(egresos_total) / dias_total if dias_total > 0 else 0
    
    return JsonResponse({
        'ok': True,
        'ingresos_total': float(ingresos_total),
        'egresos_total': float(egresos_total),
        'balance': balance,
        'movimientos_por_mes': movimientos_por_mes_lista,
        'promedio_diario_ingresos': promedio_diario_ingresos,
        'promedio_diario_egresos': promedio_diario_egresos,
        'cantidad_movimientos': movimientos.count(),
    })


@login_required
@verificar_permiso('reportes')
def api_reportes_exportar(request):
    """API para exportar un reporte integral a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        from django.db.models import Sum, Count, Avg, F, ExpressionWrapper, DecimalField
        from datetime import datetime
        
        # Crear Workbook
        wb = openpyxl.Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center')
        currency_format = '"$"#,##0.00'
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        def agregar_hoja(nombre, headers, data, column_widths=None):
            ws = wb.create_sheet(title=nombre)
            # Headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Data
            for row_num, row_data in enumerate(data, 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    # Auto-format currency
                    if isinstance(value, (int, float)) and col_num > 1: # Asumiendo col 1 es texto/fecha
                         if 'Monto' in headers[col_num-1] or 'Total' in headers[col_num-1] or 'Precio' in headers[col_num-1] or 'Valor' in headers[col_num-1] or 'Costo' in headers[col_num-1]:
                            cell.number_format = currency_format

            # Column Widths
            if column_widths:
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Eliminar hoja default
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Filtros
        fecha_desde_str = request.GET.get('fecha_desde', '')
        fecha_hasta_str = request.GET.get('fecha_hasta', '')
        
        fecha_desde = None
        fecha_hasta = None
        
        if fecha_desde_str:
            try: fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            except: pass
        if fecha_hasta_str:
            try: fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            except: pass

        # ==========================
        # 1. VENTAS
        # ==========================
        ventas = Venta.objects.all().select_related('cliente')
        if fecha_desde: ventas = ventas.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta: ventas = ventas.filter(fecha__date__lte=fecha_hasta)
        
        data_ventas = []
        for v in ventas:
            cliente_nombre = v.cliente.nombre if v.cliente else "Cliente Eliminado"
            data_ventas.append([
                v.id,
                v.fecha.strftime('%d/%m/%Y'),
                cliente_nombre,
                v.tipo_comprobante,
                v.estado,
                float(v.total)
            ])
        
        agregar_hoja(
            "Ventas", 
            ['ID', 'Fecha', 'Cliente', 'Tipo', 'Estado', 'Monto Total'], 
            data_ventas,
            [10, 15, 40, 10, 15, 20]
        )

        # ==========================
        # 2. COMPRAS
        # ==========================
        compras = Compra.objects.all().select_related('proveedor')
        if fecha_desde: compras = compras.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta: compras = compras.filter(fecha__date__lte=fecha_hasta)
        
        data_compras = []
        for c in compras:
            prov = c.proveedor.nombre if c.proveedor else 'Sin Proveedor'
            data_compras.append([
                c.id,
                c.fecha.strftime('%d/%m/%Y'),
                prov,
                c.nro_comprobante,
                float(c.total)
            ])
            
        agregar_hoja(
            "Compras", 
            ['ID', 'Fecha', 'Proveedor', 'Nro Factura', 'Monto Total'], 
            data_compras,
            [10, 15, 40, 20, 20]
        )

        # ==========================
        # 3. CAJA
        # ==========================
        movimientos = MovimientoCaja.objects.all()
        if fecha_desde: movimientos = movimientos.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta: movimientos = movimientos.filter(fecha__date__lte=fecha_hasta)
        
        data_caja = []
        for m in movimientos:
            data_caja.append([
                m.id,
                m.fecha.strftime('%d/%m/%Y %H:%M'),
                m.tipo,
                m.descripcion,
                float(m.monto)
            ])
            
        agregar_hoja(
            "Caja", 
            ['ID', 'Fecha/Hora', 'Tipo', 'Descripci贸n', 'Monto'], 
            data_caja,
            [10, 20, 15, 50, 20]
        )

        # ==========================
        # 4. STOCK (Snapshot actual)
        # ==========================
        # Stock no filtra por fecha, es el estado actual
        productos = Producto.objects.all().select_related('rubro', 'marca')
        
        data_stock = []
        for p in productos:
            rubro = p.rubro.nombre if p.rubro else '-'
            marca = p.marca.nombre if p.marca else '-'
            
            # Validar valores num茅ricos
            stock = p.stock if p.stock is not None else 0
            costo = p.costo if p.costo is not None else 0
            precio = p.precio_efectivo if p.precio_efectivo is not None else 0
            
            valor_total = stock * costo
            
            data_stock.append([
                p.codigo,
                p.descripcion,
                rubro,
                marca,
                float(stock),
                float(costo),
                float(precio),
                float(valor_total)
            ])
            
        agregar_hoja(
            "Stock Actual", 
            ['C贸digo', 'Descripci贸n', 'Rubro', 'Marca', 'Stock', 'Costo Unit.', 'Precio Venta', 'Valor Total'], 
            data_stock,
            [15, 40, 20, 20, 10, 15, 15, 20]
        )

        # Generar respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Reporte_General_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error exportando Excel: {str(e)}'}, status=500)


# ==========================================
# GESTIN BANCARIA
# ==========================================

@login_required
@verificar_permiso('contabilidad') # Asumiendo permiso contabilidad o crear uno nuevo 'bancos'
def bancos(request):
    return render(request, 'administrar/bancos.html')

@login_required
def api_cheques_listar(request):
    from .models import Cheque
    from django.db.models import Sum, Q
    from datetime import datetime, date

    try:
        # Params
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        busqueda = request.GET.get('busqueda', '').strip()
        tipo = request.GET.get('tipo', '').strip()
        estado = request.GET.get('estado', '').strip()
        
        # Base Query
        cheques = Cheque.objects.all().order_by('-fecha_pago')

        # Filters
        if busqueda:
            cheques = cheques.filter(Q(numero__icontains=busqueda) | Q(banco__icontains=busqueda) | Q(cliente__nombre__icontains=busqueda) | Q(firmante__icontains=busqueda))
        
        if tipo and tipo != 'TODOS':
            cheques = cheques.filter(tipo=tipo)
            
        if estado and estado != 'TODOS':
            cheques = cheques.filter(estado=estado)

        # KPIs
        kpi_cartera_terceros_qs = Cheque.objects.filter(estado='CARTERA', tipo='TERCERO')
        kpi_cartera_terceros_total = kpi_cartera_terceros_qs.aggregate(t=Sum('monto'))['t'] or 0
        kpi_cartera_terceros_count = kpi_cartera_terceros_qs.count()

        kpi_propios_qs = Cheque.objects.filter(tipo='PROPIO', estado__in=['CARTERA', 'ENTREGADO']) 
        kpi_propios_total = kpi_propios_qs.aggregate(t=Sum('monto'))['t'] or 0
        
        today = date.today()
        # Filtramos depositados en el mes actual (aprox)
        kpi_depositados_qs = Cheque.objects.filter(estado='DEPOSITADO', fecha_modificacion__month=today.month, fecha_modificacion__year=today.year)
        kpi_depositados_total = kpi_depositados_qs.aggregate(t=Sum('monto'))['t'] or 0

        kpi_rechazados_qs = Cheque.objects.filter(estado='RECHAZADO')
        kpi_rechazados_total = kpi_rechazados_qs.aggregate(t=Sum('monto'))['t'] or 0

        # Pagination
        total = cheques.count()
        cheques_page = cheques[(page - 1) * per_page : page * per_page]

        cheques_qs = cheques
        paginator = Paginator(cheques_qs, per_page)
        cheques_page = paginator.get_page(page)

        data = []
        for c in cheques_page:
            data.append({
                'id': c.id,
                'fecha_pago': c.fecha_pago.strftime('%d/%m/%Y') if c.fecha_pago else '-',
                'banco': c.banco,
                'numero': c.numero,
                'origen_destino': c.cliente.nombre if c.cliente else (c.destinatario or c.firmante or '-'),
                'monto': float(c.monto),
                'tipo': c.tipo,
                'estado': c.estado,
            })

        print(f"DEBUG: Found {total} cheques. Page has {len(data)} items.")

        return JsonResponse({
            'ok': True,
            'data': data,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page,
            'kpis': {
                'cartera_terceros': {'total': kpi_cartera_terceros_total, 'count': kpi_cartera_terceros_count},
                'apagar_propios': {'total': kpi_propios_total},
                'depositados_mes': {'total': kpi_depositados_total},
                'rechazados': {'total': kpi_rechazados_total}
            }
        })

    except Exception as e:
        print(f"Error api_cheques_listar: {e}")
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@login_required
def api_cheque_cambiar_estado(request, id):
    from .models import Cheque, CuentaBancaria
    from .services import AccountingService
    import json
    
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'M茅todo no permitido'}, status=405)
        
    try:
        cheque = Cheque.objects.get(id=id)
        data = json.loads(request.body)
        nuevo_estado = data.get('estado')
        cuenta_destino_id = data.get('cuenta_destino_id') # Optional, for deposits
        
        valid_states = ['CARTERA', 'DEPOSITADO', 'COBRADO', 'ENTREGADO', 'RECHAZADO', 'ANULADO']
        
        if nuevo_estado not in valid_states:
            return JsonResponse({'ok': False, 'error': 'Estado inv谩lido'}, status=400)
            
        estado_anterior = cheque.estado
        
        # Validation for Deposit
        cuenta_destino = None
        if nuevo_estado == 'DEPOSITADO' and estado_anterior == 'CARTERA':
            if not cuenta_destino_id:
                return JsonResponse({'ok': False, 'error': 'Debe seleccionar una cuenta bancaria para depositar'}, status=400)
            try:
                cuenta_destino = CuentaBancaria.objects.get(id=cuenta_destino_id)
            except CuentaBancaria.DoesNotExist:
                return JsonResponse({'ok': False, 'error': 'Cuenta bancaria no encontrada'}, status=404)

        cheque.estado = nuevo_estado
        cheque.save()
        
        # Generate Accounting Entry
        try:
            AccountingService.registrar_cambio_estado(cheque, estado_anterior, cuenta_destino)
        except Exception as e:
            print(f"Error generando asiento contable para cheque {cheque.id}: {e}")
            # Non-blocking error, but worth logging/notifying
        
        return JsonResponse({'ok': True, 'message': f'Estado actualizado a {nuevo_estado}'})
        
    except Cheque.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Cheque no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@login_required
def api_bancos_listar(request):
    from .models import CuentaBancaria
    cuentas = CuentaBancaria.objects.filter(activo=True)
    data = []
    for c in cuentas:
        data.append({
            'id': c.id,
            'banco': c.banco,
            'cbu': c.cbu,
            'alias': c.alias,
            'moneda': c.moneda,
            'saldo': float(c.saldo_actual)
        })
    return JsonResponse({'ok': True, 'cuentas': data})

@require_POST
@csrf_exempt
@login_required
def api_bancos_crear(request):
    import json
    from .models import CuentaBancaria
    try:
        data = json.loads(request.body)
        c = CuentaBancaria.objects.create(
            banco=data['banco'],
            cbu=data.get('cbu',''),
            alias=data.get('alias',''),
            moneda=data.get('moneda','ARS'),
            saldo_inicial=data.get('saldo_inicial',0),
            saldo_actual=data.get('saldo_inicial',0)
        )
        return JsonResponse({'ok': True, 'id': c.id})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@login_required
def api_bancos_movimientos(request):
    from .models import MovimientoBanco, CuentaBancaria
    
    cuenta_id = request.GET.get('cuenta_id')
    conciliado = request.GET.get('conciliado') # 'si', 'no', 'todos'
    
    if not cuenta_id:
        return JsonResponse({'ok': False, 'error': 'Falta cuenta_id'})
        
    movs = MovimientoBanco.objects.filter(cuenta_id=cuenta_id).order_by('-fecha', '-id')
    
    if conciliado == 'si':
        movs = movs.filter(conciliado=True)
    elif conciliado == 'no':
        movs = movs.filter(conciliado=False)
        
    data = []
    for m in movs:
        data.append({
            'id': m.id,
            'fecha': m.fecha.strftime('%d/%m/%Y'),
            'descripcion': m.descripcion,
            'monto': float(m.monto),
            'conciliado': m.conciliado,
            'referencia': m.referencia_interna
        })
        
    return JsonResponse({'ok': True, 'movimientos': data})

@require_POST
@csrf_exempt
@login_required
def api_bancos_movimiento_crear(request):
    import json
    from .models import MovimientoBanco, CuentaBancaria
    try:
        data = json.loads(request.body)
        cuenta = CuentaBancaria.objects.get(id=data['cuenta_id'])
        
        monto = float(data['monto'])
        if data['tipo'] == 'DEBITO':
            monto = -abs(monto)
        else:
            monto = abs(monto)
            
        m = MovimientoBanco.objects.create(
            cuenta=cuenta,
            fecha=data['fecha'],
            descripcion=data['descripcion'],
            monto=monto,
            referencia_interna=data.get('referencia','')
        )
        
        # Actualizar saldo cuenta
        cuenta.saldo_actual = float(cuenta.saldo_actual) + monto
        cuenta.save()
        
        # Registrar en contabilidad
        try:
            from .services import AccountingService
            # Agregamos tipo 'DEBITO'/'CREDITO' temporalmente al objeto para que el servicio sepa
            m.tipo = data['tipo'] 
            AccountingService.registrar_movimiento_banco(m)
        except Exception as e:
            print(f"Error contabilidad: {e}")
        
        return JsonResponse({'ok': True, 'id': m.id})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@require_POST
@csrf_exempt
@login_required
def api_bancos_conciliar(request):
    import json
    from .models import MovimientoBanco
    try:
        data = json.loads(request.body)
        movimiento_id = data['id']
        conciliado = data['conciliado'] # True/False
        
        m = MovimientoBanco.objects.get(id=movimiento_id)
        m.conciliado = conciliado
        if conciliado:
            from datetime import date
            m.fecha_conciliado = date.today()
        else:
            m.fecha_conciliado = None
        m.save()
        
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@csrf_exempt


# =======================================
#  PRESUPUESTOS (COTIZACIONES)
# =======================================

@login_required
@verificar_permiso('ventas')
def presupuestos_lista(request):
    return render(request, 'administrar/presupuestos.html')

@login_required
@verificar_permiso('ventas')
def presupuesto_nuevo(request):
    return render(request, 'administrar/presupuesto_nuevo.html', {
        'items_json': '[]',
        'cliente_json': 'null'
    })

@login_required
def api_presupuestos_listar(request):
    try:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        q = request.GET.get('q', '').strip()
        estado = request.GET.get('estado', '')

        presupuestos = Presupuesto.objects.select_related('cliente').order_by('-fecha')

        if q:
            presupuestos = presupuestos.filter(cliente__nombre__icontains=q)
        if estado:
            presupuestos = presupuestos.filter(estado=estado)

        total = presupuestos.count()
        start = (page - 1) * per_page
        end = start + per_page
        
        data = []
        for p in presupuestos[start:end]:
            # Calcular vencimiento
            fecha_venc = p.fecha.date() + datetime.timedelta(days=p.validez)
            is_vencido = fecha_venc < datetime.date.today() and p.estado == 'PENDIENTE'
            
            estado_real = 'VENCIDO' if is_vencido else p.estado

            data.append({
                'id': p.id,
                'fecha': p.fecha.strftime('%d/%m/%Y'),
                'cliente': p.cliente.nombre,
                'vencimiento': fecha_venc.strftime('%d/%m/%Y'),
                'total': float(p.total),
                'estado': estado_real
            })

        return JsonResponse({'ok': True, 'data': data, 'total': total})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@csrf_exempt
@require_POST
@login_required
@transaction.atomic
def api_presupuesto_guardar(request):
    try:
        data = json.loads(request.body)
        presupuesto_id = data.get('id')
        cliente_id = data.get('cliente_id')
        items = data.get('items', [])
        total = float(data.get('total', 0))
        validez = int(data.get('validez', 15))
        observaciones = data.get('observaciones', '')

        if not items:
            return JsonResponse({'ok': False, 'error': 'Sin 铆tems'})

        # Cliente
        if cliente_id:
            cliente = Cliente.objects.get(id=cliente_id)
        else:
            cliente = Cliente.objects.get_or_create(nombre="Consumidor Final", defaults={'condicion_fiscal': 'CF'})[0]

        if presupuesto_id:
            # EDICION
            presupuesto = Presupuesto.objects.get(pk=presupuesto_id)
            if presupuesto.estado != 'PENDIENTE':
                return JsonResponse({'ok': False, 'error': 'Solo se pueden editar presupuestos PENDIENTES'})
            
            presupuesto.cliente = cliente
            presupuesto.total = total
            presupuesto.validez = validez
            presupuesto.observaciones = observaciones
            presupuesto.save()
            
            # Borrar items viejos
            presupuesto.detalles.all().delete()
        else:
            # CREACION
            presupuesto = Presupuesto.objects.create(
                cliente=cliente,
                total=total,
                validez=validez,
                observaciones=observaciones,
                estado='PENDIENTE'
            )

        for item in items:
            DetallePresupuesto.objects.create(
                presupuesto=presupuesto,
                producto_id=item['id'],
                descripcion_producto=item['descripcion'], # Asegurar que esto venga del JS
                cantidad=item['cantidad'],
                precio_unitario=item['precio'],
                subtotal=item['subtotal']
            )

        return JsonResponse({'ok': True, 'presupuesto_id': presupuesto.id})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@csrf_exempt
@require_POST
@login_required
@transaction.atomic
def api_presupuesto_cancelar(request, id):
    try:
        p = Presupuesto.objects.get(id=id)
        if p.estado != 'PENDIENTE':
            return JsonResponse({'ok': False, 'error': 'Solo se pueden cancelar presupuestos pendientes'})
        
        p.estado = 'CANCELADO'
        p.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})

@login_required
def api_presupuesto_detalle(request, id):
    """API para obtener detalle de un presupuesto"""
    try:
        p = Presupuesto.objects.select_related('cliente', 'venta').prefetch_related('detalles__producto').get(id=id)
    except Presupuesto.DoesNotExist:
        return JsonResponse({'error': 'Presupuesto no encontrado'}, status=404)
    
    # Serializar detalles
    detalles = []
    for d in p.detalles.all():
        detalles.append({
            'id': d.id,
            'producto_id': d.producto.id,
            'producto_codigo': d.producto.codigo,
            'producto_descripcion': d.producto.descripcion,
            'cantidad': float(d.cantidad),
            'precio_unitario': float(d.precio_unitario),
            'subtotal': float(d.subtotal),
        })
    
    return JsonResponse({
        'id': p.id,
        'fecha': p.fecha.strftime('%d/%m/%Y'),
        'validez': p.validez,
        'vencimiento': (p.fecha + datetime.timedelta(days=p.validez)).strftime('%d/%m/%Y'),
        'cliente_id': p.cliente.id,
        'cliente_nombre': p.cliente.nombre,
        'cliente_cuit': p.cliente.cuit or '',
        'cliente_telefono': p.cliente.telefono or '',
        'cliente_email': p.cliente.email or '',
        'estado': p.estado,
        'total': float(p.total),
        'observaciones': p.observaciones or '',
        'venta_id': p.venta_id if p.venta else None,
        'pedido_id': p.pedido_id if hasattr(p, 'pedido') and p.pedido else None, # Safely access pedido
        'detalles': detalles,
    })

@csrf_exempt
@require_POST
@login_required
@transaction.atomic
@csrf_exempt
@require_POST
@login_required
@transaction.atomic
def api_presupuesto_convertir_a_pedido(request, id):
    """
    Convierte un Presupuesto en PEDIDO.
    1. NO Verifica Stock estricto (permite backorder).
    2. Crea Pedido.
    3. NO Descuenta Stock (se descuenta al facturar/remitear el pedido).
    """
    try:
        p = Presupuesto.objects.get(id=id)
        if p.estado != 'PENDIENTE':
            return JsonResponse({'ok': False, 'error': f'El presupuesto est谩 {p.estado}'})

        # 1. Crear Pedido (Sin validar stock bloqueante)
        pedido = Pedido.objects.create(
            cliente=p.cliente,
            total=p.total,
            estado='PENDIENTE',
            observaciones=f"Generado desde Presupuesto #{p.id}"
        )

        # 2. Detalles del Pedido
        for det in p.detalles.all():
            DetallePedido.objects.create(
                pedido=pedido,
                producto=det.producto,
                cantidad=det.cantidad,
                precio_unitario=det.precio_unitario, # Precio congelado del presupuesto
                subtotal=det.subtotal
            )

        # 3. Actualizar Presupuesto
        p.estado = 'APROBADO'
        p.pedido = pedido
        p.save()

        return JsonResponse({'ok': True, 'pedido_id': pedido.id, 'message': f'Pedido #{pedido.id} creado correctamente'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'ok': False, 'error': str(e)})

@login_required
def presupuesto_pdf(request, id):
    # Stub para PDF (Podemos reutilizar la l贸gica de Factura PDF si existe)
    # Por ahora devolvemos un mensaje o un HTML simple
    p = get_object_or_404(Presupuesto, id=id)
    # Renderizamos mismo template de factura pero con t铆tulo Presupuesto?
    # Para simplicidad, uso un template stub
    return HttpResponse(f"Generaci贸n de PDF para Presupuesto #{p.id} pendiente de implementaci贸n visual.", content_type="text/plain")


# API - Provincias y Localidades
@login_required
def api_provincias_listar(request):
    provincias = Provincia.objects.all().order_by('nombre')
    data = [{'id': p.id, 'nombre': p.nombre} for p in provincias]
    return JsonResponse(data, safe=False)

@login_required
def api_localidades_listar(request):
    localidades = Localidad.objects.all().order_by('nombre')
    data = [{'id': l.id, 'nombre': l.nombre, 'cp': l.codigo_postal} for l in localidades]
    return JsonResponse(data, safe=False)


# =========================================
# API REMITOS (Nuevo)
# =========================================
@login_required
def api_remitos_listar(request):
    """API para listar remitos"""
    try:
        remitos = Remito.objects.select_related('cliente', 'venta_asociada').all().order_by('-fecha')
        
        data = []
        for r in remitos:
            data.append({
                'id': r.id,
                'numero': r.numero_formateado(),
                'fecha': r.fecha.strftime('%d/%m/%Y %H:%M'),
                'cliente': r.cliente.nombre,
                'venta_id': r.venta_asociada.id if r.venta_asociada else None,
                'venta_str': r.venta_asociada.numero_factura_formateado() if r.venta_asociada else '-',
                'estado': r.estado,
                'direccion': r.direccion_entrega,
                'item_count': r.detalles.count()
            })
            
        return JsonResponse({'remitos': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# =========================================
# API NOTAS DE CRDITO
# =========================================
@login_required
def api_notas_credito_listar(request):
    """API para listar notas de cr茅dito"""
    try:
        ncs = NotaCredito.objects.select_related('cliente', 'venta_asociada').all().order_by('-fecha')
        
        data = []
        for nc in ncs:
            data.append({
                'id': nc.id,
                'numero': nc.numero_formateado(),
                'fecha': nc.fecha.strftime('%d/%m/%Y %H:%M'),
                'cliente': nc.cliente.nombre,
                'venta_id': nc.venta_asociada.id if nc.venta_asociada else None,
                'venta_str': nc.venta_asociada.numero_factura_formateado() if nc.venta_asociada else '-',
                'total': float(nc.total),
                'estado': nc.estado,
                'tipo': nc.tipo_comprobante
            })
            
        return JsonResponse({'notas_credito': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# =========================================
# API CONFIGURACIN EMPRESA
# =========================================
@login_required
def api_empresa_config(request):
    """Obtiene configuraci贸n p煤blica de la empresa"""
    try:
        empresa = Empresa.objects.first()
        if not empresa:
            return JsonResponse({'error': 'Empresa no configurada'}, status=404)
        
        return JsonResponse({
            'nombre': empresa.nombre,
            'cuit': empresa.cuit,
            'direccion': empresa.direccion,
            'localidad': empresa.localidad,
            'provincia': empresa.provincia,
            'condicion_fiscal': empresa.condicion_fiscal,
            'iibb': empresa.iibb,
            'inicio_actividades': empresa.inicio_actividades.strftime('%Y-%m-%d') if empresa.inicio_actividades else '',
            'moneda_predeterminada': empresa.moneda_predeterminada,
            'items_por_pagina': empresa.items_por_pagina,
            'habilita_remitos': empresa.habilita_remitos,
            'logo': empresa.logo.url if empresa.logo else None,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@csrf_exempt
@require_POST
def api_empresa_config_guardar(request):
    """Guarda configuraci贸n global (soporta multipart para logo)"""
    try:
        # Si es multipart/form-data (subida de logo) usamos request.POST, si no request.body (JSON)
        if request.content_type.startswith('multipart/form-data'):
            data = request.POST
        else:
            data = json.loads(request.body)

        empresa = Empresa.objects.first()
        if not empresa:
            empresa = Empresa.objects.create(nombre="Mi Empresa") # Fallback
            
        # UI & Sistema
        if 'items_por_pagina' in data:
            empresa.items_por_pagina = int(data['items_por_pagina'])
        if 'habilita_remitos' in data:
            # En multipart los booleanos llegan como strings
            val = data['habilita_remitos']
            empresa.habilita_remitos = str(val).lower() == 'true' if isinstance(val, str) else bool(val)

        # Datos Empresa
        empresa.nombre = data.get('nombre', empresa.nombre)
        empresa.cuit = data.get('cuit', empresa.cuit)
        empresa.direccion = data.get('direccion', empresa.direccion)
        empresa.localidad = data.get('localidad', empresa.localidad)
        empresa.provincia = data.get('provincia', empresa.provincia)
        empresa.condicion_fiscal = data.get('condicion_fiscal', empresa.condicion_fiscal)
        empresa.iibb = data.get('iibb', empresa.iibb)
        empresa.moneda_predeterminada = data.get('moneda_predeterminada', empresa.moneda_predeterminada)
        
        if 'inicio_actividades' in data and data['inicio_actividades']:
            empresa.inicio_actividades = data['inicio_actividades']

        # Manejo de Logo
        if 'logo' in request.FILES:
            empresa.logo = request.FILES['logo']

        empresa.save()
        return JsonResponse({
            'ok': True,
            'logo_url': empresa.logo.url if empresa.logo else None
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ==========================================
# API REMITOS Y NOTAS DE CRDITO (PAGINADOS)
# ==========================================

@login_required
def api_remitos_listar(request):
    try:
        from django.core.paginator import Paginator
        
        page_number = request.GET.get('page', 1)
        per_page = request.GET.get('per_page', 10)
        q = request.GET.get('q', '')
        fecha = request.GET.get('fecha', '')

        remitos = Remito.objects.all().order_by('-fecha', '-id')

        if q:
            remitos = remitos.filter(
                Q(cliente__nombre__icontains=q) | 
                Q(numero__icontains=q)
            )

        if fecha:
            remitos = remitos.filter(fecha=fecha)

        paginator = Paginator(remitos, per_page)
        page_obj = paginator.get_page(page_number)

        data = []
        for r in page_obj:
            data.append({
                'id': r.id,
                'fecha': r.fecha.strftime('%d/%m/%Y'),
                'numero': r.numero_formateado(),
                'cliente': r.cliente.nombre,
                'venta_id': r.venta_asociada.id if r.venta_asociada else None,
                'venta_str': r.venta_asociada.numero_factura_formateado() if r.venta_asociada else '-',
                'estado': r.estado
            })

        return JsonResponse({
            'remitos': data,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number
        })

    except Exception as e:
        print(f"Error api_remitos_listar: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_notas_credito_listar(request):
    try:
        from django.core.paginator import Paginator
        
        page_number = request.GET.get('page', 1)
        per_page = request.GET.get('per_page', 10)
        q = request.GET.get('q', '')
        fecha = request.GET.get('fecha', '')

        notas = NotaCredito.objects.all().order_by('-fecha', '-id')

        if q:
            notas = notas.filter(
                Q(cliente__nombre__icontains=q) | 
                Q(numero__icontains=q)
            )
            
        if fecha:
            notas = notas.filter(fecha=fecha)

        paginator = Paginator(notas, per_page)
        page_obj = paginator.get_page(page_number)

        data = []
        for n in page_obj:
            data.append({
                'id': n.id,
                'fecha': n.fecha.strftime('%d/%m/%Y'),
                'numero': n.numero_formateado(),
                'cliente': n.cliente.nombre,
                'venta_id': n.venta_asociada.id if n.venta_asociada else None,
                'venta_str': n.venta_asociada.numero_factura_formateado() if n.venta_asociada else '-',
                'total': float(n.total),
                'estado': n.estado
            })

        return JsonResponse({
            'notas_credito': data,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number
        })

    except Exception as e:
        print(f"Error api_notas_credito_listar: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required

def api_remito_detalle(request, id):

    try:

        r = get_object_or_404(Remito, pk=id)

        

        items = []

        for item in r.detalles.all():

            items.append({

                'id': item.id,

                'producto': item.producto.descripcion,

                'cantidad': float(item.cantidad),

            })



        data = {

            'id': r.id,

            'numero': r.numero_formateado(),

            'fecha': r.fecha.strftime('%d/%m/%Y %H:%M'),

            'cliente': r.cliente.nombre,

            'direccion': r.cliente.domicilio, # O la direcci贸n de entrega espec铆fica si existiera en el modelo

            'venta_asociada': r.venta_asociada.numero_factura_formateado() if r.venta_asociada else '-',

            'estado': r.estado,

            'items': items

        }

        return JsonResponse(data)

    except Exception as e:

        print(f"Error api_remito_detalle: {e}")

        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@transaction.atomic
def api_nota_credito_crear(request, venta_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'M茅todo no permitido. Use POST.'}, status=405)
        
    try:
        venta = Venta.objects.get(pk=venta_id)
        
        # Verificar si ya tiene nota de cr茅dito
        if NotaCredito.objects.filter(venta_asociada=venta, estado='EMITIDA').exists():
             return JsonResponse({'ok': False, 'error': 'Esta venta ya tiene una Nota de Cr茅dito asociada.'})

        # Logic copied and adapted from view_comprobantes.crear_nota_credito
        motivo = "Anulaci贸n desde Listado de Ventas"
        
        # Crear NC
        nc = NotaCredito.objects.create(
            cliente=venta.cliente,
            venta_asociada=venta,
            tipo_comprobante=f"NC{venta.tipo_comprobante}",
            total=venta.total,
            motivo=motivo,
            estado='EMITIDA'
        )
        
        # Copiar detalles
        for det in venta.detalles.all():
            DetalleNotaCredito.objects.create(
                nota_credito=nc,
                producto=det.producto,
                cantidad=det.cantidad,
                precio_unitario=det.precio_unitario,
                subtotal=det.subtotal
            )
            
            # Devolver stock
            det.producto.stock += int(det.cantidad)
            det.producto.save()
            MovimientoStock.objects.create(
                producto=det.producto,
                tipo='IN',
                cantidad=det.cantidad,
                referencia=f"NC {nc.id} (Anula Venta {venta.id})",
                observaciones="Devoluci贸n por Nota de Cr茅dito (API)"
            )

        # Generar Asiento Contable
        try:
            from .services import AccountingService
            AccountingService.registrar_nota_credito(nc)
        except Exception as e:
            print(f"Error generando asiento de NC {nc.id}: {e}")
            
        # Actualizar estado de la venta? Generalmente queda como "Emitida" pero con NC asociada.
        # Opcional: venta.estado = 'Anulada'
        
        return JsonResponse({
            'ok': True, 
            'id': nc.id, 
            'message': f'Nota de Cr茅dito {nc.numero_formateado()} generada correctamente.'
        })

    except Venta.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Venta no encontrada.'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)



@login_required
def api_nota_credito_detalle(request, id):
    try:
        nc = NotaCredito.objects.get(pk=id)
        
        detalles = []
        for det in nc.detalles.all():
            detalles.append({
                'id': det.id,
                'producto': det.producto.descripcion,
                'cantidad': det.cantidad,
                'precio_unitario': float(det.precio_unitario),
                'subtotal': float(det.subtotal)
            })
            
        data = {
            'ok': True,
            'header': {
                'id': nc.id,
                'fecha': nc.fecha.strftime('%d/%m/%Y %H:%M'),
                'numero': nc.numero_formateado(),
                'cliente': nc.cliente.nombre,
                'venta_asociada': f"#{nc.venta_asociada.id} ({nc.venta_asociada.tipo_comprobante})" if nc.venta_asociada else "-",
                'total': float(nc.total),
                'estado': nc.estado,
                'motivo': nc.motivo,
                'tipo_comprobante': nc.tipo_comprobante 
            },
            'items': detalles
        }
        return JsonResponse(data)
    except NotaCredito.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Nota de Cr茅dito no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)



@login_required
def api_notas_debito_listar(request):
    """API para listar notas de d茅bito"""
    try:
        nds = NotaDebito.objects.select_related('cliente', 'venta_asociada').all().order_by('-fecha')
        
        data = []
        for nd in nds:
            data.append({
                'id': nd.id,
                'numero': nd.numero_formateado(),
                'fecha': nd.fecha.strftime('%d/%m/%Y %H:%M'),
                'cliente': nd.cliente.nombre,
                'venta_id': nd.venta_asociada.id if nd.venta_asociada else None,
                'venta_str': nd.venta_asociada.numero_factura_formateado() if nd.venta_asociada else '-',
                'total': float(nd.total),
                'estado': nd.estado,
                'tipo': nd.tipo_comprobante
            })
            
        return JsonResponse({'notas_debito': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def api_nota_debito_detalle(request, id):
    try:
        nd = NotaDebito.objects.get(pk=id)
        
        detalles = []
        for det in nd.detalles.all():
            detalles.append({
                'id': det.id,
                'producto': det.producto.descripcion,
                'cantidad': det.cantidad,
                'precio_unitario': float(det.precio_unitario),
                'subtotal': float(det.subtotal)
            })
            
        data = {
            'ok': True,
            'header': {
                'id': nd.id,
                'fecha': nd.fecha.strftime('%d/%m/%Y %H:%M'),
                'numero': nd.numero_formateado(),
                'cliente': nd.cliente.nombre,
                'venta_asociada': f"#{nd.venta_asociada.id} ({nd.venta_asociada.tipo_comprobante})" if nd.venta_asociada else "-",
                'total': float(nd.total),
                'estado': nd.estado,
                'motivo': nd.motivo,
                'tipo_comprobante': nd.tipo_comprobante 
            },
            'items': detalles
        }
        return JsonResponse(data)
    except NotaDebito.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Nota de D茅bito no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@csrf_exempt
@transaction.atomic
def api_nota_debito_crear(request, venta_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'M茅todo no permitido. Use POST.'}, status=405)
        
    try:
        data = json.loads(request.body)
        monto = Decimal(str(data.get('monto', 0)))
        motivo = data.get('motivo', 'Nota de D茅bito Gen茅rica')

        if monto <= 0:
             return JsonResponse({'ok': False, 'error': 'El monto debe ser mayor a 0.'})

        venta = Venta.objects.get(pk=venta_id)
        
        # Crear ND
        nd = NotaDebito.objects.create(
            cliente=venta.cliente,
            venta_asociada=venta,
            tipo_comprobante=f"ND{venta.tipo_comprobante[-1] if venta.tipo_comprobante else 'X'}", # e.g., NO -> NDX, FA -> FDA? Better: NDA/NDB/NDC
            total=monto,
            motivo=motivo,
            estado='EMITIDA'
        )
        
        # Ajustar tipo comprobante mejorado
        tipo_letra = venta.tipo_comprobante[-1] if venta.tipo_comprobante else 'X' # A, B, C
        if tipo_letra not in ['A', 'B', 'C']: tipo_letra = 'X'
        nd.tipo_comprobante = f"ND{tipo_letra}"
        nd.save()

        # Crear Detalle Gen茅rico
        DetalleNotaDebito.objects.create(
            nota_debito=nd,
            producto=Producto.objects.first(), # Fallback product necessary? Or make nullable? Assuming first product exists or generic service product.
                                             # Better: find or create a generic 'Concepto ND' product if needed, but let's use first product for now as placeholder is risky.
                                             # Actually, models says 'producto' is ForeignKey(Producto). We need a product.
            cantidad=1,
            precio_unitario=monto,
            subtotal=monto
        )
        # Note: If no product exists, this will fail. Assuming system has products.
        # Ideally we should have a 'Servicio / Concepto' product.

        # Generar Asiento Contable
        try:
            from .services import AccountingService
            AccountingService.registrar_nota_debito(nd)
        except Exception as e:
            print(f"Error generando asiento de ND {nd.id}: {e}")
            
        return JsonResponse({
            'ok': True, 
            'id': nd.id, 
            'message': f'Nota de D茅bito {nd.numero_formateado()} generada correctamente.'
        })

    except Venta.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Venta no encontrada.'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


# =======================================
#  API DASHBOARD
# =======================================
@login_required
@verificar_permiso('reportes')
def api_dashboard_stats(request):
    from django.db.models import Sum, Count, F, Q
    from datetime import date, timedelta, datetime
    from django.utils import timezone
    
    # 1. Determine Date Range (Default: Current Month)
    local_now = timezone.localtime(timezone.now())
    hoy_date = local_now.date()
    
    fecha_start_str = request.GET.get('fecha_start')
    fecha_end_str = request.GET.get('fecha_end')
    
    if fecha_start_str and fecha_end_str:
        try:
            dt_start_date = datetime.strptime(fecha_start_str, '%Y-%m-%d').date()
            dt_start = timezone.make_aware(datetime.combine(dt_start_date, datetime.min.time()))
            
            dt_end_date = datetime.strptime(fecha_end_str, '%Y-%m-%d').date()
            dt_end = timezone.make_aware(datetime.combine(dt_end_date, datetime.max.time()))
        except ValueError:
            dt_start = timezone.make_aware(datetime.combine(hoy_date.replace(day=1), datetime.min.time()))
            dt_end = timezone.make_aware(datetime.combine(hoy_date, datetime.max.time()))
    else:
        # Default: Current Month (1st to Now)
        dt_start = timezone.make_aware(datetime.combine(hoy_date.replace(day=1), datetime.min.time()))
        dt_end = timezone.make_aware(datetime.combine(hoy_date, datetime.max.time()))

    # --- 1. RENTABILIDAD (Dynamic Date Range) ---
    # Ventas en el rango
    ventas_rango = Venta.objects.filter(fecha__range=(dt_start, dt_end))
    agregado_ventas = ventas_rango.aggregate(total=Sum('total'), cantidad=Count('id'))
    total_ventas = agregado_ventas['total'] or 0
    cantidad_ventas = agregado_ventas['cantidad'] or 0
    
    ticket_promedio = float(total_ventas) / cantidad_ventas if cantidad_ventas > 0 else 0
    
    # Costo Estimado (Approximation using current Product cost)
    detalles_rango = DetalleVenta.objects.filter(venta__in=ventas_rango).select_related('producto')
    total_costo = sum((d.cantidad * d.producto.costo) for d in detalles_rango)
    
    ganancia_bruta = float(total_ventas) - float(total_costo)
    
    # Gastos Operativos (Egresos de Caja en el rango)
    total_gastos = MovimientoCaja.objects.filter(
        fecha__range=(dt_start, dt_end), 
        tipo='Egreso'
    ).aggregate(total=Sum('monto'))['total'] or 0
    
    ganancia_neta = ganancia_bruta - float(total_gastos)
    
    rentabilidad = {
        'start': dt_start.strftime('%d/%m/%Y'),
        'end': dt_end.strftime('%d/%m/%Y'),
        'ventas': float(total_ventas),
        'cantidad_ventas': cantidad_ventas,
        'ticket_promedio': float(ticket_promedio),
        'costos': float(total_costo),
        'ganancia_bruta': float(ganancia_bruta),
        'gastos': float(total_gastos),
        'ganancia_neta': float(ganancia_neta),
        'margen': round((float(ganancia_neta) / float(total_ventas) * 100), 1) if total_ventas > 0 else 0
    }
    
    # --- 1.1 Top Clientes (Dynamic Date Range) ---
    top_clientes_qs = ventas_rango.values('cliente__nombre')\
        .annotate(total_comprado=Sum('total'), cantidad_compras=Count('id'))\
        .order_by('-total_comprado')[:5]
        
    top_clientes = []
    for c in top_clientes_qs:
        top_clientes.append({
            'cliente': c['cliente__nombre'],
            'total': float(c['total_comprado']),
            'cantidad': c['cantidad_compras']
        })

    # --- 2. KPIs Principales ---
    # We keep these as TODAY/REALTIME for operational awareness, independent of the profitability filter
    today_start = timezone.make_aware(datetime.combine(hoy_date, datetime.min.time()))
    today_end = today_start + timedelta(days=1)
    
    ventas_hoy = Venta.objects.filter(fecha__gte=today_start, fecha__lt=today_end).aggregate(total=Sum('total'))['total'] or 0
    ingresos_caja = MovimientoCaja.objects.filter(fecha__gte=today_start, fecha__lt=today_end, tipo='Ingreso').aggregate(total=Sum('monto'))['total'] or 0
    egresos_caja = MovimientoCaja.objects.filter(fecha__gte=today_start, fecha__lt=today_end, tipo='Egreso').aggregate(total=Sum('monto'))['total'] or 0
    caja_hoy_neto = ingresos_caja - egresos_caja
    
    # Pendientes Query & List
    pedidos_pendientes_qs = Pedido.objects.filter(estado__in=['PENDIENTE', 'PREPARACION']).order_by('-fecha')
    pedidos_pendientes_count = pedidos_pendientes_qs.count()
    pedidos_pendientes_list = [{
        'id': p.id,
        'cliente': p.cliente.nombre,
        'total': float(p.total),
        'fecha': p.fecha.strftime('%d/%m %H:%M'),
        'estado': p.estado
    } for p in pedidos_pendientes_qs[:10]] # Limit to 10 for dashboard

    # Low Stock Query & List
    stock_bajo_qs = Producto.objects.filter(Q(stock__lte=F('stock_minimo')) | Q(stock__lte=5)).order_by('stock')
    stock_bajo_count = stock_bajo_qs.count()
    stock_bajo_list = [{
        'id': p.id,
        'nombre': p.descripcion, # Fixed: was p.nombre
        'stock': float(p.stock),
        'minimo': float(p.stock_minimo)
    } for p in stock_bajo_qs[:10]] # Limit to 10 for dashboard
    
    kpi = {
        'ventas_hoy': float(ventas_hoy),
        'caja_hoy': float(caja_hoy_neto),
        'pedidos_pendientes': pedidos_pendientes_count,
        'stock_bajo': stock_bajo_count,
    }

    # --- 3. Chart: Ventas vs Compras (Last 7 days fixed) ---
    fecha_inicio_chart = hoy_date - timedelta(days=6)
    start_chart_dt = timezone.make_aware(datetime.combine(fecha_inicio_chart, datetime.min.time()))
    
    ventas_range = Venta.objects.filter(fecha__gte=start_chart_dt).values('fecha', 'total')
    compras_range = Compra.objects.filter(fecha__gte=start_chart_dt).values('fecha', 'total')
    
    daily_ventas = {}
    daily_compras = {}
    
    for v in ventas_range:
        local_dt = v['fecha'] - timedelta(hours=3)
        day_str = local_dt.strftime('%Y-%m-%d')
        daily_ventas[day_str] = daily_ventas.get(day_str, 0) + float(v['total'])
        
    for c in compras_range:
        local_dt = c['fecha'] - timedelta(hours=3)
        day_str = local_dt.strftime('%Y-%m-%d')
        daily_compras[day_str] = daily_compras.get(day_str, 0) + float(c['total'])

    chart_labels = []
    data_ventas = []
    data_compras = []
    
    for i in range(7):
        dia_iter = fecha_inicio_chart + timedelta(days=i)
        dia_str = dia_iter.strftime('%Y-%m-%d')
        chart_labels.append(dia_iter.strftime('%d/%m'))
        data_ventas.append(daily_ventas.get(dia_str, 0))
        data_compras.append(daily_compras.get(dia_str, 0))

    chart_datasets = [
        {'label': 'Ventas', 'data': data_ventas, 'borderColor': '#0d6efd', 'backgroundColor': 'rgba(13, 110, 253, 0.1)'},
        {'label': 'Compras', 'data': data_compras, 'borderColor': '#dc3545', 'backgroundColor': 'rgba(220, 53, 69, 0.1)'}
    ]
    
    # --- 4. Actividad Reciente ---
    recientes = []
    ultimas_ventas = Venta.objects.select_related('cliente').order_by('-fecha')[:8]
    for v in ultimas_ventas:
        recientes.append({
            'timestamp': v.fecha.timestamp(),
            'tipo': 'VENTA',
            'icono': 'ShoppingCart',
            'color': 'primary',
            'fecha': (v.fecha - timedelta(hours=3)).strftime('%d/%m %H:%M'),
            'texto': f"Venta #{v.id} - ${v.total}",
            'subtexto': v.cliente.nombre if v.cliente else 'Cliente Final'
        })
    ultimas_compras = Compra.objects.select_related('proveedor').order_by('-fecha')[:8]
    for c in ultimas_compras:
        recientes.append({
            'timestamp': c.fecha.timestamp(),
            'tipo': 'COMPRA',
            'icono': 'Truck', 
            'color': 'danger',
            'fecha': (c.fecha - timedelta(hours=3)).strftime('%d/%m %H:%M'),
            'texto': f"Compra #{c.id} - ${c.total}",
            'subtexto': c.proveedor.nombre if c.proveedor else 'Proveedor General'
        })
    recientes.sort(key=lambda x: x['timestamp'], reverse=True)
    recientes = recientes[:10]

    # --- 5. Top Productos (Last 30 days fixed) ---
    fecha_inicio_top_date = hoy_date - timedelta(days=30)
    start_top_dt = timezone.make_aware(datetime.combine(fecha_inicio_top_date, datetime.min.time()))
    
    top_productos = DetalleVenta.objects.filter(venta__fecha__gte=start_top_dt)\
        .values('producto__descripcion')\
        .annotate(total_vendido=Sum('cantidad'))\
        .order_by('-total_vendido')[:5]
    
    top_list = []
    for p in top_productos:
        top_list.append({
            'producto': p['producto__descripcion'] or "Producto Desconocido",
            'total': float(p['total_vendido'] or 0)
        })

    data = {
        'kpi': kpi,
        'rentabilidad': rentabilidad,
        'chart': {
            'labels': chart_labels,
            'data': data_ventas,
            'datasets': chart_datasets
        },
        'actividad_reciente': recientes,
        'top_productos': top_list,
        'top_clientes': top_clientes,
        'stock_bajo_list': stock_bajo_list, # New
        'pedidos_pendientes_list': pedidos_pendientes_list # New
    }
    
    return JsonResponse(data)




